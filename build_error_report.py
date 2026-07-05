"""Build a per-company error report (one Excel sheet per company).
Runs the full engine on each company/scope, diffs against gt_master_corrected.csv,
and lists every error with page, item, model value, GT value, confidence, reason.

Detailed run logging (for post-run forensics, all under logs/run_<ts>/):
  llm_calls.jsonl              one record per API call: company/scope, schema, pages sent,
                               tokens in/out, cost, duration, status
  datapoints_<comp>_<scope>.json  EVERY datapoint (correct ones too): value, confidence,
                               pages, grounded, evidence — lets you trace any error to the
                               exact read without re-running
  errors.json                  the diff rows in machine-readable form
  summary.json                 totals: per-company/per-section error counts, tokens, cost
"""
import csv, json, os, re, threading, time
from datetime import datetime

import src.llm as llm

# --- run log ------------------------------------------------------------------
RUN_DIR = os.path.join("logs", f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
os.makedirs(RUN_DIR, exist_ok=True)
_LOG_LOCK = threading.Lock()
CURRENT = {"company": "", "scope": ""}          # set by the outer loop, read by the wrapper


def _log(rec: dict):
    with _LOG_LOCK:
        with open(os.path.join(RUN_DIR, "llm_calls.jsonl"), "a") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# --- cost meter + call logger ---------------------------------------------------
_cli = llm.client(); _real = _cli.responses.create
ACC = {"calls": 0, "in": 0, "out": 0}
PRICE_IN, PRICE_OUT = 0.75 / 1e6, 4.5 / 1e6     # $/token (keep in sync with the model)


def _w(*a, **k):
    t0 = time.time()
    r = _real(*a, **k)
    dt = time.time() - t0
    u = getattr(r, "usage", None)
    ti = (getattr(u, "input_tokens", 0) or 0) if u else 0
    to = (getattr(u, "output_tokens", 0) or 0) if u else 0
    with _LOG_LOCK:
        ACC["calls"] += 1; ACC["in"] += ti; ACC["out"] += to
    inp = k.get("input", "")
    txt = inp if isinstance(inp, str) else " ".join(
        c.get("text", "") for m in inp for c in m.get("content", []) if isinstance(c, dict))
    n_imgs = 0 if isinstance(inp, str) else sum(
        1 for m in inp for c in m.get("content", []) if c.get("type") == "input_image")
    sec = re.search(r"SECTION: ([^\n]*)", txt)
    _log({
        "ts": round(time.time(), 2), "company": CURRENT["company"], "scope": CURRENT["scope"],
        "schema": (k.get("text") or {}).get("format", {}).get("name", "?"),
        "section": sec.group(1).strip() if sec else "",
        "pages": [int(p) for p in re.findall(r"=== PAGE (\d+) ===", txt)],
        "images": n_imgs, "model": k.get("model", ""),
        "input_tokens": ti, "output_tokens": to,
        "cost": round(ti * PRICE_IN + to * PRICE_OUT, 5),
        "duration_s": round(dt, 2), "status": getattr(r, "status", ""),
        "instr": (k.get("instructions") or "")[:90],
    })
    return r


_cli.responses.create = _w

from src.engine.index import PageIndex
from src.engine import datapoints as dp

def num(s):
    if s is None: return None
    s = str(s).replace("−", "-").replace("–", "-")
    neg = "(" in s and ")" in s
    m = re.search(r"-?\d[\d,]*\.?\d*", s.replace(" ", ""))
    if not m: return None
    v = float(m.group(0).replace(",", "")); return -abs(v) if neg else v

def vmatch(gt, ev):
    x, y = num(gt), num(ev)
    if x is None or y is None: return False
    return abs(x) < 0.5 if abs(y) < 1e-9 else abs(x - y) / abs(y) < 0.01

# --- load GT ----------------------------------------------------------------
GT = {}   # (company, scope, key) -> {value, evidence}
for r in csv.DictReader(open("data/gt_master_corrected.csv")):
    GT[(r["company"], r["scope"], r["key"])] = {"value": r["corrected_value"], "evidence": r.get("evidence", "")}

COMPANIES = ["reliance", "hindalco", "itc", "infosys", "adani"]
SCOPES = [("standalone", "standalone"), ("consolidated", "consolidated")]
concepts = dp.load_concepts()
sec_of = {c.key: c.section for c in concepts}

def gt_is_absent(v):
    """GT cell means 'not present in the report' (no numeric truth)."""
    if num(v) is not None: return False
    t = (v or "").strip().lower()
    return t in ("", "-", "na", "n/a", "not disclosed", "not applicable", "absent", "none", "nil")

def classify(gt_val, dp_obj):
    present = dp_obj is not None and dp_obj.present
    model_v = dp_obj.value if present else None
    if gt_is_absent(gt_val):
        if present:
            return "false_positive", model_v, f"Model returned a value but GT says '{gt_val or 'absent'}' — not disclosed in the report"
        return "ok", None, ""
    # GT has a numeric truth
    if not present:
        return "miss", None, "Model returned ABSENT — datapoint not located/extracted (locate or extraction gap)"
    if vmatch(gt_val, model_v):
        return "ok", model_v, ""
    return "mismatch", model_v, "Value mismatch — model extracted a different figure than the report"

# --- run --------------------------------------------------------------------
errors = {c: [] for c in COMPANIES}
all_rows = {c: [] for c in COMPANIES}
per_company_cost = {}
for comp in COMPANIES:
    calls0, in0, out0 = ACC["calls"], ACC["in"], ACC["out"]
    idx = PageIndex(os.path.expanduser(f"~/Downloads/{comp}.pdf"))
    got = {}   # (scope, key) -> Datapoint
    for eng_scope, gt_scope in SCOPES:
        CURRENT["company"], CURRENT["scope"] = comp, eng_scope
        t0 = time.time()
        res = dp.extract_datapoints(idx, eng_scope, concepts)
        for k, d in res.items():
            got[(gt_scope, k)] = d
        # full per-datapoint dump — correct values included (forensics without re-running)
        with open(os.path.join(RUN_DIR, f"datapoints_{comp}_{eng_scope}.json"), "w") as f:
            json.dump({k: {"section": d.section, "present": d.present, "value": d.value,
                           "confidence": d.confidence, "grounded": d.grounded,
                           "pages": d.pages, "evidence": d.evidence}
                       for k, d in res.items()}, f, indent=1, ensure_ascii=False)
        print(f"  {comp}/{eng_scope}: {len(res)} datapoints in {time.time()-t0:.0f}s "
              f"(running cost ~${ACC['in']*PRICE_IN + ACC['out']*PRICE_OUT:.2f})", flush=True)
    # diff every GT row for this company — ALL rows recorded (correct ones included)
    for (cc, scope, key), info in GT.items():
        if cc != comp: continue
        lookups = [scope] if scope != "both" else ["standalone", "consolidated"]
        # for 'both', report once; pick the worse outcome across scopes
        chosen = None
        for sc in lookups:
            d = got.get((sc, key))
            etype, mv, reason = classify(info["value"], d)
            cand = (etype, mv, reason, d, sc)
            order = {"mismatch": 0, "miss": 1, "false_positive": 2, "ok": 3}
            if chosen is None or order[etype] < order[chosen[0]]:
                chosen = cand
        etype, mv, reason, d, sc = chosen
        row = {
            "section": sec_of.get(key, ""), "scope": sc, "item": key, "status": etype,
            "model_value": (d.value if d is not None and d.present else None),
            "gt_value": info["value"],
            "confidence": d.confidence if d is not None else "absent",
            "grounded": d.grounded if d is not None else "",
            "pages": ",".join(str(p) for p in (d.pages or [])) if d is not None else "",
            "model_evidence": (d.evidence or "")[:300] if d is not None else "",
            "gt_evidence": (info["evidence"] or "")[:300],
            "reason": reason,
        }
        all_rows[comp].append(row)
        if etype != "ok":
            errors[comp].append(row)
    dc, din, dout = ACC["calls"] - calls0, ACC["in"] - in0, ACC["out"] - out0
    per_company_cost[comp] = {"calls": dc, "input_tokens": din, "output_tokens": dout,
                              "cost": round(din * PRICE_IN + dout * PRICE_OUT, 3)}
    print(f"{comp:9}: {len(errors[comp])} errors | {dc} calls, {din:,} in / {dout:,} out tok, "
          f"~${per_company_cost[comp]['cost']:.2f}", flush=True)

# --- machine-readable outputs -------------------------------------------------
with open(os.path.join(RUN_DIR, "errors.json"), "w") as f:
    json.dump(errors, f, indent=1, ensure_ascii=False)
sec_counts = {}
for comp, es in errors.items():
    for e in es:
        sec_counts[e["section"]] = sec_counts.get(e["section"], 0) + 1
summary = {
    "total_errors": sum(len(v) for v in errors.values()),
    "per_company": {c: len(errors[c]) for c in COMPANIES},
    "per_section": dict(sorted(sec_counts.items(), key=lambda x: -x[1])),
    "api": {"calls": ACC["calls"], "input_tokens": ACC["in"], "output_tokens": ACC["out"],
            "cost": round(ACC["in"] * PRICE_IN + ACC["out"] * PRICE_OUT, 2)},
    "per_company_cost": per_company_cost,
}
with open(os.path.join(RUN_DIR, "summary.json"), "w") as f:
    json.dump(summary, f, indent=1)

# --- write Excel (FULL report: every GT datapoint, correct rows included) ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
hfill = PatternFill("solid", fgColor="1F4E78"); hfont = Font(color="FFFFFF", bold=True)
sfill = {"ok": "E2EFDA", "mismatch": "FCE4D6", "miss": "FFF2CC", "false_positive": "F8CBAD"}
_ORDER = {"mismatch": 0, "miss": 1, "false_positive": 2, "ok": 3}

# Summary: per-company accuracy + API cost
ws = wb.active; ws.title = "Summary"
ws.append(["Company", "GT rows", "Correct", "Errors", "Mismatch", "Miss", "False positive",
           "Accuracy %", "API calls", "Input tokens", "Output tokens", "Cost $"])
for c in range(1, 13):
    ws.cell(1, c).fill = hfill; ws.cell(1, c).font = hfont
tot = dict(n=0, ok=0, mm=0, mi=0, fp=0)
for comp in COMPANIES:
    rows = all_rows[comp]
    n = len(rows); ok = sum(r["status"] == "ok" for r in rows)
    mm = sum(r["status"] == "mismatch" for r in rows)
    mi = sum(r["status"] == "miss" for r in rows)
    fp = sum(r["status"] == "false_positive" for r in rows)
    tot["n"] += n; tot["ok"] += ok; tot["mm"] += mm; tot["mi"] += mi; tot["fp"] += fp
    api = per_company_cost[comp]
    ws.append([comp, n, ok, n - ok, mm, mi, fp, round(ok / n * 100, 1),
               api["calls"], api["input_tokens"], api["output_tokens"], api["cost"]])
ws.append(["TOTAL", tot["n"], tot["ok"], tot["n"] - tot["ok"], tot["mm"], tot["mi"], tot["fp"],
           round(tot["ok"] / tot["n"] * 100, 1), ACC["calls"], ACC["in"], ACC["out"],
           summary["api"]["cost"]])
for c in range(1, 13):
    ws.cell(ws.max_row, c).font = Font(bold=True)
for i, w in enumerate([12, 8, 8, 7, 9, 6, 13, 11, 9, 13, 14, 8], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.append([]); ws.append(["Errors by section"]); ws.cell(ws.max_row, 1).font = Font(bold=True)
for s, n in summary["per_section"].items():
    ws.append([s, n])

# per-company sheets: EVERY datapoint, errors first, status colour-coded
HEAD = ["section", "scope", "item", "status", "model_value", "gt_value", "confidence",
        "grounded", "pages", "model_evidence", "gt_evidence"]
for comp in COMPANIES:
    w2 = wb.create_sheet(comp.capitalize())
    w2.append(HEAD)
    for c in range(1, len(HEAD) + 1):
        w2.cell(1, c).fill = hfill; w2.cell(1, c).font = hfont
    rows = sorted(all_rows[comp], key=lambda r: (_ORDER[r["status"]], r["section"], r["item"]))
    for r in rows:
        w2.append([r[h] for h in HEAD])
        w2.cell(w2.max_row, HEAD.index("status") + 1).fill = \
            PatternFill("solid", fgColor=sfill[r["status"]])
    for i, w in enumerate([16, 12, 34, 13, 14, 14, 11, 8, 14, 55, 45], 1):
        w2.column_dimensions[get_column_letter(i)].width = w
    w2.freeze_panes = "A2"
    for row in w2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

out = "error_report.xlsx"
wb.save(out)
print(f"\nWrote {out}  ({summary['total_errors']} errors / {tot['n']} rows, "
      f"accuracy {tot['ok']/tot['n']*100:.1f}%)")
print(f"Per-section: {summary['per_section']}")
print(f"API: calls={ACC['calls']} in={ACC['in']:,} out={ACC['out']:,} "
      f"cost~${summary['api']['cost']:.2f}")
print(f"Logs: {RUN_DIR}/")
