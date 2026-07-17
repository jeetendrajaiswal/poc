"""Raw table extraction — every table in the report, verbatim, to Excel.

Unlike the datapoint layer (which maps a fixed taxonomy by meaning), this module
transcribes EVERY table exactly as printed: exact labels, exact values, no
renaming, no interpretation. One worksheet per table plus an Index sheet
(page, scope, title, dimensions) for navigation.

Fully deterministic and local — word coordinates from PyMuPDF, whitespace-gutter
column detection (the two-column reflow idea generalised to N columns), zero
model calls, zero run-to-run variance, nothing leaves the host.

    from src.engine.tables import extract_tables, write_workbook
    tables = extract_tables("annual_report.pdf")
    write_workbook(tables, "out.xlsx")

    .venv/bin/python -m src.engine.tables report.pdf out.xlsx
"""
from __future__ import annotations

import re
import sys
from dataclasses import dataclass, field

import pymupdf


# --------------------------------------------------------------------------- words

def _rotated_boxes(page) -> tuple[list[tuple], dict]:
    """(bboxes of non-horizontal text lines, dir->line-count) for the page."""
    boxes, dirs = [], {}
    for blk in page.get_text("dict")["blocks"]:
        if blk.get("type") != 0:
            continue
        for ln in blk["lines"]:
            key = (round(ln["dir"][0]), round(ln["dir"][1]))
            dirs[key] = dirs.get(key, 0) + 1
            if abs(ln["dir"][0] - 1) > 0.01:
                boxes.append(ln["bbox"])
    return boxes, dirs


_ROT_BOILER = re.compile(r"annual report|in (millions?|crores?|lakhs?)|unless otherwise",
                         re.IGNORECASE)


def _page_fallback_title(page, first_row_text: str) -> str:
    """Title for tables with no adjacent heading: some layouts (e.g. Wipro)
    print the statement title as rotated text in the margin. Not applicable on
    rotation-dominant pages, where rotated lines ARE the table content."""
    rot, dirs = _rotated_boxes(page)
    if sum(c for d, c in dirs.items() if d != (1, 0)) > max(10, dirs.get((1, 0), 0)):
        return first_row_text[:80]
    cands = []
    for blk in page.get_text("dict")["blocks"]:
        if blk.get("type") != 0:
            continue
        for ln in blk["lines"]:
            if abs(ln["dir"][0] - 1) <= 0.01:
                continue
            txt = "".join(sp["text"] for sp in ln["spans"]).strip()
            if txt and not _ROT_BOILER.search(txt):
                cands.append(txt)
    if cands:
        return max(cands, key=len)[:80]
    return first_row_text[:80]


def _fallback_from_rows(rows: list[dict]) -> str:
    """First usable heading-like line on the page, top-down."""
    for r in rows[:12]:
        txt = _row_text(r)
        if len(_row_segments(r)) < 2 and _good_title(txt):
            return txt[:80]
    return ""


def _page_words(page) -> list[tuple]:
    """Words as (x0, y0, x1, y1, text) in READING coordinates.

    Horizontal pages: horizontal words, rotated sidebar text dropped.
    Rotation-DOMINANT pages (landscape PPE/SOCIE grids printed sideways): the
    rotated words are re-mapped into their reading frame and the (sideways)
    leftovers are dropped, so the same pipeline handles both."""
    rot, dirs = _rotated_boxes(page)
    n_hor = dirs.get((1, 0), 0)
    n_rot = sum(c for d, c in dirs.items() if d != (1, 0))

    def in_rot(w):
        cx, cy = (w[0] + w[2]) / 2, (w[1] + w[3]) / 2
        return any(bx0 <= cx <= bx1 and by0 <= cy <= by1 for bx0, by0, bx1, by1 in rot)

    words = page.get_text("words")
    if n_rot > max(10, n_hor):
        up = dirs.get((0, -1), 0)       # bottom-to-top text
        down = dirs.get((0, 1), 0)      # top-to-bottom text
        H, W = page.rect.height, page.rect.width
        out = []
        for w in words:
            if not in_rot(w):
                continue
            x0, y0, x1, y1 = w[:4]
            if up >= down:              # rotate page 90° clockwise to read
                out.append((H - y1, x0, H - y0, x1, w[4]))
            else:                       # rotate 90° counter-clockwise
                out.append((y0, W - x1, y1, W - x0, w[4]))
        return out
    return [(w[0], w[1], w[2], w[3], w[4]) for w in words if not in_rot(w)]


# --------------------------------------------------------------------------- rows

def _cluster_rows(words: list[tuple]) -> list[dict]:
    """Cluster words into physical text lines by y-center (fixed tolerance)."""
    if not words:
        return []
    hs = sorted(w[3] - w[1] for w in words)
    ytol = max(2.0, 0.45 * hs[len(hs) // 2])
    rows: list[dict] = []
    for w in sorted(words, key=lambda w: ((w[1] + w[3]) / 2, w[0])):
        cy = (w[1] + w[3]) / 2
        if rows and abs(rows[-1]["y"] - cy) <= ytol:
            r = rows[-1]
            n = len(r["words"])
            r["words"].append(w)
            r["y"] = (r["y"] * n + cy) / (n + 1)
            r["y0"] = min(r["y0"], w[1])
            r["y1"] = max(r["y1"], w[3])
        else:
            rows.append({"y": cy, "y0": w[1], "y1": w[3], "words": [w]})
    for r in rows:
        r["words"].sort(key=lambda w: w[0])
    return rows


def _row_segments(row: dict, gap: float = 8.0) -> list[tuple]:
    """Group a row's words into segments split on x-gaps > `gap`."""
    segs = []
    cur = [row["words"][0]]
    for w in row["words"][1:]:
        if w[0] - cur[-1][2] > gap:
            segs.append(cur)
            cur = [w]
        else:
            cur.append(w)
    segs.append(cur)
    return [(s[0][0], s[-1][2], " ".join(w[4] for w in s)) for s in segs]


def _row_text(row: dict) -> str:
    return " ".join(w[4] for w in row["words"])


# --------------------------------------------------------------------------- numeric tests

_DIGITS = set("0123456789")
_VALUE_TOKENS = {"-", "–", "—", "nil", "*", "^"}


def _is_numericish(cell: str) -> bool:
    """True when the cell carries a numeric token (value, note ref, count, year)."""
    if not cell:
        return False
    for tok in cell.split():
        alnum = [c for c in tok if c.isalnum()]
        if not alnum:
            continue
        d = sum(c in _DIGITS for c in alnum)
        if d >= 1 and d >= 0.5 * len(alnum):
            return True
    return False


def _is_value_side(rows: list[dict], run: int = 4) -> bool:
    """True when the region looks like the detached value-column half of ONE
    table: it contains a contiguous run of rows made (almost) only of numeric
    tokens. Prose columns and side-by-side sub-tables carry label words in
    every row, so they never produce such a run."""
    streak = 0
    for r in rows:
        toks = [w[4] for w in r["words"]]
        hits = sum(1 for t in toks if _is_numericish(t) or t.lower() in _VALUE_TOKENS)
        if toks and hits >= 0.8 * len(toks):
            streak += 1
            if streak >= run:
                return True
        else:
            streak = 0
    return False


# --------------------------------------------------------------------------- geometry

def _find_gutters(rows: list[dict], x0: float, x1: float,
                  min_width: float = 5.0, max_cross_frac: float = 0.12) -> list[tuple]:
    """Vertical x-intervals inside [x0,x1] crossed by (almost) no row's words.
    Returns [(a, b, fully_empty)]."""
    n = len(rows)
    step = 2.0
    xs = []
    x = x0
    while x < x1:
        xs.append(x)
        x += step
    cross = [0] * len(xs)
    for r in rows:
        for w in r["words"]:
            for i, x in enumerate(xs):
                if w[0] < x < w[2]:
                    cross[i] += 1
    maxc = max(1, int(max_cross_frac * n))
    gutters = []
    i = 0
    while i < len(xs):
        if cross[i] <= maxc:
            j = i
            while j < len(xs) and cross[j] <= maxc:
                j += 1
            a, b = xs[i], xs[min(j, len(xs) - 1)]
            if b - a >= min_width:
                gutters.append((a, b, all(cross[k] == 0 for k in range(i, j))))
            i = j
        else:
            i += 1
    return gutters


def _col_signature(seg: list[dict]) -> list[float] | None:
    """Column-gutter midpoints of a segment's multi-segment rows (None if too
    little structure to tell)."""
    mrows = [r for r in seg if len(_row_segments(r)) >= 2]
    if len(mrows) < 2:
        return None
    x0 = min(w[0] for r in mrows for w in r["words"])
    x1 = max(w[2] for r in mrows for w in r["words"])
    return [(a + b) / 2 for a, b, _ in _find_gutters(mrows, x0, x1)]


def _is_total_row(r: dict) -> bool:
    """A ruled-off totals line: few tokens, mostly numeric ('Total 5,568,164,645
    4,968,542,229'). Bank schedules print these detached from the table body."""
    toks = [w[4] for w in r["words"]]
    if len(toks) < 2 or len(toks) > 8:
        return False
    numeric = sum(1 for t in toks if _is_numericish(t) or t.lower() in _VALUE_TOKENS)
    return numeric >= 2 and numeric >= 0.5 * len(toks)


def _merge_compatible_segments(segs: list[list[dict]],
                               tol: float = 18.0) -> list[list[dict]]:
    """Re-join vertically-split segments that share the same column skeleton —
    a tall note table (e.g. a PPE schedule) has spacer gaps between its
    sub-sections but identical columns throughout, whereas a signature block
    or a following unrelated table has different columns. A detached one-row
    totals line re-joins the table above it."""
    out = [segs[0]]
    for nxt in segs[1:]:
        prev = out[-1]
        # detached table-row(s) — ruled-off totals like 'Total (a+b)' /
        # 'Total (I+II)' — belong to the table above when they fit inside its
        # column extent; a following heading etc. stays its own segment
        if len(prev) >= 2 and _col_signature(prev):
            lead = []
            for r in nxt[:2]:
                if (len(_row_segments(r)) >= 2
                        and (_is_total_row(r) or _is_numericish(_row_text(r)))):
                    lead.append(r)
                else:
                    break
            # if what follows the peeled rows is still table-like, nxt is a
            # genuine table block (e.g. a continuation page) — don't peel it
            rest0 = nxt[len(lead)] if len(lead) < len(nxt) else None
            if rest0 is not None and len(_row_segments(rest0)) >= 2:
                lead = []
            if lead:
                px0 = min(w[0] for r in prev for w in r["words"]) - 20
                px1 = max(w[2] for r in prev for w in r["words"]) + 20
                nx0 = min(w[0] for r in lead for w in r["words"])
                nx1 = max(w[2] for r in lead for w in r["words"])
                if px0 <= nx0 and nx1 <= px1:
                    out[-1] = prev + lead
                    nxt = nxt[len(lead):]
                    if not nxt:
                        continue
                    prev = out[-1]
        sa, sb = _col_signature(prev), _col_signature(nxt)
        if (sa and sb and len(sa) == len(sb) and len(sa) >= 1
                and all(abs(a - b) <= tol for a, b in zip(sa, sb))):
            out[-1] = prev + nxt
        else:
            out.append(nxt)
    return out


def _split_at_vertical_gaps(rows: list[dict], factor: float = 2.5) -> list[list[dict]]:
    """Split a run of rows into segments wherever a big vertical gap occurs."""
    if len(rows) < 4:
        return [rows]
    gaps = [rows[i + 1]["y0"] - rows[i]["y1"] for i in range(len(rows) - 1)]
    med = sorted(gaps)[len(gaps) // 2]
    thresh = max(14.0, factor * max(med, 1.0))
    segs, cur = [], [rows[0]]
    for i, g in enumerate(gaps):
        if g > thresh:
            segs.append(cur)
            cur = []
        cur.append(rows[i + 1])
    segs.append(cur)
    return segs


def _side(rows: list[dict], xa: float, xb: float) -> list[dict]:
    out = []
    for r in rows:
        ws = [w for w in r["words"] if xa <= (w[0] + w[2]) / 2 < xb]
        if ws:
            out.append({"y": r["y"], "y0": r["y0"], "y1": r["y1"], "words": ws})
    return out


# --------------------------------------------------------------------------- blocks -> grids

def _blocks_from_rows(rows: list[dict], max_single: int = 3,
                      min_multi: int = 2) -> list[tuple[int, list[dict]]]:
    """Maximal runs of multi-segment rows (tolerating a few 1-segment rows
    inside, e.g. section headers). Returns (start_index, block_rows) pairs."""
    tagged = [(r, len(_row_segments(r)) >= 2) for r in rows]
    blocks: list[tuple[int, list[dict]]] = []
    cur: list[dict] = []
    start = 0
    single_run = 0

    def flush():
        if cur and sum(1 for r in cur if len(_row_segments(r)) >= 2) >= min_multi:
            b = list(cur)
            while b and len(_row_segments(b[-1])) < 2:
                b.pop()
            blocks.append((start, b))

    for i, (r, multi) in enumerate(tagged):
        if multi:
            if not cur:
                start = i
            cur.append(r)
            single_run = 0
        elif cur and single_run < max_single:
            cur.append(r)
            single_run += 1
        else:
            flush()
            cur, single_run = [], 0
    flush()
    return blocks


def _merge_sparse_columns(grid: list[list[str]]) -> list[list[str]]:
    """Drop all-empty columns; fold nearly-empty ones into the right neighbour."""
    if not grid:
        return grid
    n = len(grid[0])
    thresh = max(2, int(0.05 * len(grid)))
    while True:
        counts = [sum(1 for row in grid if row[i]) for i in range(n)]
        victim = next((i for i in range(n) if counts[i] == 0), None)
        if victim is None:
            # sparse = few cells AND clearly under-populated vs the row count
            victim = next((i for i in range(n)
                           if counts[i] <= thresh and counts[i] <= 0.5 * len(grid)), None)
            if victim is None or n <= 2:
                break
        tgt = victim + 1 if victim + 1 < n else victim - 1
        for row in grid:
            row[tgt] = (row[min(victim, tgt)] + " " + row[max(victim, tgt)]).strip()
            del row[victim]
        n -= 1
    return grid


def _looks_tabular(grid: list[list[str]]) -> bool:
    """Keep only grids that look like data tables, not multi-column prose."""
    if len(grid) < 2 or len(grid[0]) < 2:
        return False
    cells = [c for row in grid for c in row if c]
    if not cells:
        return False
    numeric = sum(1 for c in cells if _is_numericish(c))
    if numeric < max(3, 0.12 * len(cells)):
        return False
    lens = sorted(len(c) for c in cells)
    if lens[len(lens) // 2] > 45:            # median cell is a prose sentence
        return False
    return True


def _assign_row(r: dict, edges: list[float], gap: float = 8.0) -> list[str]:
    """Place a row's segments into columns by maximal overlap (keeps wrapped
    headers whole); a segment straddling columns falls back to per-word centers."""
    ncol = len(edges) - 1
    cells = ["" for _ in range(ncol)]
    for sx0, sx1, txt in _row_segments(r, gap):
        best, besto = None, 0.0
        for i in range(ncol):
            o = min(sx1, edges[i + 1]) - max(sx0, edges[i])
            if o > besto:
                best, besto = i, o
        if best is not None and besto >= 0.7 * (sx1 - sx0):
            cells[best] = (cells[best] + " " + txt).strip()
        else:
            for w in r["words"]:
                if sx0 <= w[0] and w[2] <= sx1 + 0.1:
                    cx = (w[0] + w[2]) / 2
                    for i in range(ncol):
                        if edges[i] <= cx < edges[i + 1]:
                            cells[i] = (cells[i] + " " + w[4]).strip()
                            break
    return cells


def _build_grid(rows: list[dict]) -> list[list[str]] | None:
    if len(rows) < 2:
        return None
    x0 = min(w[0] for r in rows for w in r["words"])
    x1 = max(w[2] for r in rows for w in r["words"])
    mrows = [r for r in rows if len(_row_segments(r)) >= 2] or rows
    gutters = _find_gutters(mrows, x0, x1)
    edges = [x0 - 1] + [(a + b) / 2 for a, b, _ in gutters] + [x1 + 1]
    if len(edges) < 3:
        return None
    grid = _merge_sparse_columns([_assign_row(r, edges) for r in rows])
    return grid if _looks_tabular(grid) else None


# --------------------------------------------------------------------------- regions

def _process_region(rows: list[dict], depth: int = 0,
                    segmented: bool = False) -> list[tuple[str, list[list[str]]]]:
    """Vertical-gap segments first, then side-by-side column splits, then table
    blocks. Returns (title, grid) pairs in reading order."""
    rows = [r for r in rows if r["words"]]
    if len(rows) < 2:
        return []
    if not segmented:
        segs = _split_at_vertical_gaps(rows)
        if len(segs) > 1:
            segs = _merge_compatible_segments(segs)
        if len(segs) > 1:
            out = []
            for s in segs:
                out.extend(_process_region(s, depth, segmented=True))
            return out
        rows = segs[0]
    x0 = min(w[0] for r in rows for w in r["words"])
    x1 = max(w[2] for r in rows for w in r["words"])
    if depth < 2:
        cands = [g for g in _find_gutters(rows, x0, x1, min_width=10.0, max_cross_frac=0.12)
                 if g[0] > x0 + 40 and g[1] < x1 - 40]
        cands.sort(key=lambda g: g[1] - g[0], reverse=True)
        for a, b, _zero in cands:
            mid = (a + b) / 2
            left, right = _side(rows, x0 - 1, mid), _side(rows, mid, x1 + 1)
            if len(left) < 3 or len(right) < 3:
                continue
            # a pure value-column side means this is ONE table's label/number
            # whitespace gap, not a page-layout split — never break it apart
            if _is_value_side(left) or _is_value_side(right):
                continue
            # try both readings, keep whichever preserves more content — a
            # short table's label/value gap can masquerade as a layout gutter
            split_res = (_process_region(left, depth + 1)
                         + _process_region(right, depth + 1))
            unsplit_res = _blocks_to_tables(rows)
            if _content_tokens(split_res) >= 0.98 * _content_tokens(unsplit_res):
                return split_res
            return unsplit_res
    return _blocks_to_tables(rows)


def _content_tokens(tables) -> int:
    return sum(len(c.split()) for t in tables for row in t[-1] for c in row if c)


def _blocks_to_tables(rows: list[dict]) -> list[tuple]:
    """(title, y_top, x0, x1, grid) per detected block."""
    out = []
    for start, block in _blocks_from_rows(rows):
        grid = _build_grid(block)
        if grid is None:
            continue
        # title: nearest usable single-segment heading lines above the block
        title_parts = []
        for r in reversed(rows[max(0, start - 4):start]):
            if len(_row_segments(r)) >= 2 or block[0]["y0"] - r["y1"] > 60:
                break
            txt = _row_text(r)
            if _good_title(txt):
                title_parts.append(txt)
                if len(title_parts) == 2:
                    break
        bx0 = min(w[0] for r in block for w in r["words"])
        bx1 = max(w[2] for r in block for w in r["words"])
        out.append((" ".join(reversed(title_parts)).strip(), block[0]["y0"], bx0, bx1, grid))
    return out


def _good_title(txt: str) -> bool:
    """Reject printed page numbers, unit boilerplate, and running banners."""
    t = txt.strip()
    letters = [c for c in t if c.isalpha()]
    if len(t) < 4 or not letters:
        return False
    if _ROT_BOILER.search(t):
        return False
    # running banner: long, digit-led, shouting ("230 STATUTORY REPORTS AND ...")
    if (t[0].isdigit() and len(t.split()) >= 5
            and sum(c.isupper() for c in letters) > 0.7 * len(letters)):
        return False
    return True


_STATEMENT_TITLE = re.compile(
    r"balance sheet|statements? of (profit|cash flows?|changes in equity)|profit and loss"
    r"|revenue account|receipts and payments", re.IGNORECASE)
_STATEMENT_HEAD = re.compile(
    r"^\s*(consolidated\s+|standalone\s+)?(balance sheet|statements? of profit"
    r"|statements? of cash flows?|statements? of changes in equity|profit and loss"
    r"|revenue account|receipts and payments|cash flow statement)\b", re.IGNORECASE)
_NOTE_HEAD = re.compile(r"^\s*(?:\d{1,3}|[A-Z])\s*[.)]\s+(.{3,70})$")
_SCHED_HEAD = re.compile(r"^\s*SCHEDULE\s+\d+", re.IGNORECASE)
_UNITS_LINE = re.compile(
    r"(?:\bin\s+|\brupees\s+|\brs\.?\s+in\s+|^`?\s*)(millions?|crores?|lakhs?|thousands?"
    r"|['‘’`]?000s?)\b", re.IGNORECASE)


_ROLE_WORDS = re.compile(
    r"secretary|director|officer|chairman|chairperson|partner|accountant|chief|"
    r"executive|president|membership", re.IGNORECASE)


def _is_note_heading(txt: str) -> bool:
    """Numbered note / schedule headings: '5. RIGHT-OF-USE ASSETS',
    'SCHEDULE 2 - RESERVES AND SURPLUS', 'B. Other Equity'."""
    t = txt.strip()
    if len(t) > 80 or _UNITS_LINE.search(t):
        return False
    if _SCHED_HEAD.match(t):
        return True
    m = _NOTE_HEAD.match(t)
    if not m:
        return False
    if re.match(r"^\s*[IVX]\s*[.)]", t):      # roman numerals are row markers, not headings
        return False
    body = m.group(1)
    letters = sum(c.isalpha() for c in body)
    digits = sum(c.isdigit() for c in body)
    if not (letters >= 3 and letters > 2 * digits):
        return False
    # letter-led headings ('B. Other Equity') share their shape with signatory
    # initials ('M. Sanaulla Khan Company Secretary') — keep them short and
    # free of people-role words
    if not t[0].isdigit():
        toks = body.split()
        if (len(toks) > 4 or _ROLE_WORDS.search(body)
                or not all(all(c.isalpha() or c in "&-'()," for c in w) for w in toks)):
            return False
    return True


def _is_banner_row(txt: str, top: bool) -> bool:
    """Rows to strip before table detection: printed page numbers, running
    top-of-page banners, and unit-boilerplate lines."""
    t = txt.strip()
    if t.replace(" ", "").isdigit() and len(t) <= 5:
        return True                                     # bare printed page number
    if (len(t) <= 90 and _UNITS_LINE.search(t)
            and t.lstrip("([").rstrip(")]")[:1] in "₹`RIKCH(r"
            and not any(c.isdigit() for c in re.sub(r"['‘’`]?000s?", "", t))):
        return True                                     # '(₹ in millions, ...)' / 'Rupees crores'
    if top and not _good_title(t):
        letters = [c for c in t if c.isalpha()]
        if (letters and len(t.split()) >= 4
                and sum(c.isupper() for c in letters) > 0.7 * len(letters)):
            return True                                 # shouting nav banner
    return False


def _prefilter_rows(rows: list[dict]) -> tuple[list[dict], str]:
    """Drop banner/page-number/units rows (they garble grids and titles);
    return the surviving rows plus the page's units line if one was seen."""
    units = ""
    out = []
    for i, r in enumerate(rows):
        txt = _row_text(r)
        if _is_banner_row(txt, top=i < 3):
            if not units and _UNITS_LINE.search(txt):
                units = txt.strip()
            continue
        out.append(r)
    return out, units


# --------------------------------------------------------------------------- public API

@dataclass
class RawTable:
    page: int                       # 1-based page number
    n: int                          # 1-based table number on the page
    title: str                      # nearest heading above the table ('' if none)
    scope: str                      # standalone | consolidated | unknown
    section: str                    # note/schedule heading in force ('5. RIGHT-OF-USE ASSETS')
    page_head: str                  # page-level heading ('Notes to the Standalone ...')
    units: str                      # units line if printed ('₹ in millions, ...')
    grid: list[list[str]] = field(repr=False)


def extract_tables(path: str, progress=None, financial_only: bool = True) -> list[RawTable]:
    """Every table in the PDF, verbatim, in reading order. Deterministic, local.

    financial_only=True (default) keeps only pages inside the standalone /
    consolidated financial-statements block (statements, notes, schedules) —
    identified by the engine's deterministic scope tagging — and skips the
    front-of-book sections (corporate overview, board's report, BRSR,
    governance). Pass financial_only=False for every table in the document."""
    scopes = _scope_tags(path)
    doc = pymupdf.open(path)
    tables: list[RawTable] = []
    cur_section = ""
    cur_units = ""
    prev_scope = None
    for i in range(len(doc)):
        scope = scopes[i] if i < len(scopes) else "unknown"
        if financial_only and scopes and scope not in ("standalone", "consolidated"):
            continue
        if scope != prev_scope:          # new scope block: its notes restart
            cur_section = ""
            prev_scope = scope
        page = doc[i]
        rows, units = _prefilter_rows(_cluster_rows(_page_words(page)))
        if not rows and len(page.get_text().strip()) < 40 and page.get_images(full=True):
            # scanned/graphic page inside the financial section: never skip silently
            tables.append(RawTable(
                page=i + 1, n=1, title="(image-only page — NOT extracted)",
                scope=scope, section="(image-only page — needs vision/OCR)",
                page_head="", units=cur_units,
                grid=[["This page has no text layer (scanned image); "
                       "its tables were NOT extracted."]]))
            continue
        if units:
            cur_units = units
        page_head = _fallback_from_rows(rows)
        fallback = _page_fallback_title(page, page_head)
        # a primary-statement page starts a new section; a stale note heading
        # from earlier pages (e.g. the auditor's report) must not leak in
        if _STATEMENT_TITLE.search(fallback) or _STATEMENT_TITLE.search(page_head):
            cur_section = fallback if _STATEMENT_TITLE.search(fallback) else page_head
        headings = [(r["y0"], _row_text(r).strip()) for r in rows
                    if _is_note_heading(_row_text(r))]
        # statement titles are section headings too — detected per row SEGMENT,
        # with x-extent, so a two-up page ('Balance Sheet ... | Statement of
        # Profit and Loss ...') titles each side separately
        stmt_heads = []
        for r in rows:
            for sx0, sx1, txt in _row_segments(r):
                if len(txt) <= 90 and _STATEMENT_HEAD.match(txt):
                    stmt_heads.append((r["y0"], sx0, sx1, txt.strip()))
        found = _process_region(rows)
        # walk headings and tables in reading order so each table gets the
        # note/schedule heading in force (carried across continuation pages)
        events = ([(y, 0, txt, None) for y, txt in headings]
                  + [(y, 1, t, tup) for t, y, _x0, _x1, tup in
                     [(t, y, x0, x1, (x0, x1, g)) for t, y, x0, x1, g in found]])
        events.sort(key=lambda e: (e[0], e[1]))
        n = 0
        for y, kind, payload, tup in events:
            if kind == 0:
                cur_section = payload
                continue
            n += 1
            tx0, tx1, grid = tup
            section = cur_section
            # nearest statement heading ABOVE the table that overlaps it in x
            best = None
            for hy, sx0, sx1, txt in stmt_heads:
                if hy <= y + 2 and min(sx1, tx1) - max(sx0, tx0) > 0:
                    if best is None or hy >= best[0]:
                        best = (hy, txt)
            if best:
                section = best[1]
                cur_section = best[1]
            title = payload or fallback
            tables.append(RawTable(page=i + 1, n=n, title=title, scope=scope,
                                   section=section, page_head=page_head,
                                   units=cur_units, grid=grid))
        if progress and (i + 1) % 50 == 0:
            progress(i + 1, len(doc))
    doc.close()
    return tables


def _scope_tags(path: str) -> list[str]:
    """standalone/consolidated tag per page via the engine's page_scopes
    (deterministic, no model calls). Soft dependency: on any failure every
    page is 'unknown' and extraction proceeds."""
    try:
        from .index import PageIndex
        from .datapoints import page_scopes
        return page_scopes(PageIndex(path))
    except Exception:
        return []


# --------------------------------------------------------------------------- excel

_SHEET_BAD = re.compile(r"[\[\]:*?/\\']")


def _sheet_name(t: RawTable, used: set[str]) -> str:
    base = f"p{t.page}_{t.n}"
    label = _SHEET_BAD.sub(" ", t.section or t.title).strip()
    name = (f"{base} {label}" if label else base)[:31].rstrip()
    k, cand = 2, name
    while cand.lower() in used:
        suffix = f" ({k})"
        cand = name[:31 - len(suffix)].rstrip() + suffix
        k += 1
    used.add(cand.lower())
    return cand


def _group_tables(tables: list[RawTable]) -> list[list[RawTable]]:
    """Club consecutive tables of the same scope + section into one sheet
    (a Balance Sheet spanning two pages, a note and its sub-tables). Tables
    with no section stay on their own sheet."""
    groups: list[list[RawTable]] = []
    for t in tables:
        key = (t.scope, t.section.strip().lower()) if t.section.strip() else None
        prev = groups[-1][-1] if groups else None
        prev_key = ((prev.scope, prev.section.strip().lower())
                    if prev and prev.section.strip() else None)
        if key and prev_key == key and t.page - prev.page <= 1:
            groups[-1].append(t)
        else:
            groups.append([t])
    return groups


def write_workbook(tables: list[RawTable], out_path: str) -> None:
    """One sheet per SECTION (consecutive same-section tables clubbed,
    multi-page statements joined) + an Index sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    index = wb.active
    index.title = "Index"
    index.append(["Sheet", "Pages", "Scope", "Section", "Tables", "Rows", "Cols"])
    for c in index[1]:
        c.font = Font(bold=True)

    meta_font = Font(color="666666", size=9)
    part_font = Font(color="1F4E78", italic=True, size=9)
    used: set[str] = set()
    for group in _group_tables(tables):
        t0 = group[0]
        name = _sheet_name(t0, used)
        ws = wb.create_sheet(name)
        pages = (str(t0.page) if group[-1].page == t0.page
                 else f"{t0.page}-{group[-1].page}")
        ws.append([t0.section or t0.title])
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.append([f"Scope: {t0.scope}   |   PDF page(s): {pages}   |   "
                   f"{t0.page_head}" + (f"   |   Units: {t0.units}" if t0.units else "")])
        ws.cell(row=2, column=1).font = meta_font
        widths: dict[int, int] = {}
        total_rows = 0
        for t in group:
            ws.append([])
            if len(group) > 1:
                ws.append([f"· p{t.page}: {t.title}" if t.title else f"· p{t.page}"])
                ws.cell(row=ws.max_row, column=1).font = part_font
            elif "⚠" in t.title:
                # a verification flag must never be invisible
                ws.append([t.title])
                ws.cell(row=ws.max_row, column=1).font = Font(color="B00000", bold=True)
            for row in t.grid:
                ws.append(row)
            total_rows += len(t.grid)
            for row in t.grid:
                for j, cell in enumerate(row, 1):
                    widths[j] = max(widths.get(j, 8), min(len(cell), 60))
        for j, wdt in widths.items():
            ws.column_dimensions[get_column_letter(j)].width = wdt + 2
        ws.freeze_panes = "A3"
        r = index.max_row + 1
        section_label = t0.section or t0.title
        if any("⚠" in t.title for t in group):
            section_label = f"⚠ {section_label}"
        index.append([name, pages, t0.scope, section_label,
                      len(group), total_rows, max(len(t.grid[0]) for t in group)])
        index.cell(row=r, column=1).hyperlink = f"#'{name}'!A3"
        index.cell(row=r, column=1).style = "Hyperlink"
    for j, wdt in zip(range(1, 8), (34, 9, 14, 52, 7, 6, 6)):
        index.column_dimensions[get_column_letter(j)].width = wdt
    index.freeze_panes = "A2"
    wb.save(out_path)


def extract_to_excel(pdf_path: str, out_path: str, progress=None,
                     financial_only: bool = True) -> int:
    """Convenience: extract + write. Returns the number of tables."""
    tables = extract_tables(pdf_path, progress=progress, financial_only=financial_only)
    write_workbook(tables, out_path)
    return len(tables)


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--all"]
    fin_only = "--all" not in sys.argv
    if not args:
        sys.exit("usage: python -m src.engine.tables report.pdf [out.xlsx] [--all]\n"
                 "  default: financial-statements section only; --all extracts every table")
    pdf = args[0]
    out = args[1] if len(args) > 1 else pdf.rsplit(".", 1)[0] + "_tables.xlsx"
    n = extract_to_excel(pdf, out, progress=lambda p, t: print(f"  {p}/{t} pages"),
                         financial_only=fin_only)
    print(f"{n} tables -> {out}")
