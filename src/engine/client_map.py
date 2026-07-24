"""Client-format mapping layer — map raw extracted statements to a client's
field taxonomy BY MEANING, not by label (the proven taxonomy/definitions.yaml
approach from the datapoint engine, applied to client fields).

Pieces:
  * config/<sector>/taxonomy.yaml — per client field: fid, name,
    concept (economic DEFINITION with inclusion/exclusion/total-vs-component
    guards), illustrative aliases. Generated from evidence (client template +
    human mapping files + filing corpus), human-reviewable, versioned.
  * The client template's 'Field Calculation / Logic' formulas are used for
    VERIFICATION: after mapping, each aggregate's mapped components must sum
    to its mapped printed value. A wrong mapping breaks arithmetic and flags.
  * Every value carries (value, unit, currency); periods are normalised to
    3M/6M/9M/FY with end dates.
"""
from __future__ import annotations

import re
from dataclasses import dataclass


# --------------------------------------------------------------------------- template (structure + verification formulas)

@dataclass
class ClientField:
    fid: str
    name: str
    order: int
    row: int
    formula: list[tuple[int, int]] | None    # [(sign, template_row), ...]  additive only
    group: str = ""                          # parent aggregate name(s), derived from formulas
    ratio: tuple[int, int, float] | None = None   # (num_row, den_row, scale) for =Cn/Cd*scale


class Taxonomy(dict):
    """Statement fields plus declarative sector-level mapping policy."""

    def __init__(self, *args, location_vocabulary=None,
                 statement_sections=None, section_locations=None,
                 identities=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.location_vocabulary = set(location_vocabulary or [])
        self.statement_sections = statement_sections or {}
        self.section_locations = section_locations or {}
        self.identities = identities or {}


def _parse_formula(expr) -> list[tuple[int, int]] | None:
    if not expr or not str(expr).startswith("="):
        return None
    s = str(expr)
    # Ratio/percentage formulas (=Cn/Cd*scale) are NOT additive — a signed C-ref
    # sum would silently turn a division into an addition. They are handled
    # separately via _parse_ratio, so decline them here.
    if "*" in s or "/" in s:
        return None
    out = [(-1 if sg == "-" else 1, int(r)) for sg, r in re.findall(r"([+\-]?)C(\d+)", s)]
    return out or None


def _parse_ratio(expr) -> tuple[int, int, float] | None:
    """Ratio/percentage formulas of the form =C<num>/C<den> with an optional
    *<scale> (all margins in the template are =Cn/Cd*100; a plain =Cn/Cd is a
    bare ratio with scale 1). Returns (num_row, den_row, scale) or None."""
    if not expr:
        return None
    m = re.fullmatch(r"=\s*C(\d+)\s*/\s*C(\d+)\s*(?:\*\s*(\d+(?:\.\d+)?))?",
                     str(expr).strip())
    if not m:
        return None
    return (int(m.group(1)), int(m.group(2)), float(m.group(3)) if m.group(3) else 1.0)


def _derive_template_groups(
        template: dict[tuple[str, str], list[ClientField]]) -> None:
    """Populate formula-derived parent chains used for location resolution."""
    for fields in template.values():
        byrow = {f.row: f for f in fields}
        parent: dict[int, int] = {}
        for f in fields:
            if f.formula:
                for _sign, rr in f.formula:
                    if rr in byrow and rr not in parent:
                        parent[rr] = f.row
        for f in fields:
            chain = []
            cur = f.row
            for _ in range(3):
                p = parent.get(cur)
                if p is None:
                    break
                chain.append(byrow[p].name)
                cur = p
            f.group = " < ".join(chain)


def template_from_taxonomy(
        taxonomy: dict[str, list[dict]]) -> dict[tuple[str, str], list[ClientField]]:
    """Compile presentation order and FID-based formulas from taxonomy v2."""
    out: dict[tuple[str, str], list[ClientField]] = {}
    calculations: dict[tuple[str, str, str], dict] = {}
    for statement, items in taxonomy.items():
        for item in items:
            for scope, scoped in item.get("scopes", {}).items():
                position = int(scoped["position"])
                key = (statement, scope)
                out.setdefault(key, []).append(ClientField(
                    fid=item["fid"],
                    name=item["name"],
                    order=position,
                    row=position,
                    formula=None,
                ))
                calculations[(statement, scope, item["fid"])] = (
                    scoped.get("calculation") or {"type": "reported"})
    for key, fields in out.items():
        fields.sort(key=lambda field: (field.order, field.fid))
        byfid = {field.fid: field for field in fields}
        if len(byfid) != len(fields):
            raise ValueError(f"duplicate field id in taxonomy scope {key}")
        positions = [field.order for field in fields]
        if len(positions) != len(set(positions)):
            raise ValueError(f"duplicate field position in taxonomy scope {key}")
        for field in fields:
            calculation = calculations[(key[0], key[1], field.fid)]
            kind = calculation.get("type", "reported")
            if kind == "sum":
                field.formula = [
                    (int(term.get("sign", 1)), byfid[str(term["fid"])].row)
                    for term in calculation.get("terms", [])
                ]
                if not field.formula:
                    raise ValueError(f"empty sum formula for {key}/{field.fid}")
            elif kind == "ratio":
                numerator = byfid[str(calculation["numerator"])].row
                denominator = byfid[str(calculation["denominator"])].row
                field.ratio = (
                    numerator, denominator, float(calculation.get("scale", 1)))
            elif kind != "reported":
                raise ValueError(
                    f"unknown calculation type {kind!r} for {key}/{field.fid}")
    _derive_template_groups(out)
    return out


def derived_section_pairs(fields: list[ClientField],
                          defs: list[dict] | None = None) -> dict[str, dict[str, str]]:
    """For labels that belong to MORE THAN ONE field (duplicate display names,
    or a name/alias shared across fids — 'same concept, different id'), derive
    which fid is the CURRENT vs NON-CURRENT (or OPERATING/INVESTING/FINANCING)
    variant from the taxonomy's declared locations. Fully generic — no
    hand-written FID or caption table."""
    declared = {
        str(item["fid"]): item.get("locations", [])[0]
        for item in (defs or [])
        if len(item.get("locations", [])) == 1
    }
    side = {f.fid: declared.get(f.fid, "") for f in fields}
    bylabel: dict[str, dict[str, ClientField]] = {}
    byfid = {f.fid: f for f in fields}
    for f in fields:
        bylabel.setdefault(norm_label(f.name), {})[f.fid] = f
    for d in defs or []:
        f = byfid.get(d["fid"])
        if not f:
            continue
        for a in [d.get("name", "")] + list(d.get("aliases", [])):
            nl = norm_label(a)
            if nl:
                bylabel.setdefault(nl, {})[f.fid] = f
    pairs: dict[str, dict[str, str]] = {}
    for name, fs in bylabel.items():
        if len(fs) < 2:
            continue
        sides: dict[str, set[str]] = {}
        for f in fs.values():
            s = side[f.fid]
            if s:
                sides.setdefault(s, set()).add(f.fid)
        # pin only when unambiguous: every side has exactly ONE candidate fid;
        # a side with 2+ claimants means the label alone can't decide — leave
        # those lines to the definitions layer
        if len(sides) >= 2 and all(len(v) == 1 for v in sides.values()):
            pairs[name] = {s: next(iter(v)) for s, v in sides.items()}
    return pairs


def statement_of(section: str, title: str) -> str | None:
    """Which client statement a raw table belongs to.

    Most-specific match FIRST: many filings title every statement
    '... Financial Results — Balance Sheet / Cash Flows', so 'results' alone
    must never claim a table that names a more specific statement.
    """
    s = f"{section} {title}".lower()
    if "cash flow" in s:
        return "cashflow"
    if "segment" in s:            # before balance: 'Segment Assets and Liabilities'
        return "segment"          # names the more specific statement
    if "assets and liabilities" in s or "balance sheet" in s:
        return "balance"
    if "results" in s or "profit and loss" in s:
        return "income"
    return None


def _infer_scope(*texts) -> str:
    """Recover scope when extraction left it 'unknown' — scanned/OCR'd headings
    like 'Standelone' (Standalone) or 'Consolldated' fail exact keyword matching.
    Returns a scope ONLY on a confident fuzzy match to exactly one of the two;
    genuinely ambiguous tables stay 'unknown' and are skipped as before."""
    from difflib import SequenceMatcher
    words = re.findall(r"[a-z]+", " ".join(str(t) for t in texts).lower())
    near = lambda w, t: SequenceMatcher(None, w, t).ratio() >= 0.8
    con = any(near(w, "consolidated") for w in words)
    std = any(near(w, "standalone") for w in words)
    if con and not std:
        return "consolidated"
    if std and not con:
        return "standalone"
    return "unknown"


def map_quarter(tables, template, taxonomy, model: str | None = None,
                default_unit: str = "", period_hint: tuple | None = None,
                log=None):
    """Map one filing's raw tables to the client taxonomy.

    tables: iterable of (page, n, title, scope, section, grid).
    Merges ALL tables of the same (statement, scope) — filings routinely print
    one statement as several blocks (P&L + OCI/EPS, segment revenue + results).
    IFRS versions and unknown scopes are excluded. Returns {key: MappedStatement}.

    period_hint: (quarter 1-4, fiscal-year end year) from the job metadata —
    used only when a statement's own header cannot resolve its periods.
    Extraction ⚠ flags in table titles are carried onto the MappedStatement;
    after mapping, each statement is re-verified (verify_mapped) so a mapping
    that breaks the statement's printed identities is flagged, never silent.
    """
    grids: dict[tuple[str, str], list] = {}
    raw_flags: dict[tuple[str, str], list] = {}
    for _page, _n, title, scope, section, grid in tables:
        s = f"{section} {title}".lower()
        stmt = statement_of(section, title)
        if scope not in ("standalone", "consolidated"):    # recover OCR/vision
            scope = _infer_scope(title, section)            # 'unknown' scope
        if (stmt is None or "ifrs" in s or len(grid) < 3
                or scope not in ("standalone", "consolidated")):
            continue
        key = (stmt, scope)
        if key in template:
            grids.setdefault(key, []).append(grid)
            if "⚠" in title:
                raw_flags.setdefault(key, []).append(
                    "extraction: " + title.split("⚠", 1)[1].strip())
    out = {}
    for key, gl in grids.items():
        width = max(len(r) for g in gl for r in g)
        merged = [r + [""] * (width - len(r)) for g in gl for r in g]
        ms = map_statement(
            merged, key[0], taxonomy, template[key], model=model, scope=key[1])
        _nu = lambda u: {"mn": "million", "'000": "thousand"}.get(u, u).rstrip("s")
        if default_unit and not ms.unit:
            ms.unit = default_unit        # units line is printed above the
        elif default_unit and ms.unit and _nu(ms.unit) != _nu(default_unit):
            # the grid names one denomination, the filing-level detector
            # another — one of them is wrong; a wrong unit scales EVERY value
            ms.flag(f"denomination conflict: statement says '{ms.unit}', the "
                    f"filing-level detection says '{default_unit}' — verify")
        if key[0] == "cashflow":          # table, not inside the grid
            reconcile_cashflow_opening(ms, template[key])
        for fl in raw_flags.get(key, []):
            ms.flag(fl)
        out[key] = ms
    resolve_bare_periods(out, period_hint)
    for key, ms in out.items():
        verify_mapped(ms, key[0], key[1], taxonomy)
        # completeness: distinct value columns must map to DISTINCT periods.
        # If two columns share a period, the wide workbook silently merges them
        # and drops data — flag it rather than lose a column unseen.
        keyed = [(p.span, p.end) for p in ms.periods if p.end]
        if len(keyed) != len(set(keyed)):
            ms.flag("period columns collapsed — two columns resolved to the same "
                    "period, so a column is dropped in the workbook; verify the "
                    "period header against the filing")
        if log:
            log(f"  map[{key[0]}/{key[1]}]: {len(ms.facts)} fields, "
                f"{len(ms.unmapped)} unmapped lines, formula checks "
                f"{ms.n_ok}/{ms.n_checks}, periods "
                f"{[(p.span, p.end or p.raw[:18]) for p in ms.periods]}")
            for fl in (getattr(ms, 'flags', None) or []):
                log(f"    ⚠ {fl[:160]}")
    return out


_SPAN_RANK = {"3M": 1, "6M": 2, "9M": 3, "FY": 4}


def resolve_bare_periods(mapped: dict, period_hint: tuple | None = None) -> None:
    """Resolve periods a statement's own header could not (bare-year columns
    like '2026 | 2025' under 'Statement of Cash Flows' — the span/end context
    lives outside the table). Evidence, in order:

      1. the OTHER statements of the SAME filing: a bare year names the
         filing's reporting date for that year, and an interim cash flow is
         CUMULATIVE, so the longest span the filing reports for that year is
         the right one;
      2. the job metadata (quarter, fiscal year) when a statement has exactly
         current + prior-year columns and no year tokens at all.

    Anything still unresolved keeps its raw column label and the statement is
    FLAGGED — an empty sheet is never a legal outcome (values exist; only the
    period label is uncertain)."""
    evidence: dict[str, set] = {}
    for ms in mapped.values():
        for p in ms.periods:
            if p.end:
                evidence.setdefault(p.end[:4], set()).add((p.end, p.span))

    _QEND = {1: "06-30", 2: "09-30", 3: "12-31", 4: "03-31"}
    _QSPAN = {1: "3M", 2: "6M", 3: "9M", 4: "FY"}

    for (stmt, scope), ms in mapped.items():
        # a statement with facts must expose its value columns as periods even
        # when the header gave nothing — placeholder first, resolve below
        if not ms.periods and ms.facts:
            cols = sorted({c for f in ms.facts.values() for c in f})
            ms.periods = [Period("?", "", "", c, raw=f"(unresolved period — column {c})")
                          for c in cols]
        unresolved = [p for p in ms.periods if not p.end]
        if not unresolved:
            continue
        for p in unresolved:
            m = re.fullmatch(r"\(?\s*((?:19|20)\d{2})\s*\)?[.,]?\s*(\(?(un)?audited\)?)?",
                             " ".join(str(p.raw).split()), re.IGNORECASE)
            yr = m.group(1) if m else None
            if yr and yr in evidence:
                cands = sorted(evidence[yr],
                               key=lambda es: _SPAN_RANK.get(es[1], 0))
                end, span = (cands[-1] if stmt in ("cashflow", "income")
                             else cands[0])
                p.end = end
                if p.span == "?":
                    p.span = span if stmt != "balance" else "?"
                ms.flag(f"period '{p.raw}' resolved to {p.span} ended {p.end} "
                        "from the filing's other statements — review")
        still = [p for p in ms.periods if not p.end]
        if still and period_hint and len(ms.periods) == 2 and \
                not any(re.search(r"(19|20)\d{2}", str(p.raw)) for p in ms.periods):
            q, fy = period_hint
            if fy < 100:                          # 'FY26' → 2026
                fy += 2000
            cal = fy if q == 4 else fy - 1
            ends = [f"{cal}-{_QEND[q]}", f"{cal - 1}-{_QEND[q]}"]
            span = _QSPAN[q] if stmt in ("cashflow", "income") else "?"
            for p, end in zip(ms.periods, ends):
                if not p.end:
                    p.end, p.span = end, (span if p.span == "?" else p.span)
            ms.flag(f"periods inferred from the filing's quarter/FY metadata "
                    f"(Q{q} FY{fy}) — review")
        if any(not p.end for p in ms.periods):
            ms.flag("periods could NOT be resolved — column labels are shown "
                    "as printed; verify against the filing")


def compare_reads(rows, tables2, keys) -> tuple[list[dict], list[str], list[dict]]:
    """Cell-level comparison of two INDEPENDENT reads of the same filing, for
    the statements in `keys` ({(scope, stmt)}) — the ones with no text-layer
    authority (scans), where optical misreads are random and a disagreement
    between reads is the only per-cell error signal.

    A disagreeing cell whose original value is arithmetically PROVEN (swapping
    in the other read breaks an identity that currently ties) is dropped.
    Returns (suspects, notes, broad): cell-level items to mark in the
    deliverable, human-readable notes, and statements where the reads disagree
    wholesale (layout mismatch — flag the whole statement)."""
    from src.engine import identities
    bykey: dict = {}
    for t in tables2:
        bykey.setdefault((t.scope, statement_of(t.section, t.title)), t.grid)
    suspects, notes, broad = [], [], []
    for pg, n, tt, sc, sec, g in rows:
        st = statement_of(sec, tt)
        g2 = bykey.get((sc, st))
        if (sc, st) not in keys or not g2:
            continue
        if max(len(r) for r in g) != max(len(r) for r in g2):
            continue                        # layouts differ: cells can't be paired
        occ2, rows2 = {}, {}
        for r in g2:
            l2, v2 = _label_and_vals(r)
            k2 = " ".join(l2.lower().split())
            if k2 and v2:
                rows2[(k2, occ2.get(k2, 0))] = v2
                occ2[k2] = occ2.get(k2, 0) + 1
        hdr = next((r for r in g if sum(1 for c in r if str(c).strip()) > 2), g[0])
        occ1, cells = {}, []
        for ri, r in enumerate(g):
            l1, v1 = _label_and_vals(r)
            k1 = " ".join(l1.lower().split())
            if not (k1 and v1):
                continue
            v2m = rows2.get((k1, occ1.get(k1, 0)))
            occ1[k1] = occ1.get(k1, 0) + 1
            if not v2m:
                continue
            for j, a in v1.items():
                b = v2m.get(j)
                if b is not None and abs(a - b) > 0.02:
                    cells.append({"stmt": st, "scope": sc, "label": l1,
                                  "col": str(hdr[j] if j < len(hdr) else ""),
                                  "v1": a, "v2": b, "page": pg,
                                  "_ri": ri, "_j": j})
        unproven = []
        for c in cells:
            gc = [list(r) for r in g]
            gc[c["_ri"]][c["_j"]] = f"{c['v2']:g}"
            if identities.failing(sec, tt, gc):
                continue                    # original value arithmetically proven
            unproven.append(c)
        if len(unproven) > 8:               # wholesale disagreement = layout
            notes.append(f"[{sc[:4]}] {tt[:40]}: two reads disagree broadly "
                         f"({len(unproven)} cells) — verify statement")
            broad.append({"stmt": st, "scope": sc, "title": tt, "page": pg})
        else:
            suspects.extend(unproven)
    notes += [f"[{s['scope'][:4]}] {s['stmt']}: reads disagree on "
              f"'{s['label'][:40]}' [{s['col'][:30]}]: {s['v1']} vs {s['v2']}"
              for s in suspects]
    return suspects, notes, broad


def adopt_verified_second_reads(rows, tables2, keys, page_forms, page_lines,
                                untrusted_pages) -> tuple[list, list[dict]]:
    """Replace a failing first read with an independently verified second read.

    This is intentionally stricter than :func:`compare_reads`: replacement is
    allowed only when the first statement already fails a printed identity and
    a same-scope/same-statement/accounting-basis candidate:

      * has strictly fewer identity failures;
      * is fully grounded in a trusted PDF text layer;
      * has the same number and identity of reporting periods;
      * has the same numeric width and substantial row-label overlap.

    Healthy statements are never replaced.  A clean-looking but ungrounded
    second model response is never adopted.  These gates make the recovery
    generic while preventing a standalone/consolidated or short/wide table
    swap.
    """
    from collections import Counter
    from src.engine import identities, source_align

    def _clean_title(title: str) -> str:
        return re.sub(
            r"\s+⚠\s+(?:verification failed|arithmetic does not tie)"
            r"\s+—\s+review\s*$", "", title or "", flags=re.IGNORECASE,
        ).strip()

    def _numeric_width(grid) -> int:
        return max((len(vals) for row in grid
                    for _label, vals in [_label_and_vals(row)]), default=0)

    def _labels(grid) -> Counter:
        out: Counter = Counter()
        for row in grid:
            label, vals = _label_and_vals(row)
            if not vals:
                continue
            norm = re.sub(r"[^a-z0-9]+", " ", label.lower()).strip()
            # Enumeration prefixes vary harmlessly between independent reads.
            norm = re.sub(r"^(?:[ivxlcdm]+|[a-z]|\d+)\s+", "", norm)
            if norm:
                out[norm] += 1
        return out

    def _period_signature(grid):
        periods = parse_periods(grid)
        if not periods or any(not p.end or p.span == "?" for p in periods):
            return None
        return tuple((p.span, p.end) for p in periods)

    def _basis(*texts) -> str:
        text = " ".join(str(t) for t in texts).upper()
        if "IFRS" in text:
            return "ifrs"
        if re.search(r"\bIND[\s-]*AS\b", text):
            return "ind_as"
        return ""

    candidates: dict[tuple[str, str], list] = {}
    for table in tables2:
        stmt = statement_of(table.section, table.title)
        if stmt:
            candidates.setdefault((table.scope, stmt), []).append(table)

    out, notes = [], []
    for row in rows:
        page, n, title, scope, section, grid = row
        stmt = statement_of(section, title)
        key = (scope, stmt)
        old_bad = identities.failing(section, title, grid)
        if key not in keys or not stmt or not old_bad:
            out.append(row)
            continue

        old_width = _numeric_width(grid)
        old_labels = _labels(grid)
        old_periods = _period_signature(grid)
        old_basis = _basis(section, title)
        old_label_count = sum(old_labels.values())
        best = None
        for table in candidates.get(key, []):
            if "truncated" in f"{table.title} {table.page_head}".lower():
                continue
            candidate_basis = _basis(table.section, table.title)
            if (old_basis or candidate_basis) and old_basis != candidate_basis:
                continue
            if not source_align.has_text_authority(
                    table.grid, page_forms, untrusted_pages):
                continue
            grid2, report = source_align.reconcile_with_source(
                table.grid, table.section, table.title, page_forms, page_lines,
                scan_pages=untrusted_pages)
            if (not report or report["abstained"] or report["conservative"]
                    or report["structure_mismatch"] or report["unverified_cols"]):
                continue
            new_checks = identities.run_checks(table.section, table.title, grid2)
            new_bad = [name for name, ok in new_checks if not ok]
            if not new_checks or len(new_bad) >= len(old_bad):
                continue
            if _numeric_width(grid2) != old_width:
                continue
            new_periods = _period_signature(grid2)
            # A replacement must prove the column identity, not merely width.
            if old_periods is None or new_periods is None or new_periods != old_periods:
                continue
            new_labels = _labels(grid2)
            overlap = sum((old_labels & new_labels).values())
            required = max(5, int(0.80 * old_label_count + 0.999))
            if (overlap < required
                    or sum(new_labels.values()) < 0.90 * old_label_count):
                continue
            score = (len(old_bad) - len(new_bad), overlap, len(new_labels))
            if best is None or score > best[0]:
                best = (score, table, grid2, new_bad, overlap)

        if best is None:
            out.append(row)
            continue

        _score, table, grid2, new_bad, overlap = best
        clean = _clean_title(table.title or title)
        if new_bad:
            clean += "  ⚠ verification failed — review"
        out.append((page, n, clean, scope, table.section or section, grid2))
        notes.append({
            "scope": scope, "stmt": stmt, "title": clean, "page": page,
            "old_failures": old_bad, "new_failures": new_bad,
            "label_overlap": overlap,
        })
    return out, notes


def verify_mapped(ms: "MappedStatement", stmt: str, scope: str,
                  taxonomy: Taxonomy) -> None:
    """Post-mapping CORRECTNESS check on the MAPPED numbers.

    Correctness is decided by the statement's declarative cross-identities in
    the active sector taxonomy. A break means a value was genuinely misread or
    mis-mapped, so it becomes a statement flag.

    It deliberately does NOT flag 'mapped components don't sum to a template
    total' — that is a mapping-GRANULARITY signal (a printed line had no
    dedicated client field), not a value error: the reported total is itself a
    printed figure and is correct. Those remain in ms.verification as Audit
    notes only (see map_statement), never as review flags. This is what stops
    a 100%-correct statement (TCS, HM) from being orange-tabbed."""
    def fact(fid):
        return ms.facts.get(fid, {}) if fid else {}

    def check(name, terms, result_fid):
        """Signed terms == result reported fact, per column.

        Only runs when every term is a REPORTED fact (printed figure) — never
        forces a verdict off a partially-mapped aggregate. Optional terms
        participate only when the filing prints them.
        """
        result = fact(result_fid)
        if not result:
            return
        required = [
            (coefficient, fid)
            for coefficient, fid, presence in terms
            if presence == "required"
        ]
        optional = [
            (coefficient, fid)
            for coefficient, fid, presence in terms
            if presence == "optional"
        ]
        for c in sorted(result):
            if not all(c in fact(fid) for _coefficient, fid in required):
                continue
            present_optional = [
                (coefficient, fid) for coefficient, fid in optional
                if c in fact(fid)
            ]
            total = sum(
                coefficient * fact(fid)[c]
                for coefficient, fid in required + present_optional
            )
            if abs(total - result[c]) <= max(
                    1.0, 0.005 * abs(result[c])):
                continue
            per = next((p for p in ms.periods if p.col == c), None)
            lbl = _period_label(per) if per else f"column {c}"
            ms.flag(f"{name} does not hold ({lbl}) — a value is misread "
                    "or mis-mapped; verify against the filing")
            ms.verification.append(f"⚠ {name} fails at {lbl}")
            return

    for identity in taxonomy.identities.get(stmt, []):
        if scope not in identity["scopes"]:
            continue
        terms = [
            (
                float(term["coefficient"]),
                str(term["fid"]),
                str(term["presence"]),
            )
            for term in identity["terms"]
        ]
        check(str(identity["name"]), terms, str(identity["result_fid"]))


# --------------------------------------------------------------------------- taxonomy (definitions)

def norm_label(s: str) -> str:
    s = str(s or "").lower()
    # Preserve economic tokens before punctuation is removed. Without this,
    # "Gross NPA (%)" collapsed to "Gross NPA" and became indistinguishable
    # from the amount field.
    s = s.replace("%", " percent ")
    s = re.sub(r"\(refer[^)]*\)", " ", s)
    s = re.sub(r"^\(?[a-z][.)]\s*", " ", s)             # 'a)' 'b.' '(c)' enumerators
    # roman enumerators ('VI Total tax expense') — VALID numerals only, so real
    # words like 'LIC', 'CC', 'IT' are never stripped
    s = re.sub(r"^\(?(x{0,3}(ix|iv|v?i{1,3}|v|x))[\s.,):]+", " ", s)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", s).split())


def _label_value_type(label: str) -> str:
    """Infer only explicit type cues; return empty when the caption is silent."""
    text = str(label or "")
    normalized = norm_label(text)
    if "%" in text or re.search(r"\bpercent(?:age)?\b", normalized):
        return "percentage"
    if re.search(r"\b(?:eps|earnings per (?:equity )?share)\b", normalized):
        return "per_share"
    if (re.search(r"\b(?:number|no|total)\s+of\s+shares\b", normalized)
            or re.search(r"\bshares?\s+(?:issued|outstanding|held)\b", normalized)
            or re.search(r"\b(?:issued|outstanding)\s+shares?\b", normalized)):
        return "count"
    return ""


def _compile_definition(definition, statement: str, fid: str) -> str:
    if not isinstance(definition, dict):
        raise ValueError(
            f"definition must be an object for {statement}/{fid}")
    required = {
        "meaning", "includes", "excludes", "mapping_notes",
        "distinguish_from",
    }
    absent = sorted(required - set(definition))
    if absent:
        raise ValueError(
            f"definition contract for {statement}/{fid} is missing "
            f"{', '.join(absent)}")
    meaning = str(definition.get("meaning", "")).strip()
    if not meaning:
        raise ValueError(
            f"definition meaning cannot be empty for {statement}/{fid}")
    for key in ("includes", "excludes", "mapping_notes", "distinguish_from"):
        values = definition.get(key)
        if not isinstance(values, list):
            raise ValueError(
                f"definition.{key} must be a list for {statement}/{fid}")
        if any(not str(value).strip() for value in values):
            raise ValueError(
                f"definition.{key} contains an empty value for "
                f"{statement}/{fid}")
    details = [meaning]
    for key, heading in (
            ("includes", "Includes"), ("excludes", "Excludes"),
            ("mapping_notes", "Mapping notes")):
        if definition[key]:
            details.append(f"{heading}: " + "; ".join(definition[key]))
    if definition["distinguish_from"]:
        details.append(
            "Competing field FIDs: "
            + ", ".join(str(value)
                        for value in definition["distinguish_from"]))
    details.append(
        "A different unit, time nature, statement/location, or granularity "
        "is not this field. If more than one field fits, reject the mapping.")
    return " ".join(details)


def _validate_taxonomy_semantics(
        item: dict, statement: str, fid: str) -> None:
    value_type = str(item.get("value_type", "")).strip()
    allowed_value_types = {
        "amount", "percentage", "count", "per_share", "text",
    }
    if value_type not in allowed_value_types:
        raise ValueError(
            f"invalid value_type {value_type!r} for {statement}/{fid}")
    scopes = item.get("scopes")
    if not isinstance(scopes, dict) or not scopes:
        raise ValueError(
            f"scopes must be a non-empty object for {statement}/{fid}")
    invalid_scopes = set(scopes) - {"standalone", "consolidated"}
    if invalid_scopes:
        raise ValueError(
            f"invalid scopes for {statement}/{fid}: "
            f"{sorted(invalid_scopes)}")

    unit = str(item.get("unit", "")).strip()
    time_nature = str(item.get("time_nature", "")).strip()
    units_for_type = {
        "amount": {"statement_currency"},
        "percentage": {"percent"},
        "count": {"count", "shares"},
        "per_share": {"currency_per_share"},
        "text": {"text"},
    }
    allowed_time_natures = {
        "point_in_time", "duration", "period_average", "context_dependent",
    }
    if unit not in units_for_type[value_type]:
        raise ValueError(
            f"invalid unit {unit!r} for {value_type} field "
            f"{statement}/{fid}")
    if time_nature not in allowed_time_natures:
        raise ValueError(
            f"invalid time_nature {time_nature!r} for {statement}/{fid}")
    if item.get("evidence") not in {
            "client_mapping", "template_inferred"}:
        raise ValueError(
            f"invalid evidence status for {statement}/{fid}: "
            f"{item.get('evidence')!r}")


def _read_taxonomy_yaml(path: str) -> dict:
    """Load YAML while rejecting duplicate keys instead of silently keeping
    the last value. A duplicate contract key can otherwise change mapping
    behavior while still appearing valid during review."""
    import yaml

    class UniqueKeyLoader(yaml.SafeLoader):
        pass

    def construct_mapping(loader, node, deep=False):
        loader.flatten_mapping(node)
        mapping = {}
        for key_node, value_node in node.value:
            key = loader.construct_object(key_node, deep=deep)
            if key in mapping:
                mark = key_node.start_mark
                raise ValueError(
                    f"duplicate taxonomy key {key!r} at "
                    f"{path}:{mark.line + 1}")
            mapping[key] = loader.construct_object(value_node, deep=deep)
        return mapping

    UniqueKeyLoader.add_constructor(
        yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
        construct_mapping,
    )
    with open(path, encoding="utf-8") as handle:
        return yaml.load(handle, Loader=UniqueKeyLoader) or {}


def load_taxonomy(path: str) -> Taxonomy:
    """{statement: [ {fid, name, concept, aliases, ...} ]}"""
    doc = _read_taxonomy_yaml(path)
    if not isinstance(doc, dict) or not isinstance(doc.get("items"), list):
        raise ValueError(f"taxonomy must contain an items list: {path}")
    allowed_locations = {
        str(value).strip().upper()
        for value in (doc.get("location_vocabulary") or [])
        if str(value).strip()
    }
    if not allowed_locations or "STATEMENT-WIDE" not in allowed_locations:
        raise ValueError(
            "location_vocabulary must explicitly include STATEMENT-WIDE")
    statement_sections = doc.get("statement_sections")
    if not isinstance(statement_sections, dict):
        raise ValueError("statement_sections must be an object")
    compiled_sections = {}
    section_locations = {}
    for statement, entries in statement_sections.items():
        if not isinstance(entries, list):
            raise ValueError(
                f"statement_sections.{statement} must be a list")
        compiled = []
        for entry in entries:
            if not isinstance(entry, dict):
                raise ValueError(
                    f"statement section entry must be an object: {entry!r}")
            pattern = str(entry.get("pattern", "")).strip()
            section = str(entry.get("section", "")).strip().upper()
            location = str(entry.get("location", "")).strip().upper()
            if not pattern or not section or not location:
                raise ValueError(
                    f"statement section requires pattern, section, and "
                    f"location: "
                    f"{entry!r}")
            try:
                re.compile(pattern)
            except re.error as exc:
                raise ValueError(
                    f"invalid section pattern {pattern!r}") from exc
            if location not in allowed_locations:
                raise ValueError(
                    f"section {section!r} uses an unknown location "
                    f"{location!r} outside "
                    "location_vocabulary")
            compiled.append((pattern, section, location))
            statement_key = str(statement).lower()
            prior = section_locations.setdefault(
                statement_key, {}).get(section)
            if prior and prior != location:
                raise ValueError(
                    f"section {statement_key}/{section} has conflicting "
                    f"locations {prior!r} and {location!r}")
            section_locations[statement_key][section] = location
        compiled_sections[str(statement).lower()] = compiled
    identities = doc.get("identities") or {}
    if not isinstance(identities, dict):
        raise ValueError("identities must be an object")
    out: Taxonomy = Taxonomy(
        location_vocabulary=allowed_locations,
        statement_sections=compiled_sections,
        section_locations=section_locations,
        identities=identities,
    )
    seen: set[tuple[str, str]] = set()
    for it in doc.get("items", []):
        if not isinstance(it, dict):
            raise ValueError(f"taxonomy item must be an object: {it!r}")
        definition = it.get("definition")
        missing = [key for key in ("fid", "name", "statement")
                   if not str(it.get(key, "")).strip()]
        if not definition:
            missing.append("definition")
        if missing:
            raise ValueError(
                f"taxonomy item is missing {', '.join(missing)}: {it!r}")
        stmt = str(it["statement"]).strip().lower()
        fid = str(it["fid"]).strip()
        key = (stmt, fid)
        if key in seen:
            raise ValueError(f"duplicate taxonomy field {stmt}/{fid}")
        seen.add(key)
        item = dict(it)
        item["statement"] = stmt
        item["fid"] = fid
        mapping = item.get("mapping", {})
        if not isinstance(mapping, dict):
            raise ValueError(f"mapping must be an object for {stmt}/{fid}")
        aliases = mapping.get("aliases", [])
        if aliases is None:
            aliases = []
        if not isinstance(aliases, list):
            raise ValueError(f"aliases must be a list for {stmt}/{fid}")
        item["aliases"] = [str(alias).strip() for alias in aliases
                           if str(alias).strip()]
        if "match_name" not in mapping:
            raise ValueError(
                f"mapping.match_name must be explicit for "
                f"{stmt}/{fid}")
        match_name = mapping.get("match_name", True)
        if not isinstance(match_name, bool):
            raise ValueError(
                f"mapping.match_name must be boolean for {stmt}/{fid}")
        item["match_name"] = match_name
        mode = str(mapping.get("mode", "")).strip()
        expected_mode = (
            "canonical_name_and_aliases" if match_name
            else "aliases_only" if item["aliases"]
            else "rules_only" if mapping.get("rules")
            else "disabled"
        )
        if mode != expected_mode:
            raise ValueError(
                f"mapping.mode for {stmt}/{fid} must be "
                f"{expected_mode!r}, got {mode!r}")
        item["concept"] = _compile_definition(definition, stmt, fid)
        locations = mapping.get("locations", [])
        if not isinstance(locations, list):
            raise ValueError(f"mapping.locations must be a list for {stmt}/{fid}")
        item["locations"] = [
            str(location).strip().upper() for location in locations
            if str(location).strip()
        ]
        invalid_locations = set(item["locations"]) - allowed_locations
        if invalid_locations:
            raise ValueError(
                f"invalid mapping.locations for {stmt}/{fid}: "
                f"{sorted(invalid_locations)}")
        if not item["locations"]:
            raise ValueError(
                f"mapping.locations cannot be empty for "
                f"{stmt}/{fid}; use STATEMENT-WIDE explicitly")
        if mapping.get("location_source") != "declared":
            raise ValueError(
                f"mapping.location_source must be declared for "
                f"{stmt}/{fid}")
        alias_norms = [norm_label(alias) for alias in item["aliases"]]
        if len(alias_norms) != len(set(alias_norms)):
            raise ValueError(
                f"mapping.aliases contains normalized duplicates for "
                f"{stmt}/{fid}")
        rules = mapping.get("rules", [])
        if rules is None:
            rules = []
        if not isinstance(rules, list):
            raise ValueError(f"mapping.rules must be a list for {stmt}/{fid}")
        compiled_rules = []
        for rule in rules:
            if not isinstance(rule, dict):
                raise ValueError(
                    f"mapping rule must be an object for {stmt}/{fid}")
            rule_id = str(rule.get("id", "")).strip()
            status = str(rule.get("status", "")).strip().lower()
            rule_aliases = rule.get("aliases", [])
            rule_locations = [
                str(value).strip().upper()
                for value in (rule.get("locations", []) or [])
                if str(value).strip()
            ]
            rule_scopes = [
                str(value).strip().lower()
                for value in (rule.get("scopes", []) or [])
                if str(value).strip()
            ]
            occurrence = str(rule.get("occurrence", "")).strip().lower()
            if not rule_id or status != "reviewed":
                raise ValueError(
                    f"mapping rule for {stmt}/{fid} requires a stable id and "
                    "status: reviewed")
            if not isinstance(rule_aliases, list) or not rule_aliases:
                raise ValueError(
                    f"mapping rule {rule_id} requires aliases for {stmt}/{fid}")
            if set(rule_locations) - allowed_locations:
                raise ValueError(
                    f"invalid locations in mapping rule {rule_id} for "
                    f"{stmt}/{fid}")
            if set(rule_scopes) - {"standalone", "consolidated"}:
                raise ValueError(
                    f"invalid scopes in mapping rule {rule_id} for {stmt}/{fid}")
            if occurrence not in {"", "first", "last", "only"}:
                raise ValueError(
                    f"invalid occurrence in mapping rule {rule_id} for "
                    f"{stmt}/{fid}")
            compiled = dict(rule)
            compiled["id"] = rule_id
            compiled["status"] = status
            compiled["aliases"] = [
                str(value).strip() for value in rule_aliases
                if str(value).strip()
            ]
            compiled["locations"] = rule_locations
            compiled["scopes"] = rule_scopes
            compiled["occurrence"] = occurrence
            compiled["min_occurrences"] = int(rule.get("min_occurrences", 1))
            for condition in (
                    "locations_present", "locations_absent",
                    "labels_present", "labels_absent"):
                values = [
                    (str(value).strip().upper()
                     if condition.startswith("locations")
                     else norm_label(value))
                    for value in (rule.get(condition, []) or [])
                    if str(value).strip()
                ]
                if (condition.startswith("locations")
                        and set(values) - allowed_locations):
                    raise ValueError(
                        f"invalid {condition} in mapping rule {rule_id} for "
                        f"{stmt}/{fid}")
                compiled[condition] = values
            compiled_rules.append(compiled)
        item["rules"] = compiled_rules
        _validate_taxonomy_semantics(item, stmt, fid)
        out.setdefault(stmt, []).append(item)
    expected_fields = int(doc.get("expected_unique_field_count", -1))
    expected_scoped = int(doc.get("expected_scope_assignment_count", -1))
    actual_fields = sum(len(items) for items in out.values())
    actual_scoped = sum(
        len(item.get("scopes", {}))
        for items in out.values() for item in items)
    if expected_fields != actual_fields or expected_scoped != actual_scoped:
        raise ValueError(
            "taxonomy count guard failed: "
            f"fields {actual_fields}/{expected_fields}, "
            f"scoped fields {actual_scoped}/{expected_scoped}")

    expected_policy = {
        "ambiguity": "reject",
        "sign": "preserve_source",
        "unit_and_time_nature": "strict",
        "total_component_boundary": "exact",
        "model_authority": "proposal_only",
    }
    if doc.get("mapping_policy") != expected_policy:
        raise ValueError(
            f"mapping_policy must be exactly {expected_policy!r}")
    by_key = {
        (statement, item["fid"]): item
        for statement, items in out.items() for item in items
    }
    for statement, items in out.items():
        for item in items:
            fid = item["fid"]
            competitors = {
                str(value).strip()
                for value in item["definition"]["distinguish_from"]
            }
            for competitor in competitors:
                other = by_key.get((statement, competitor))
                if other is None:
                    raise ValueError(
                        f"definition.distinguish_from {competitor!r} for "
                        f"{statement}/{fid} is not a field in the same "
                        "statement")
                reverse = {
                    str(value).strip()
                    for value in other["definition"]["distinguish_from"]
                }
                if fid not in reverse:
                    raise ValueError(
                        f"definition.distinguish_from must be symmetric: "
                        f"{statement}/{fid} -> {competitor}")
    identity_ids = set()
    for statement, checks in identities.items():
        if statement not in out or not isinstance(checks, list):
            raise ValueError(
                f"identities.{statement} must be a list for a known statement")
        valid_fids = {item["fid"] for item in out[statement]}
        scopes_by_fid = {
            item["fid"]: set(item["scopes"]) for item in out[statement]}
        for check in checks:
            if not isinstance(check, dict):
                raise ValueError(
                    f"invalid identity declaration for {statement}: {check!r}")
            expected_keys = {
                "id", "name", "scopes", "result_fid", "terms"}
            if set(check) != expected_keys:
                raise ValueError(
                    f"identity for {statement} must contain exactly "
                    f"{sorted(expected_keys)}: {check!r}")
            identity_id = str(check.get("id", "")).strip()
            name = str(check.get("name", "")).strip()
            scopes = check.get("scopes")
            result_fid = str(check.get("result_fid", "")).strip()
            terms = check.get("terms")
            if (not re.fullmatch(r"[a-z0-9][a-z0-9-]*", identity_id)
                    or identity_id in identity_ids or not name):
                raise ValueError(
                    f"identity for {statement} has an invalid or duplicate id "
                    f"or an empty name: {check!r}")
            identity_ids.add(identity_id)
            if (not isinstance(scopes, list) or not scopes
                    or len(scopes) != len(set(scopes))
                    or set(scopes) - {"standalone", "consolidated"}):
                raise ValueError(
                    f"identity {identity_id!r} has invalid scopes")
            if result_fid not in valid_fids:
                raise ValueError(
                    f"identity {identity_id!r} has invalid result_fid "
                    f"{result_fid!r}")
            if not isinstance(terms, list) or not terms:
                raise ValueError(
                    f"identity {identity_id!r} requires terms")
            term_fids = set()
            required_count = 0
            for term in terms:
                if not isinstance(term, dict) or set(term) != {
                        "fid", "coefficient", "presence"}:
                    raise ValueError(
                        f"identity {identity_id!r} has invalid term "
                        f"shape: {term!r}")
                term_fid = str(term["fid"])
                coefficient = term["coefficient"]
                presence = str(term["presence"])
                if (term_fid not in valid_fids or term_fid in term_fids
                        or term_fid == result_fid
                        or not isinstance(coefficient, (int, float))
                        or isinstance(coefficient, bool)
                        or not float("-inf") < float(coefficient) < float("inf")
                        or float(coefficient) == 0
                        or presence not in {"required", "optional"}):
                    raise ValueError(
                        f"identity {identity_id!r} has invalid term "
                        f"{term!r}")
                term_fids.add(term_fid)
                required_count += presence == "required"
            if not required_count:
                raise ValueError(
                    f"identity {identity_id!r} requires at least one required "
                    "term")
            for identity_scope in scopes:
                unavailable = [
                    fid for fid in [result_fid, *term_fids]
                    if identity_scope not in scopes_by_fid[fid]
                ]
                if unavailable:
                    raise ValueError(
                        f"identity {identity_id!r} is declared for "
                        f"{identity_scope}, but these fields are unavailable: "
                        f"{sorted(unavailable)}")

    configured_statements = set(compiled_sections)
    unknown_statements = configured_statements - set(out)
    if unknown_statements:
        raise ValueError(
            "section configuration references unknown statements: "
            + ", ".join(sorted(unknown_statements)))
    return out


def validate_template_taxonomy(
        template: dict[tuple[str, str], list[ClientField]],
        taxonomy: dict[str, list[dict]]) -> None:
    """Fail fast when the structural template and semantic taxonomy diverge.

    Position, scope and formula role remain authoritative in the template;
    meaning and aliases remain authoritative in the taxonomy. Duplicating
    either source into the other would create two independently stale copies.
    """
    template_fids: dict[str, set[str]] = {}
    for (stmt, _scope), fields in template.items():
        template_fids.setdefault(stmt, set()).update(f.fid for f in fields)
    taxonomy_fids = {
        stmt: {str(item["fid"]) for item in items}
        for stmt, items in taxonomy.items()
    }
    problems = []
    for stmt in sorted(set(template_fids) | set(taxonomy_fids)):
        missing = sorted(template_fids.get(stmt, set())
                         - taxonomy_fids.get(stmt, set()))
        extra = sorted(taxonomy_fids.get(stmt, set())
                       - template_fids.get(stmt, set()))
        if missing:
            problems.append(f"{stmt}: template fields missing from taxonomy: "
                            + ", ".join(missing))
        if extra:
            problems.append(f"{stmt}: taxonomy fields absent from template: "
                            + ", ".join(extra))
    if problems:
        raise ValueError("template/taxonomy mismatch — " + "; ".join(problems))

    # Every active caption collision must have a deterministic separator.
    # Declared location or a reviewed rule must separate candidates. Otherwise
    # startup fails instead of allowing dictionary/set order or a model to
    # choose a field.
    taxonomy_by_key = {
        (statement, item["fid"]): item
        for statement, items in taxonomy.items() for item in items
    }
    collision_problems = []
    for (statement, scope), fields in template.items():
        field_by_fid = {field.fid: field for field in fields}
        vocabulary: dict[str, list[dict]] = {}
        for fid in field_by_fid:
            item = taxonomy_by_key[(statement, fid)]
            labels = list(item.get("aliases", []))
            if item.get("match_name", True):
                labels.append(item["name"])
            for label in labels:
                vocabulary.setdefault(norm_label(label), []).append(item)
        for label, candidates in vocabulary.items():
            unique = {item["fid"]: item for item in candidates}
            values = list(unique.values())
            for index, first in enumerate(values):
                for second in values[index + 1:]:
                    first_locations = set(first.get("locations", []))
                    second_locations = set(second.get("locations", []))
                    location_separates = (
                        "STATEMENT-WIDE" not in
                        first_locations | second_locations
                        and first_locations.isdisjoint(second_locations)
                    )
                    first_rule_labels = {
                        norm_label(alias)
                        for rule in first.get("rules", [])
                        for alias in rule.get("aliases", [])
                    }
                    second_rule_labels = {
                        norm_label(alias)
                        for rule in second.get("rules", [])
                        for alias in rule.get("aliases", [])
                    }
                    rules_separate = (
                        label in first_rule_labels
                        and label in second_rule_labels
                    )
                    if not (location_separates or rules_separate):
                        collision_problems.append(
                            f"{statement}/{scope} caption {label!r}: "
                            f"{first['fid']} vs {second['fid']}")
    if collision_problems:
        raise ValueError(
            "ambiguous active taxonomy aliases — "
            + "; ".join(collision_problems))


# --------------------------------------------------------------------------- periods, values, units

_MONTHS = ("january|february|march|april|may|june|july|august|september|"
           "october|november|december")
_MABBR = "jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec"
# month-day-year accepts abbreviated months ('Jun 30, 2025'); day-month order
# accepts a MISSING separator ('31March 2026') — both appear in the corpus
_DATE = re.compile(rf"({_MONTHS}|{_MABBR})\.?\s+(\d{{1,2}}),?\s+(\d{{4}})|"
                   rf"(\d{{1,2}})[\s.-]*({_MONTHS}|{_MABBR})[\s.,-]+(\d{{2,4}})|"
                   rf"(\d{{2}})[./-](\d{{2}})[./-](\d{{4}})", re.IGNORECASE)
_M2N = {m: i + 1 for i, m in enumerate(_MONTHS.split("|"))}
_M2N.update({m: i + 1 for i, m in enumerate(_MABBR.split("|"))})


@dataclass
class Period:
    span: str
    end: str
    audited: str
    col: int
    raw: str = ""


def _parse_period(text: str, col: int) -> Period:
    t = " ".join(text.split()).lower()
    span = "?"
    if re.search(r"three\s*months|\bquarter\b|\b3\s*months", t):
        span = "3M"
    elif re.search(r"six\s*months|half\s*year|\b6\s*months", t):
        span = "6M"
    elif re.search(r"nine\s*months|\b9\s*months", t):
        span = "9M"
    elif re.search(r"year\s*ended|twelve\s*months", t):
        span = "FY"
    end = ""
    m = _DATE.search(t)
    if m:
        g = m.groups()
        if g[0]:
            end = f"{g[2]}-{_M2N[g[0].lower()]:02d}-{int(g[1]):02d}"
        elif g[3]:
            yr = int(g[5])
            yr = yr + 2000 if yr < 100 else yr
            end = f"{yr}-{_M2N[g[4].lower()]:02d}-{int(g[3]):02d}"
        else:
            end = f"{g[8]}-{int(g[7]):02d}-{int(g[6]):02d}"
    if not end:
        # Indian shorthand headers: 'Q1 FY25', 'FY 2024', 'FY25' (FY ends March 31)
        _QEND = {1: "06-30", 2: "09-30", 3: "12-31", 4: "03-31"}
        m = re.search(r"\bq([1-4])\s*fy\s*(\d{2,4})\b", t)
        if m:
            qn, yr = int(m.group(1)), int(m.group(2))
            yr = yr + 2000 if yr < 100 else yr           # FY25 -> FY ending Mar 2025
            cal = yr if qn == 4 else yr - 1              # Q1-Q3 fall in the prior calendar year
            end = f"{cal}-{_QEND[qn]}"
            span = "3M"
        else:
            m = re.search(r"\bfy\s*(\d{2,4})\b", t)
            if m:
                yr = int(m.group(1))
                yr = yr + 2000 if yr < 100 else yr
                end = f"{yr}-03-31"
                span = "FY"
    audited = ("unaudited" if "unaudited" in t else "audited" if "audited" in t else "")
    return Period(span, end, audited, col, text.strip()[:70])


def infer_spans(periods: list[Period]) -> list[Period]:
    """SEBI results layouts often print only dates. Infer spans: a duplicated
    end-date's FIRST occurrence is the 3M column, the SECOND the cumulative
    (6M for Sep, 9M for Dec, FY for Mar); a solitary March end is FY; other
    solitary ends are 3M."""
    seen: dict[str, int] = {}
    for p in periods:
        if not p.end:
            continue
        seen[p.end] = seen.get(p.end, 0) + 1
        if p.span != "?":
            continue
        month = int(p.end[5:7])
        if seen[p.end] == 1:
            p.span = "FY" if (month == 3 and sum(1 for q in periods if q.end == p.end) == 1) else "3M"
        else:
            p.span = {6: "3M", 9: "6M", 12: "9M", 3: "FY"}.get(month, "?")
    return periods


_GROUPED_NUMBER = re.compile(
    r"-?(?:\d{1,3}(?:,\d{3})+|\d{1,2}(?:,\d{2})+,\d{3}|\d+)"
    r"(?:\.\d+)?$"
)
_SPACE_GROUPED_NUMBER = re.compile(
    r"-?(?:\d{1,3}(?:\s+\d{3})+|\d{1,2}(?:\s+\d{2})+\s+\d{3})"
    r"(?:\.\d+)?$"
)


def _normalize_numeric_spacing(value: str) -> str:
    """Repair whitespace inserted *inside one printed numeric token*.

    Accept only standard western/Indian thousands grouping, so OCR forms such
    as ``11 ,778`` and ``90 828`` become ``11,778`` and ``90828`` while two
    separate values accidentally sharing a cell (``2026 2025``) stay invalid.
    """
    s = str(value)
    punct = re.sub(r"\s*([,.])\s*", r"\1", s)
    if punct != s and _GROUPED_NUMBER.fullmatch(punct):
        return punct
    if _SPACE_GROUPED_NUMBER.fullmatch(s):
        return re.sub(r"\s+", "", s)
    return s


def _num(cell, mode: str = "clean") -> float | None:
    """Parse a numeric cell. mode='clean' is the original behaviour (commas are
    thousands separators). mode='repair' is used ONLY for statements whose text
    layer is detected as corrupted (decimal point rendered as a comma or space,
    e.g. '296,99' → 296.99, '218 92' → 218.92); clean tokens are still parsed
    exactly as before, so a repair-mode grid's already-correct cells are safe."""
    s = str(cell or "").strip().strip("*^#@")
    if not s or s in "-–—" or s.lower() in ("nil", "na", "n.a.", "-"):
        return None
    # a value in parentheses is negative; tolerate an OCR-mangled CLOSING paren
    # ('(17,253'' or '(17,253`' for '(17,253)') so the cell parses instead of
    # silently dropping — the opening '(' plus a digits body is unambiguous.
    neg = s.startswith("(") and (s.endswith(")") or bool(re.search(r"\d[)'`\"]?$", s)))
    s = s.strip("()'`\"").replace("₹", "").replace("`", "").strip()
    s = _normalize_numeric_spacing(s)
    if mode == "repair":
        v = _num_repair(s)
        return None if v is None else (-v if neg else v)
    s = s.replace(",", "").strip()
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


def _num_repair(s: str) -> float | None:
    """Decimal-safe parse for a filing with a corrupted text layer. Indian
    statement amounts are 2 dp and a real thousands group is always 3 digits, so
    a separator followed by EXACTLY 2 trailing digits is the decimal point; any
    other separators are thousands. Clean/standard numbers are returned as-is."""
    s = s.strip()
    if not s:
        return None
    # 1) already a clean standard number (incl. plain integers and 1+ dp) — untouched
    if re.fullmatch(r"\d{1,3}(,\d{3})*(\.\d+)?", s) or re.fullmatch(r"\d+(\.\d+)?", s):
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    # 2) corrupted decimal: final separator + exactly 2 digits is the fraction
    m = re.search(r"[.,\s](\d{2})$", s)
    if m:
        head = re.sub(r"[.,\s]", "", s[:m.start()]) or "0"
        try:
            return float(f"{head}.{m.group(1)}")
        except ValueError:
            return None
    # 3) no 2-digit fraction → integer with thousands/space separators (or noise)
    digits = re.sub(r"[.,\s]", "", s)
    return float(digits) if re.fullmatch(r"\d+", digits) else None


def detect_number_format(grid) -> str:
    """Classify a statement grid's number format from its own cells. Returns
    'repair' when the text layer shows the corrupted-decimal signature (a comma
    or space followed by exactly 2 trailing digits with no proper '.'-decimal,
    or a multi-dot token like '3.485.21'); otherwise 'clean'. Clean and
    integer-only filings (e.g. '12,651') never trip the signature, so they stay
    on the untouched clean path."""
    def is_numeric(cell):
        s = str(cell or "").strip()
        return bool(re.search(r"\d", s)) and bool(re.fullmatch(r"[()\d.,\s₹%+\-]+", s))

    def corrupted(cell):
        s = str(cell or "").strip().strip("()").replace("₹", "").strip()
        if not re.search(r"\d", s):
            return False
        if re.search(r"\.\d", s):                       # has a normal .decimal …
            return len(re.findall(r"\.", s)) >= 2       # … unless it's 3.485.21 style
        return bool(re.search(r"[,\s]\d{2}$", s))       # ',dd' or ' dd' trailing → corrupted

    total = corrupt = 0
    for row in grid:
        for cell in list(row)[1:]:
            if is_numeric(cell):
                total += 1
                corrupt += corrupted(cell)
    return "repair" if corrupt >= 2 else "clean"


_UNITS = [("crore", 1e7), ("lakh", 1e5), ("million", 1e6), ("mn", 1e6),
          ("billion", 1e9), ("'000", 1e3), ("thousand", 1e3)]


def detect_units(grid) -> tuple[str, float, str]:
    blob = " ".join(str(c) for row in grid[:6] for c in row if c).lower()
    unit, mult = "", 1.0
    for name, m in _UNITS:
        if name in blob:
            unit, mult = name, m
            break
    cur = "USD" if ("usd" in blob or "us$" in blob) else "INR"
    return unit, mult, cur


# --------------------------------------------------------------------------- definition-driven mapping

_MAP_SCHEMA = {
    "type": "object",
    "properties": {
        "assignments": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "line": {"type": "integer"},   # the L<number> of the input line
                    "fid": {"type": "string"},     # "" when no field fits
                },
                "required": ["line", "fid"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["assignments"],
    "additionalProperties": False,
}


def _mapping_schema(line_count: int, valid_fids: set[str]) -> dict:
    """Constrain model output to this exact statement invocation."""
    import copy

    schema = copy.deepcopy(_MAP_SCHEMA)
    assignments = schema["properties"]["assignments"]
    assignments["minItems"] = line_count
    assignments["maxItems"] = line_count
    item = assignments["items"]["properties"]
    item["line"]["minimum"] = 1
    item["line"]["maximum"] = line_count
    item["fid"]["enum"] = [""] + sorted(valid_fids, key=lambda x: (len(x), x))
    return schema

_PROPOSAL_INSTR = """You review unresolved lines of an Indian company's financial statement and
PROPOSE a client-taxonomy field by economic meaning. Your response is advisory only and is never
written to the authoritative report.

Rules:
- Propose the ONE field whose definition the report line appears to satisfy; return fid "" if no
  definition fits or if the evidence is ambiguous (do NOT force a fit).
- Respect total-vs-component guards in definitions: a printed subtotal/total line maps only to a
  field defined as that aggregate — NEVER to a component field.
- Lines under a 'Components of cash and cash equivalents' block (balances with banks,
  'on current account', EEFC accounts, deposits with original maturity < 3 months, cash in
  hand, cheques on hand, remittances in transit) are the BREAKDOWN of closing cash, which is
  already mapped — return "" for every one of them.
- A printed TOTAL maps ONLY to the field for exactly that aggregate; if no such field exists,
  return "" (leave it unmapped). Never map a broader total into a narrower field (e.g. 'TOTAL
  LIABILITIES' must NOT go into 'Total Current Liabilities') and never map a per-segment or
  per-category subtotal (e.g. 'Total of IT Services') into an overall-total field.
- Several lines may map to the SAME fid when its definition says components are summed there.
- Ignore heading-only lines, period-header lines and blank lines (they are not in your input).
- Lines are prefixed with their statement SECTION (e.g. [CURRENT LIABILITIES], [INVESTING]) —
  use it: the same label maps to different fields in different sections.
- Refer to lines by their L<number>; every input line must appear exactly once in assignments.
- When a label appears TWICE and no section prefixes are given, Schedule III order applies:
  the FIRST occurrence is the NON-CURRENT item, the SECOND the CURRENT one.
- The aliases in definitions are ILLUSTRATIVE examples, not a match list."""


@dataclass
class MappedStatement:
    periods: list[Period]
    facts: dict[str, dict[int, float]]
    sources: dict[str, list[str]]
    unmapped: list[str]
    verification: list[str]
    n_checks: int = 0
    n_ok: int = 0
    unit: str = ""
    multiplier: float = 1.0
    currency: str = "INR"
    # statement-level review flags. They ONLY accumulate down the pipeline —
    # extraction ⚠ titles land here, mapped-layer identity failures append
    # here, and the writer surfaces every entry in the deliverable. A flag can
    # never be silently dropped between layers.
    flags: list = None

    def flag(self, msg: str) -> None:
        cur = getattr(self, "flags", None)   # tolerate pre-flags pickles
        if cur is None:
            cur = []
            self.flags = cur
        if msg not in cur:
            cur.append(msg)


def _label_and_vals(row, mode: str = "clean"):
    """Label = first SUBSTANTIVE text cell (grids may lead with Sl-No or
    enumerator cells like 'B' / 'iv.' / '(A)'); values = numeric cells after
    the label cell. Enumerator cells are kept as a prefix."""
    cand = []
    for j, c in enumerate(row):
        s = str(c or "").strip()
        if s and any(ch.isalpha() for ch in s) and _num(s, mode) is None:
            cand.append((j, s))
    if not cand:
        return "", {}
    li, label = next(((j, s) for j, s in cand if len(s) > 4), cand[-1])
    pres = [s for j, s in cand if j < li and len(s) <= 6]
    if pres:
        label = " ".join(pres) + " " + label
    vals = {j: _num(row[j], mode) for j in range(li + 1, len(row))
            if _num(row[j], mode) is not None}
    return label, vals


def parse_periods(grid, header=None) -> list[Period]:
    """Period per value column, parsed COLUMN-FIRST.

    Each column's period comes from ITS OWN header cells; a banner row (a
    single period-worded cell spanning all columns, e.g. 'For the six months
    ended September 30,' with the years on the next row) is borrowed ONLY when
    a column has no date of its own. Critically, a document TITLE that happens
    to contain a date — 'Audited Standalone Financial Results for the three
    months and year ended March 31, 2026' — is NOT a banner; treating it as one
    stamps that single date onto every column and collapses all periods into
    one (silently dropping columns in the wide output)."""
    if header is None:
        header, data = _header_and_data(grid)
    else:
        _, data = _header_and_data(grid)
    ncol = max(len(r) for r in grid)
    # VALUE columns only: a period is created for a column that carries numeric
    # DATA, never for an enumerator ('Sl No'/'Sr No') or the label ('Particulars')
    # column. Otherwise a filing with a leading serial column makes the label
    # column look like a period and emits a spurious '(unresolved period)'.
    from collections import Counter as _Counter
    _vc: _Counter = _Counter()
    for r in data:
        _, _vals = _label_and_vals(r)
        for _c in _vals:
            _vc[_c] += 1
    value_cols = {c for c, n in _vc.items() if n >= 2}
    _TITLE = re.compile(r"results|financial statement|\blimited\b|\bltd\b|\bcin\b|"
                        r"balance sheet|cash flow|profit and loss|regd|website|corporate",
                        re.IGNORECASE)
    banner_parts = []
    for r in header:
        cells = [str(c).strip() for c in r if str(c).strip()]
        if (len(cells) == 1 and len(cells[0]) < 55
                and re.search(r"month|quarter|\byear\b|ended|as at", cells[0], re.I)
                and not _TITLE.search(cells[0])):
            banner_parts.append(cells[0])
    banner = " ".join(banner_parts)
    banner_span = _parse_period(banner, 0).span if banner else "?"
    periods = []
    for j in range(1, ncol):
        if value_cols and j not in value_cols:   # skip enumerator/label columns
            continue
        text = " ".join(str(r[j]) for r in header if j < len(r) and r[j]).strip()
        if not text:
            continue
        p = _parse_period(text, j)               # column's OWN cells first
        if not p.end and banner:                 # no date of its own -> borrow banner
            p = _parse_period((banner + " " + text).strip(), j)
        elif p.end and p.span == "?" and banner_span != "?":
            p.span = banner_span                 # has date, borrow only the span word
        periods.append(p)
    return infer_spans(periods)


def _periods_fallback(grid, data, mode: str = "clean") -> list[Period]:
    """Rebuild periods when parse_periods() found none. Vision/OCR transcriptions
    of cash-flow statements often collapse the period header into the data region
    (the banner 'For the six months ended September 30, 2025' and bare years
    '2024' land on the first row alongside a label), so _header_and_data() eats
    the header and parse_periods() returns []. Here we recover the period banner +
    year tokens from the grid's top rows and align one period to each column that
    actually carries data. Fires ONLY on the empty-periods failure, so statements
    that already parse periods are never touched."""
    from collections import Counter
    colcnt: Counter = Counter()
    for r in data:
        _, vals = _label_and_vals(r, mode)
        for c, v in vals.items():                       # ignore bare years (period
            if not (v == int(v) and 1900 <= v <= 2099): # markers), count real data
                colcnt[c] += 1
    cols = sorted(c for c, n in colcnt.items() if n >= 2)
    if not cols:
        return []
    src = " ".join(str(c) for row in grid[:5] for c in row if str(c).strip())
    banners = re.findall(r"(?:three|six|nine|twelve)\s+months?\s+ended|quarter\s+ended|"
                         r"year\s+ended", src, re.I)
    m = re.search(r"((?:three|six|nine|twelve)\s+months?|quarter|year)\s+ended\s+"
                  r"([A-Za-z]+\.?\s+\d{1,2})", src, re.I)
    yrs_raw = re.findall(r"\b(?:19|20)\d{2}\b", src)
    seen: set = set()
    years = [y for y in yrs_raw if not (y in seen or seen.add(y))]
    # SAFE-ONLY: recover only the unambiguous case — ONE banner phrasing that
    # maps 1:1 to every data column (e.g. a two-column half-year cash flow).
    # Mixed banners (Quarter + Year in one segment header) or a year/column
    # count mismatch means we cannot place columns confidently, so decline and
    # leave the sheet untouched (empty) rather than emit WRONG periods.
    if not (m and years) or len(years) != len(cols):
        return []
    if len({b.lower().replace("quarter ended", "3m").split(" ended")[0] for b in banners}) > 1:
        return []
    base = f"{m.group(1)} ended {m.group(2)}"
    periods = [_parse_period(f"{base} {yr}", col) for col, yr in zip(cols, years)]
    if not all(p.end for p in periods):
        return []
    return infer_spans(periods)


def _header_and_data(grid, mode: str = "clean"):
    first = None
    for i, row in enumerate(grid):
        label, vals = _label_and_vals(row, mode)
        # A period-header row can carry a stray bare year ('Particulars |
        # Quarter ended March 31, 2026 | ... | 2025') — its only numeric cell is
        # a YEAR, not an amount. Such a row is still a HEADER, not the first data
        # row; requiring a NON-YEAR value keeps the period header out of the data
        # region so parse_periods can read it.
        real = {c: v for c, v in vals.items()
                if not (v == int(v) and 1900 <= v <= 2099)}
        if label and real:
            first = i
            break
    if first is None:
        first = min(3, len(grid))
    return grid[:first], grid[first:]


def map_statement(grid: list[list[str]], stmt: str, taxonomy: dict[str, list[dict]],
                  template_fields: list[ClientField], model: str | None = None,
                  scope: str | None = None) -> MappedStatement:
    """Map one raw statement using reviewed taxonomy rules, then verify formulas.

    ``model`` remains in the public signature for callers from older releases,
    but model output is deliberately not consumed here.  Advisory model
    suggestions are produced by :func:`propose_unmapped_mappings` as a separate
    review artifact and therefore cannot change reported facts.
    """

    # Detect the grid's number format up front. 'repair' is used only when the
    # text layer shows the corrupted-decimal signature; clean/integer filings
    # stay on the original parse path (mode='clean'), so nothing that works today
    # is affected.
    num_mode = detect_number_format(grid)
    header, data = _header_and_data(grid, num_mode)
    periods = parse_periods(grid, header)
    # parse_periods returns [] when a mangled/vision header row got eaten, but it
    # can also return UNUSABLE periods when a consensus vision re-read leaves
    # '⚠ a | b' disagreement markers or no dates in the header. Treat both as
    # failures and rebuild from the banner + year tokens — but only ADOPT the
    # rebuild when it actually recovers dated periods, so clean statements (which
    # have real dates and no ⚠) are never altered.
    if (not periods or all(not p.end for p in periods)
            or any("⚠" in (p.raw or "") for p in periods)):
        fb = _periods_fallback(grid, data, num_mode)
        if fb:
            periods = fb
    unit, mult, cur = detect_units(grid)

    rows = []
    section = ""
    section_patterns = getattr(
        taxonomy, "statement_sections", {}).get(stmt, ())
    for hrow in header:
        for header_cell in hrow:
            hl = norm_label(header_cell)
            for pat, name, _location in section_patterns:
                if re.search(pat, hl):
                    section = name
                    break
    for row in data:
        label, vals = _label_and_vals(row, num_mode)
        if label and any(c.isalpha() for c in label):
            ll = norm_label(label)
            for pat, name, _location in section_patterns:
                if re.search(pat, ll):
                    section = name
                    break
            if vals:
                rows.append((label, section, vals))

    # deterministic layer first: duplicate names resolved by template structure
    valid_all = {d["fid"] for d in taxonomy.get(stmt, [])}
    pairs = derived_section_pairs(template_fields, taxonomy.get(stmt, []))
    exact: dict[str, set[str]] = {}
    reviewed_rules: dict[str, list[tuple[str, dict, list[str]]]] = {}
    byfid = {f.fid: f for f in template_fields}
    for d in taxonomy.get(stmt, []):
        if d["fid"] not in byfid:
            continue
        labels = list(d.get("aliases", []))
        if d.get("match_name", True):
            labels.insert(0, d.get("name", ""))
        for label in labels:
            normalized = norm_label(label)
            if normalized:
                exact.setdefault(normalized, set()).add(d["fid"])
        for rule in d.get("rules", []):
            for label in rule.get("aliases", []):
                normalized = norm_label(label)
                if normalized:
                    reviewed_rules.setdefault(normalized, []).append(
                        (d["fid"], rule, d.get("locations", [])))
    declared_locations = {
        str(item["fid"]): set(item.get("locations", []))
        for item in taxonomy.get(stmt, [])
        if item.get("locations")
    }
    locations = {
        field.fid: declared_locations.get(field.fid, set())
        for field in template_fields
    }
    value_types = {
        str(item["fid"]): str(item.get("value_type", ""))
        for item in taxonomy.get(stmt, [])
    }
    from collections import Counter
    label_sec_count = Counter((norm_label(lab), sec) for lab, sec, _v in rows)
    label_sec_positions: dict[tuple[str, str], list[int]] = {}
    for idx, (lab, sec, _vals) in enumerate(rows):
        label_sec_positions.setdefault((norm_label(lab), sec), []).append(idx)
    secs_present = {sec for _l, sec, _v in rows if sec}
    section_locations = taxonomy.section_locations.get(stmt, {})
    locations_present = {
        section_locations.get(sec, "") for sec in secs_present
        if section_locations.get(sec, "")
    }
    labels_present = {norm_label(label) for label, _sec, _vals in rows}
    # a statement with separate external-customer / inter-segment revenue
    # sub-sections prints several bare totals — revenue-side pins are unsafe there
    multi_revenue = bool(
        locations_present & {"EXTERNAL-REVENUE", "INTER-SEGMENT"})
    pre: dict[int, str] = {}
    for i, (lab, sec, vals) in enumerate(rows):
        nl = norm_label(lab)
        row_location = section_locations.get(sec, "")
        # Explicit reviewed rules are more specific than the field's default
        # alias list. They resolve the few cases where the same printed caption
        # has different meanings by scope, location, or occurrence.
        ruled: list[tuple[str, int]] = []
        positions = label_sec_positions.get((nl, sec), [])
        for fid, rule, default_locations in reviewed_rules.get(nl, []):
            rule_scopes = set(rule.get("scopes", []))
            rule_locations = set(rule.get("locations", []) or default_locations)
            if (fid not in valid_all
                    or (rule_scopes and scope not in rule_scopes)
                    or (rule_locations and row_location not in rule_locations)
                    or not set(rule.get("locations_present", [])).issubset(
                        locations_present)
                    or set(rule.get("locations_absent", [])) & locations_present
                    or not set(rule.get("labels_present", [])).issubset(
                        labels_present)
                    or set(rule.get("labels_absent", [])) & labels_present
                    or len(positions) < int(rule.get("min_occurrences", 1))):
                continue
            occurrence = rule.get("occurrence", "")
            if (occurrence == "first" and positions and i != positions[0]
                    or occurrence == "last" and positions and i != positions[-1]
                    or occurrence == "only" and len(positions) != 1):
                continue
            score = (10 + bool(rule_scopes) + bool(rule_locations)
                     + bool(occurrence)
                     + len(rule.get("locations_present", []))
                     + len(rule.get("locations_absent", []))
                     + len(rule.get("labels_present", []))
                     + len(rule.get("labels_absent", [])))
            ruled.append((fid, score))
        if ruled:
            top = max(score for _fid, score in ruled)
            candidates = {fid for fid, score in ruled if score == top}
        else:
            candidates = {
                fid for fid in exact.get(nl, set())
                if (fid in valid_all
                    and (not row_location
                         or "STATEMENT-WIDE" in locations.get(fid, set())
                         or not locations.get(fid)
                         or row_location in locations[fid]))
            }
        value_type_cue = _label_value_type(lab)
        if value_type_cue:
            typed = {
                fid for fid in candidates
                if value_types.get(fid) == value_type_cue
            }
            if typed:
                candidates = typed
        # A printed "Total ..." cannot be assigned to a leaf when an exact,
        # reviewed aggregate candidate exists. This resolves legacy component
        # captions such as "Other Current Liabilities" that historically also
        # carried the alias "Total current liabilities".
        if nl.startswith(("total ", "grand total")):
            aggregates = {
                fid for fid in candidates
                if byfid.get(fid) and byfid[fid].formula
            }
            if aggregates:
                candidates = aggregates
        if len(candidates) == 1:
            fid = next(iter(candidates))
            pre[i] = fid
            continue
        pair = pairs.get(nl)
        if pair and row_location == "SEGMENT-REVENUE" and multi_revenue:
            pair = None
        if pair and sec and label_sec_count[(nl, sec)] == 1:
            side = row_location
            fid = pair.get(side)
            if fid and fid in valid_all:
                pre[i] = fid
    defs = taxonomy.get(stmt, [])
    assign: dict[int, str] = dict(pre)

    valid_fids = {d["fid"] for d in defs}
    # a printed TOTAL line subsumes its components: if a fid was assigned both
    # a total-labelled line and component lines, keep only the total(s)
    by_fid: dict[str, list[int]] = {}
    for i in range(len(rows)):
        fid = assign.get(i, "")
        if fid and fid in valid_fids:
            by_fid.setdefault(fid, []).append(i)
    for fid, idxs in by_fid.items():
        totals = [i for i in idxs if norm_label(rows[i][0]).startswith("total")]
        if totals and len(totals) < len(idxs):
            for i in idxs:
                if i not in totals:
                    assign.pop(i, None)
        elif len(idxs) >= 3:
            # arithmetic total-detection: if ONE line equals the sum of all the
            # others (every column), it IS their printed total — keep only it
            # (e.g. 'Cash at the end' followed by its components breakdown)
            def _vec(i, cols):
                return [rows[i][2].get(c) for c in cols]
            cols = sorted({c for i in idxs for c in rows[i][2]})
            hits = []
            for i in idxs:
                rest = [j for j in idxs if j != i]
                ok = True
                for c in cols:
                    tot = rows[i][2].get(c)
                    ssum = sum(rows[j][2].get(c, 0.0) for j in rest)
                    if tot is None or abs(ssum - tot) > max(0.02, 0.005 * abs(tot)):
                        ok = False
                        break
                if ok:
                    hits.append(i)
            if len(hits) == 1:
                for j in idxs:
                    if j != hits[0]:
                        assign.pop(j, None)
    # duplicate-total pruning: filings print the same figure twice (TCI in
    # the body AND as the 'attributable to' block total) or as a running
    # subtotal chain ('PBT before exceptional items' followed by 'PBT').
    # If one fid holds two such rows it double-counts. Purely arithmetic:
    #   identical value vectors            -> same printed fact repeated
    #   B - A == sum(rows strictly between)-> B is a later running subtotal
    # keep the row whose label best matches the client field name.
    from itertools import combinations as _combos
    _fname = {tf.fid: norm_label(tf.name) for tf in template_fields}
    _byfid: dict[str, list[int]] = {}
    for _i in range(len(rows)):
        _fd = assign.get(_i)
        if _fd and _fd in valid_fids and rows[_i][2]:
            _byfid.setdefault(_fd, []).append(_i)
    for _fd, _idxs in _byfid.items():
        if len(_idxs) < 2:
            continue
        _drop: set = set()
        for _a, _b in _combos(sorted(_idxs), 2):
            if _a in _drop or _b in _drop:
                continue
            _va, _vb = rows[_a][2], rows[_b][2]
            _cc = sorted(set(_va) & set(_vb))
            if len(_cc) < 2:
                continue
            _same = all(abs(_va[c] - _vb[c]) <= 0.02 for c in _cc)
            _chain = False
            if not _same:
                _mids = [rows[_m][2] for _m in range(_a + 1, _b)]
                _chain = all(abs((_vb[c] - _va[c])
                                 - sum(_mv.get(c, 0.0) for _mv in _mids)) <= 0.02
                             for c in _cc)
            if not (_same or _chain):
                continue
            def _sim(i):
                _lt = set(norm_label(rows[i][0]).split())
                _nt = set(_fname.get(_fd, "").split())
                return (len(_lt & _nt) / max(1, len(_lt | _nt)), i)
            _keep = max((_a, _b), key=_sim)
            _drop.add(_a if _keep == _b else _b)
        for _i in _drop:
            assign.pop(_i, None)

    # A field that is a COMPUTED SUBTOTAL (has a template formula) corresponds to
    # exactly ONE printed total line — it must never SUM two different printed
    # totals. The LLM occasionally assigns a broader total to a narrower field
    # (e.g. both 'Total current liabilities' AND the grand 'TOTAL LIABILITIES'
    # land on fid 13775, summing to a doubled value). When ≥2 total-labelled
    # lines survive for such a field, keep the one the ARITHMETIC picks — the
    # candidate that equals the sum of the field's mapped formula components;
    # label similarity only breaks a tie the arithmetic cannot.
    _has_formula = {f.fid for f in template_fields if f.formula}
    _formula_of = {f.fid: f.formula for f in template_fields if f.formula}
    _row2fid_t = {f.row: f.fid for f in template_fields}
    _byfid3: dict[str, list[int]] = {}
    for _i in range(len(rows)):
        _fd = assign.get(_i)
        if _fd and _fd in valid_fids and rows[_i][2]:
            _byfid3.setdefault(_fd, []).append(_i)

    def _fid_vals(fid, skip: set):
        """Tentative per-column fact for `fid` from the current assignment."""
        acc: dict[int, float] = {}
        for _i2 in range(len(rows)):
            if _i2 in skip or assign.get(_i2) != fid:
                continue
            for _c, _v in rows[_i2][2].items():
                acc[_c] = acc.get(_c, 0.0) + _v
        return acc

    for _fd, _idxs in _byfid3.items():
        if _fd not in _has_formula:
            continue
        _tot = [i for i in _idxs
                if norm_label(rows[i][0]).startswith(("total", "grand"))]
        if len(_tot) < 2:
            continue
        _tot_set = set(_tot)
        _arith = []
        for _i in _tot:
            _comp: dict[int, float] = {}
            _terms = 0
            for _sg, _rr in _formula_of[_fd]:
                _cf = _row2fid_t.get(_rr)
                if not _cf or _cf == _fd:
                    continue
                _cv = _fid_vals(_cf, _tot_set)
                if _cv:
                    _terms += 1
                    for _c, _v in _cv.items():
                        _comp[_c] = _comp.get(_c, 0.0) + _sg * _v
            _cand = rows[_i][2]
            _cc = sorted(set(_comp) & set(_cand))
            if _terms >= 2 and _cc and all(
                    abs(_comp[_c] - _cand[_c]) <= max(1.0, 0.01 * abs(_cand[_c]))
                    for _c in _cc):
                _arith.append(_i)
        if len(_arith) == 1:
            _keep = _arith[0]
        else:
            _fn = set(_fname.get(_fd, "").split())
            def _match(i, _fn=_fn):
                _lt = set(norm_label(rows[i][0]).split())
                return (len(_lt & _fn) / max(1, len(_lt | _fn)), -i)
            _keep = max(_arith or _tot, key=_match)
        for _i in _tot:
            if _i != _keep:
                assign.pop(_i, None)

    facts: dict[str, dict[int, float]] = {}
    sources: dict[str, list[str]] = {}
    sources_vals: dict[str, dict[int, list]] = {}
    unmapped = []
    unmapped_vals: list = []
    for i, (label, sec, vals) in enumerate(rows):
        fid = assign.get(i, "")
        if fid and fid in valid_fids:
            tgt = facts.setdefault(fid, {})
            sv = sources_vals.setdefault(fid, {})
            for j, v in vals.items():
                tgt[j] = tgt.get(j, 0.0) + v
                sv.setdefault(j, []).append((label, v))
            sources.setdefault(fid, []).append(label)
        else:
            unmapped.append(label)
            unmapped_vals.append((label, dict(vals)))

    # verification via the client template's own formulas
    verification = []
    n_checks = n_ok = 0
    row2fid = {f.row: f.fid for f in template_fields}
    fidname = {tf.fid: tf.name for tf in template_fields}
    col2per = {p.col: p for p in periods}
    for f in template_fields:
        if not f.formula or f.fid not in facts:
            continue
        for j in sorted(facts[f.fid].keys()):
            total, have = 0.0, 0
            for sign, r in f.formula:
                cf = row2fid.get(r)
                if cf and cf in facts and j in facts[cf]:
                    total += sign * facts[cf][j]
                    have += 1
            # a partial sum proves nothing: verify only when most formula
            # terms are actually mapped from the filing
            if have < 2 or have < 0.6 * len(f.formula):
                continue
            printed = facts[f.fid][j]
            tol = max(1.0, 0.01 * abs(printed))
            ok = abs(total - printed) <= tol
            n_checks += 1
            n_ok += ok
            if not ok:
                per = col2per.get(j)
                plabel = _period_label(per) if per else f"column {j}"
                comps = []
                for sign, r in f.formula:
                    cf = row2fid.get(r)
                    if cf and cf in facts and j in facts[cf]:
                        comps.append(f"{'+' if sign > 0 else '-'} {fidname.get(cf, cf)} "
                                     f"({cf}) = {facts[cf][j]:,.2f}")
                # UNDERSHOOT (components < reported total) = a printed line has no
                # dedicated client field, so the reconstruction is incomplete —
                # the reported total is a printed figure and is correct. This is
                # an informational coverage note, NOT a correctness failure, so it
                # does not flag the statement. OVERSHOOT (components > total) means
                # a line was double-counted into two fields; that DOES corrupt a
                # value, and it surfaces as a broken printed cross-identity in
                # verify_mapped, which flags. Either way this message is Audit-only.
                short = printed - total
                kind = ("components incomplete — a printed line has no dedicated "
                        "field; reported total is the printed figure"
                        if short > 0 else
                        "components exceed the reported total — a line may be "
                        "double-counted (see cross-identity checks)")
                verification.append(
                    f"note: {f.name} [{f.fid}] — {plabel}: reported {printed:,.2f}, "
                    f"mapped components sum to {total:,.2f} ({kind}). "
                    f"Components: {'  '.join(comps)}.")
    ms = MappedStatement(periods=periods, facts=facts, sources=sources,
                         unmapped=unmapped, verification=verification,
                         n_checks=n_checks, n_ok=n_ok,
                         unit=unit, multiplier=mult, currency=cur)
    ms.sources_vals = sources_vals            # exact per-line provenance
    ms.unmapped_vals = unmapped_vals
    ms.unmapped_details = [
        (label, sec, dict(vals))
        for i, (label, sec, vals) in enumerate(rows)
        if not (assign.get(i, "") and assign.get(i, "") in valid_fids)
    ]
    return ms


def propose_unmapped_mappings(
        mapped: dict[tuple[str, str], MappedStatement],
        taxonomy: dict[str, list[dict]],
        template: dict[tuple[str, str], list[ClientField]],
        model: str | None = None) -> dict:
    """Return non-authoritative model suggestions for unresolved report lines.

    This function is intentionally separate from :func:`map_quarter`.  It
    receives a completed deterministic mapping and returns a standalone review
    payload; it never mutates ``MappedStatement.facts``, sources, verification,
    or flags. Consequently model sampling, retries, outages, and future model
    versions cannot change the client workbook.
    """
    from src.llm import extract_json

    proposals = []
    for stmt, scope in sorted(mapped):
        ms = mapped[(stmt, scope)]
        details = getattr(ms, "unmapped_details", None)
        if details is None:
            details = [
                (label, "", values)
                for label, values in (getattr(ms, "unmapped_vals", None) or [])
            ]
        if not details:
            continue
        fields = template.get((stmt, scope), [])
        byfid = {field.fid: field for field in fields}
        defs = [item for item in taxonomy.get(stmt, [])
                if item["fid"] in byfid]
        valid_fids = {item["fid"] for item in defs}
        definition_lines = []
        for item in defs:
            aliases = "; ".join(item.get("aliases", [])[:5])
            field = byfid[item["fid"]]
            role = ("ratio" if field.ratio else
                    "formula-verified aggregate" if field.formula else
                    "reported field")
            definition_lines.append(
                f"fid={item['fid']} | {item['name']} | ROLE: {role} | "
                f"{item['concept']}"
                + (f" | e.g.: {aliases}" if aliases else ""))
        result = extract_json(
            instructions=_PROPOSAL_INSTR,
            user_input=(
                "FIELD DEFINITIONS:\n" + "\n".join(definition_lines)
                + "\n\nUNRESOLVED STATEMENT LINES:\n"
                + "\n".join(
                    f"L{line}: " + (f"[{section}] " if section else "") + label
                    for line, (label, section, _values)
                    in enumerate(details, 1))),
            schema_name="mapping_proposals",
            schema=_mapping_schema(len(details), valid_fids),
            model=model, max_output_tokens=8000, temperature=0)
        by_line: dict[int, set[str]] = {}
        for assignment in result.get("assignments", []):
            line = assignment.get("line")
            if isinstance(line, int):
                by_line.setdefault(line, set()).add(
                    str(assignment.get("fid", "")))
        for line, (label, section, _values) in enumerate(details, 1):
            candidates = by_line.get(line, set())
            suggested = next(iter(candidates)) if len(candidates) == 1 else ""
            if suggested not in valid_fids:
                suggested = ""
            proposals.append({
                "statement": stmt,
                "scope": scope,
                "report_line": label,
                "location": section,
                "suggested_fid": suggested,
                "suggested_name": (
                    byfid[suggested].name if suggested in byfid else ""),
                "status": "unreviewed",
                "authority": "proposal_only",
            })
    return {
        "schema_version": 1,
        "authority": "proposal_only",
        "authoritative_report_affected": False,
        "proposals": proposals,
    }


def write_mapping_proposals(path: str, payload: dict) -> None:
    """Atomically write the advisory mapping-review sidecar."""
    import json
    import os
    import uuid

    os.makedirs(os.path.dirname(path), exist_ok=True)
    temporary = path + f".{uuid.uuid4().hex}.tmp"
    with open(temporary, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, sort_keys=True, ensure_ascii=False)
        handle.write("\n")
    os.replace(temporary, path)


# --------------------------------------------------------------------------- client workbook output

def _period_label(p: Period) -> str:
    lab = f"{p.span} {p.end or p.raw}"
    if p.audited:
        lab += f" ({p.audited})"
    return lab


_SPAN_MONTHS = {"3M": 3, "6M": 6, "9M": 9, "FY": 12}


def attach_provenance(ms: "MappedStatement", grid) -> None:
    """Rebuild per-period (report line, value) provenance for every mapped fid
    and every unmapped line from the raw grid — deterministic."""
    from collections import defaultdict
    num_mode = detect_number_format(grid)
    rows_by_label = defaultdict(list)
    for row in grid:
        label, vals = _label_and_vals(row, num_mode)
        if label and any(c.isalpha() for c in label):
            rows_by_label[norm_label(label)].append((label, vals))
    if not getattr(ms, "unmapped_vals", None):
        uv = []
        for lab in ms.unmapped:
            for orig, vals in rows_by_label.get(norm_label(lab), []):
                uv.append((orig, dict(vals)))
                break
        ms.unmapped_vals = uv
    if getattr(ms, "sources_vals", None):
        return                                  # exact provenance already recorded
    out: dict = {}
    for fid, labels in ms.sources.items():
        per: dict = defaultdict(list)
        seen = set()
        for lab in labels:
            nl = norm_label(lab)
            if nl in seen:
                continue
            seen.add(nl)
            for orig, vals in rows_by_label.get(nl, []):
                for col, v in vals.items():
                    per[col].append((orig, v))
        # duplicate-label rows can over-list: keep only a combination whose sum
        # matches the mapped value (try: all, keep-one, drop-one, keep-a-pair);
        # if nothing reconciles, say so rather than mislead
        from itertools import combinations
        for col, items in list(per.items()):
            fact = ms.facts.get(fid, {}).get(col)
            if fact is None or abs(sum(v for _l, v in items) - fact) <= 0.01:
                continue
            candidates = ([[it] for it in items]
                          + [items[:k] + items[k + 1:] for k in range(len(items))]
                          + ([list(c) for c in combinations(items, 2)] if len(items) <= 6 else []))
            for cand in candidates:
                if cand and abs(sum(v for _l, v in cand) - fact) <= 0.01:
                    per[col] = cand
                    break
            else:
                per[col] = [("(same-named report lines; exact split ambiguous)", fact)]
        out[fid] = dict(per)
    ms.sources_vals = out


def reconcile_cashflow_opening(ms: "MappedStatement",
                               template_fields: list[ClientField]) -> int:
    """Closing-identity completion for cash flows: if closing != beginning +
    net(+fx) and a subset of unmapped lines sums EXACTLY to the residual in
    every period, fold those lines into the beginning-cash field (they are
    printed opening-balance adjustments like 'Cash acquired on acquisition' or
    'Less: Bank overdraft'). Deterministic and arithmetic-gated. Returns the
    number of lines folded; folded rows are flagged 'adjusted' in the output."""
    from itertools import combinations

    def find(*pats):
        for f in template_fields:
            n = f.name.lower()
            if all(re.search(p, n) for p in pats):
                return f.fid
        return None

    BEG = find(r"cash and cash equivalents at the beginning")
    END = find(r"cash and cash equivalents at the end")
    FX = find(r"effect of exchange fluctuation")
    NET = find(r"net increase.*cash and cash equivalents")
    ADJ = find(r"^other adjustments$")
    f = ms.facts
    if not BEG or BEG not in f or not END or END not in f or not ADJ:
        return 0
    cols = sorted(set(f[BEG]) & set(f[END]))
    net = f.get(NET) or {}
    fx = f.get(FX) or {}
    resid = {}
    for c in cols:
        n = net.get(c)
        if n is None:
            continue
        resid[c] = round(f[END][c] - (f[BEG][c] + n + fx.get(c, 0.0)
                                      + f.get(ADJ, {}).get(c, 0.0)), 2)
    if not resid or all(abs(v) <= 0.02 for v in resid.values()):
        return 0
    uv = getattr(ms, "unmapped_vals", None) or []
    # candidates must come from the reconciliation zone: opening-balance
    # adjustments are printed BEFORE the closing-cash line. Lines from the
    # 'Components of cash and cash equivalents' breakdown (below closing) are
    # parts of the closing balance, never adjustments — without this cut the
    # subset search can grab one to absorb the filing's own rounding noise.
    end_zone = re.compile(r"cash and cash equivalents at the end|"
                          r"components of cash", re.IGNORECASE)
    zone_end = next((i for i, (lab, _v) in enumerate(uv) if end_zone.search(lab)),
                    len(uv))
    cand = [(i, lab, vals) for i, (lab, vals) in enumerate(uv[:zone_end])
            if any(c in vals for c in resid)]
    for size in range(1, min(4, len(cand)) + 1):
        for combo in combinations(cand, size):
            ok = all(abs(sum(vals.get(c, 0.0) for _i, _l, vals in combo) - r) <= max(0.02, 0.002 * abs(f[END][c]))
                     for c, r in resid.items())
            if not ok:
                continue
            sv = getattr(ms, "sources_vals", None) or {}
            svb = sv.setdefault(ADJ, {})
            for i, lab, vals in combo:
                for c, v in vals.items():
                    f.setdefault(ADJ, {})[c] = round(f.get(ADJ, {}).get(c, 0.0) + v, 2)
                    svb.setdefault(c, []).append((lab + " [adjusted]", v))
                ms.sources.setdefault(ADJ, []).append(lab)
                ms.verification.append(
                    f"~ reconciliation [{ADJ} Other Adjustments]: '{lab}' folded so that "
                    "beginning + net (+fx) + adjustments = closing ✓")
            drop = {i for i, _l, _v in combo}
            ms.unmapped_vals = [x for i, x in enumerate(uv) if i not in drop]
            ms.unmapped = [l for l in ms.unmapped
                           if l not in {lab for _i, lab, _v in combo}]
            return size
    return 0


def classify_unmapped(ms: "MappedStatement",
                      template_fields: list[ClientField]) -> list[tuple[str, str]]:
    """Deterministically classify every unmapped line by its ARITHMETIC
    relationship to the mapped fields — nothing is guessed:

      aggregate : the line equals a +/- combination of mapped fields (verified)
      component : the line belongs to a consecutive group that sums exactly to
                  a mapped field's value (verified — already included there)
      unresolved: neither — no reviewed taxonomy rule matched; a model may
                  propose a candidate in the separate review sidecar

    Returns [(class, comment)] aligned with ms.unmapped_vals."""
    from itertools import combinations
    names = {f.fid: f.name for f in template_fields}
    uv = getattr(ms, "unmapped_vals", None) or []
    facts = ms.facts

    def vec(d, cols):
        return tuple(round(d.get(c, 0.0), 2) for c in cols)

    def close(a, b):
        return all(abs(x - y) <= max(0.02, 0.005 * abs(y)) for x, y in zip(a, b))

    results: list = [None] * len(uv)

    # --- component groups: consecutive unmapped lines summing to a mapped fid
    for size in range(len(uv), 1, -1):
        for start in range(0, len(uv) - size + 1):
            idxs = list(range(start, start + size))
            if any(results[i] for i in idxs):
                continue
            cols = sorted({c for i in idxs for c in uv[i][1]})
            if not cols:
                continue
            total = {c: sum(uv[i][1].get(c, 0.0) for i in idxs) for c in cols}
            for fid, fv in facts.items():
                if all(c in fv for c in cols) and close(vec(total, cols), vec(fv, cols)):
                    for i in idxs:
                        results[i] = ("component",
                                      f"Already included in '{names.get(fid, fid)}' [{fid}] — "
                                      f"the group of {size} lines sums exactly to its value ✓")
                    break

    # --- aggregates: line = +/- combination of mapped fields
    fids = [f for f in facts if facts[f]]
    for i, (lab, vals) in enumerate(uv):
        if results[i] or not vals:
            continue
        cols = sorted(vals)
        cand = [f for f in fids if all(c in facts[f] for c in cols)]
        tv = vec(vals, cols)
        hit = None
        for f in cand:
            if close(vec(facts[f], cols), tv):
                hit = f"equals '{names.get(f, f)}' [{f}] ✓"
                break
        if hit is None:                        # sums first (clearer than differences)
            for f1, f2 in combinations(cand, 2):
                s = tuple(round(facts[f1][c] + facts[f2][c], 2) for c in cols)
                if close(s, tv):
                    hit = (f"= '{names.get(f1, f1)}' [{f1}] + '{names.get(f2, f2)}' [{f2}] ✓")
                    break
        if hit is None:
            for f1, f2 in combinations(cand, 2):
                d1 = tuple(round(facts[f1][c] - facts[f2][c], 2) for c in cols)
                d2 = tuple(round(facts[f2][c] - facts[f1][c], 2) for c in cols)
                if close(d1, tv):
                    hit = (f"= '{names.get(f1, f1)}' [{f1}] − '{names.get(f2, f2)}' [{f2}] ✓")
                elif close(d2, tv):
                    hit = (f"= '{names.get(f2, f2)}' [{f2}] − '{names.get(f1, f1)}' [{f1}] ✓")
                if hit:
                    break
        if hit:
            results[i] = ("aggregate",
                          f"No dedicated client field, but fully represented: {hit} "
                          "(shown for verification)")
    for i in range(len(uv)):
        if not results[i]:
            results[i] = (
                "unresolved",
                "No reviewed taxonomy rule matched. The value was not written "
                "to an authoritative field; any model suggestion is retained "
                "only in the separate proposal review sidecar.")
    return results


# --------------------------------------------------------------------------- shared output helpers

_UNIT_WORDS = {"lac": "lakhs", "lakh": "lakhs", "crore": "crores", "cr": "crores",
          "million": "millions", "mn": "millions", "mio": "millions",
          "billion": "billions", "bn": "billions", "thousand": "thousands"}


def filing_unit(pdf_path: str) -> str:
    """The filing's printed denomination ('(Rs. in lakhs)', '(₹ crore)').
    Anchored on the parenthetical unit text, which survives dirty OCR even
    when the currency symbol does not."""
    import collections
    import re
    import pymupdf
    votes = collections.Counter()
    doc = pymupdf.open(pdf_path)
    for page in doc:
        t = " ".join(page.get_text().split())
        if len(re.findall(r"\d[\d,]{2,}", t)) < 15:
            continue                              # statement pages only
        for m in re.finditer(r"\(([^()]{0,40})\)", t):
            span = m.group(1).lower()
            if "$" in span or "usd" in span:
                continue                          # foreign-currency note, not the denomination
            w = re.search(r"\b(lakh|lac|crore|cr|million|mn|mio|billion|bn|thousand)s?\b", span)
            if w is None:
                continue
            pre = span[:w.start()]
            # a denomination reads '(In ₹ Million)' / '(Rs. in lakhs)': the unit
            # follows 'in' or a currency marker — never a QUANTITY, as in the
            # share-count note '(Five Hundred Fifty Thousand only)'
            if re.search(r"(\d|one|two|three|four|five|six|seven|eight|nine|ten|"
                         r"twenty|thirty|forty|fifty|sixty|seventy|eighty|ninety|"
                         r"hundred|thousand|lakh|crore|million)\s*$", pre):
                continue
            if re.search(r"[a-z]{3,}", pre) and not re.search(r"\b(in|rs|inr|rupee|rupees|amount|amounts|figures|values)\b", pre):
                continue                          # words before the unit, none of them a currency marker
            votes[_UNIT_WORDS[w.group(1)]] += 1
        # unparenthesised header form '₹ in million' — custom fonts extract the
        # ₹ glyph as junk ('~in million', 'tin million'), so anchor on a currency
        # marker OR a single mangled glyph directly before 'in <unit>'
        for m in re.finditer(r"(?:\b(?:rs\.?|inr|rupees)\s*|[~t§`¹])in\s+"
                             r"(lakh|lac|crore|cr|million|mn|mio|billion|bn|thousand)s?\b",
                             t, re.I):
            votes[_UNIT_WORDS[m.group(1).lower()]] += 1
    doc.close()
    return votes.most_common(1)[0][0] if votes else ""







def company_unit(pdf_path: str, pages: list[int] | None = None) -> str:
    """filing_unit, with a sibling fallback: scanned filings may carry the
    denomination only as pixels ('In ₹ Million' on an image page). A company
    reports in ONE unit, so the same company's other filings in the folder
    decide when this filing's text layer has no vote."""
    u = filing_unit(pdf_path)
    if u:
        return u
    import collections
    import glob
    import os
    base = os.path.basename(pdf_path)
    if "_q" not in base:
        return ""
    comp = base.split("_q")[0]
    votes = collections.Counter()
    for f in glob.glob(os.path.join(os.path.dirname(pdf_path), comp + "_q*.pdf")):
        if os.path.abspath(f) == os.path.abspath(pdf_path):
            continue
        u2 = filing_unit(f)
        if u2:
            votes[u2] += 1
    if votes:
        return votes.most_common(1)[0][0]
    # every sibling is a scan too (no text layer anywhere): read the printed
    # denomination from the page pixels
    return unit_from_pixels(pdf_path, pages)


def unit_from_pixels(pdf_path: str, pages: list[int] | None = None) -> str:
    """Last-resort denomination read for fully scanned filings: the header
    strip ('Rs. in Mn') exists only as pixels, so ask for it optically.
    One tiny call (~$0.01), top strips of a few pages. `pages` (1-based)
    should be the known statement pages when the caller has them."""
    import base64

    import pymupdf

    from src import llm
    doc = pymupdf.open(pdf_path)
    if pages:
        cand = [p - 1 for p in pages if 0 < p <= len(doc)]
    else:
        # scanned pages first (the denomination lives in the image); else the
        # unit is a graphic on an otherwise-text page — use number-heavy pages
        cand = [i for i in range(min(10, len(doc)))
                if len(doc[i].get_text().strip()) < 100]
        if not cand:
            cand = [i for i in range(min(10, len(doc)))
                    if len(re.findall(r"\d[\d,]{2,}", doc[i].get_text())) >= 15]
    imgs = []
    for i in cand[:6]:
        r = doc[i].rect
        pix = doc[i].get_pixmap(dpi=150, clip=pymupdf.Rect(r.x0, r.y0, r.x1,
                                                           r.y0 + r.height * 0.25))
        imgs.append(base64.b64encode(pix.tobytes("png")).decode())
    doc.close()
    if not imgs:
        return ""
    try:
        out = llm.extract_json(
            instructions=("Each image is the top strip of a financial-results page. "
                          "Some strips may be cover/auditor pages — look for the one "
                          "with a results-table heading. Find the money denomination "
                          "printed near the table "
                          "(e.g. 'Rs. in Mn' means millions, '₹ in lakhs', "
                          "'Rs. crores'). Answer with the unit word only."),
            user_input="What denomination are the amounts stated in?",
            schema_name="denomination",
            schema={"type": "object",
                    "properties": {"unit": {"type": "string",
                                            "enum": ["lakhs", "crores", "millions",
                                                     "billions", "thousands", "unknown"]}},
                    "required": ["unit"], "additionalProperties": False},
            images_b64=imgs, max_output_tokens=200)
        u = out.get("unit", "")
        return "" if u == "unknown" else u
    except Exception:
        return ""


def _sized_comment(text: str, author: str):
    """An Excel hover note SIZED TO ITS TEXT — the default 2x1cm box truncates
    anything longer than a few words."""
    import math

    from openpyxl.comments import Comment
    width = 520                                   # px
    # Estimate wrapped lines conservatively (~44 chars/line — Excel wraps tighter
    # than the raw char count) and give the box plenty of height so long
    # calculation notes are shown in full rather than clipped.
    lines = sum(max(1, math.ceil(len(ln) / 44)) for ln in text.splitlines() or [""])
    height = max(110, 26 + lines * 18)
    return Comment(text, author, height=height, width=width)


def _review_sheet(wb):
    """The workbook's Review sheet — created at index 0 on first use, reused
    afterwards so extraction flags, scan suspects and post-write failures all
    land on ONE sheet."""
    from openpyxl.styles import Font
    if "Review" in wb.sheetnames:
        return wb["Review"]
    rv = wb.create_sheet("Review", 0)
    rv.sheet_properties.tabColor = "ED8B00"
    rv.append(["Items needing manual review — these could not be fully verified "
               "automatically. Please check them against the source PDF."])
    rv["A1"].font = Font(bold=True)
    rv.append([])
    rv.append(["Statement", "Scope", "Item", "Period column",
               "Read 1", "Read 2", "Source page"])
    for c in rv[3]:
        c.font = Font(bold=True)
    for col, w in zip("ABCDEFG", (22, 14, 80, 40, 12, 12, 11)):
        rv.column_dimensions[col].width = w
    return rv


def _review_append(rv, row) -> None:
    """Append to the Review sheet unless an identical item is already listed
    (the mapped-layer flag and the post-write finding for the same defect
    must not appear twice)."""
    key = (str(row[0]), str(row[1]), str(row[2])[:120])
    for r in rv.iter_rows(min_row=4, max_col=3, values_only=True):
        if (str(r[0]), str(r[1]), str(r[2] or "")[:120]) == key:
            return
    _ws_append(rv, row)


def annotate_flags(wide_path: str, mapped: dict) -> int:
    """Stamp every MappedStatement flag ONTO the wide deliverable: orange tab
    + header note on the statement sheet, one row per flag on the Review
    sheet. Returns the number of flags written. This is what makes a flag
    impossible to drop — the file the client opens carries it."""
    import openpyxl
    STMT_NAME = {"income": "Income Statement", "balance": "Balance Sheet",
                 "cashflow": "Cash Flow", "segment": "Segment Finance"}
    items = [(stmt, scope, f) for (stmt, scope), ms in mapped.items()
             for f in (getattr(ms, "flags", None) or [])]
    if not items:
        return 0
    wb = openpyxl.load_workbook(wide_path)
    rv = _review_sheet(wb)
    for stmt, scope, f in items:
        sn = f"{STMT_NAME.get(stmt, stmt)} - {str(scope).title()}"[:31]
        if sn in wb.sheetnames:
            ws = wb[sn]
            ws.sheet_properties.tabColor = "ED8B00"
            c = ws.cell(row=1, column=1)
            note = f"⚠ {f}"
            prev = c.comment.text if c.comment else ""
            if note not in prev:
                c.comment = _sized_comment((prev + "\n" if prev else "") + note,
                                           "Reports Radar")
        _review_append(rv, [STMT_NAME.get(stmt, stmt), str(scope).title(),
                            f"⚠ {f}", "", "", "", ""])
    wb.save(wide_path)
    return len(items)


def annotate_review(wide_path: str, suspects: list[dict], failing: list[dict]) -> None:
    """Surface manual-verification items IN the deliverable. Offline.

    suspects: cell-level items {stmt, scope, label, col, v1, v2, page} —
      two independent reads of a scanned page disagreed and no printed total
      settles it. Matching workbook cells get an orange fill + a comment; all
      items are listed on a 'Review' sheet with page references.
    failing: statement-level items {stmt, scope, title, page} — the printed
      totals of that statement could not be reconciled automatically.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    if not suspects and not failing:
        return
    STMT_NAME = {"income": "Income Statement", "balance": "Balance Sheet",
                 "cashflow": "Cash Flow", "segment": "Segment Finance"}
    _n = lambda s: " ".join(re.sub(r"[^a-z0-9 ]", " ", str(s).lower()).split())
    fill = PatternFill("solid", start_color="FFE0B2")
    wb = openpyxl.load_workbook(wide_path)
    for s in suspects:
        sn = f"{STMT_NAME.get(s['stmt'], s['stmt'])} - {str(s['scope']).title()}"[:31]
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        p = _parse_period(s["col"], 0)
        span = "12M" if p.span == "FY" else p.span   # wide headers print months
        heads = [(j, str(c.value or "")) for j, c in enumerate(ws[1], 1)]
        cols = [j for j, h in heads
                if p.end and p.end in h and (span == "?" or f"({span})" in h)]
        if len(cols) != 1:
            continue                              # only mark what we can place exactly
        lab = _n(s["label"])

        def _line_match(text):
            # sub-items read 'I Revenue from operations = 17,721' per line —
            # compare LINE LABELS, not substrings ('Total' must not light up
            # every row that contains the word 'total')
            for line in str(text or "").splitlines():
                ll = _n(line.rsplit("=", 1)[0] if "=" in line else line)
                if not ll:
                    continue
                if ll == lab:
                    return True
                lo, hi = sorted((ll, lab), key=len)
                if lo and lo in hi and len(lo) >= 0.6 * len(hi):
                    return True
            return False

        for row in ws.iter_rows(min_row=2):
            cell = row[cols[0] - 1]
            cmt = cell.comment.text if cell.comment else ""
            if lab and (_line_match(row[-1].value) or _line_match(cmt)):
                cell.fill = fill
                note = (f"Scanned source is unclear here — please verify against "
                        f"the filing (page {s['page']}). Independent reads: "
                        f"{s['v1']} / {s['v2']}.")
                cell.comment = _sized_comment((cmt + "\n\n" if cmt else "") + note,
                                              "Reports Radar")
    rv = _review_sheet(wb)
    for s in suspects:
        _ws_append(rv, [STMT_NAME.get(s["stmt"], s["stmt"]), str(s["scope"]).title(),
                        s["label"], s["col"], s["v1"], s["v2"], s["page"]])
    for f in failing:
        _review_append(rv, [STMT_NAME.get(f["stmt"], f["stmt"]), str(f["scope"]).title(),
                            "⚠ (whole statement) — values could not be fully verified; "
                            "please check this statement against the source",
                            "", "", "", f.get("page", "")])
        sn = f"{STMT_NAME.get(f['stmt'], f['stmt'])} - {str(f['scope']).title()}"[:31]
        if sn in wb.sheetnames:
            wb[sn].sheet_properties.tabColor = "ED8B00"
    wb.save(wide_path)


def _ws_append(sheet, row):
    """Append a row, stripping characters Excel/openpyxl refuse — control
    characters that leak in from a filing's text layer (e.g. a mis-decoded ₹
    glyph), which otherwise raise IllegalCharacterError."""
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    sheet.append([ILLEGAL_CHARACTERS_RE.sub("", c) if isinstance(c, str) else c
                  for c in row])


def to_wide(long_path: str, wide_path: str) -> None:
    """Pivot a long-format workbook (one row per field x period) into the
    DEFAULT wide format: one row per field, one column per period.
    Deterministic — works offline from the long file alone."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    src = openpyxl.load_workbook(long_path, read_only=True)
    out = openpyxl.Workbook()
    out.remove(out.active)
    for sn in src.sheetnames:
        rows = list(src[sn].iter_rows(values_only=True))
        ws = out.create_sheet(sn)
        if sn == "Audit" or not rows:
            for r in rows:
                ws.append(list(r))
            for c in ws[1]:
                c.font = Font(bold=True)
            if sn == "Audit":
                from openpyxl.styles import Alignment
                for j, w in zip(range(1, 10), (12, 12, 10, 44, 10, 70, 12, 8, 110)):
                    ws.column_dimensions[get_column_letter(j)].width = w
                for r_ in ws.iter_rows(min_row=2, min_col=9, max_col=9):   # wrap Verification
                    for c in r_:
                        c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.freeze_panes = "A2"
            continue
        if sn == "Unmapped":
            key_cols, val_i = (0, 1, 2, 6, 7), 5      # stmt, scope, label, denom, reason
            per_i, mon_i, aud_i = 3, 4, None
            lead = ["Statement", "Scope", "Report line (as printed)"]
            tail = ["Denomination", "Reason"]
        else:
            key_cols, val_i = (0, 1, 6, 7, 8), 5      # fid, name, denom, cur, method
            per_i, mon_i, aud_i = 2, 3, 4
            lead = ["Field id", "Display Name"]
            tail = ["Denomination", "Currency", "Method",
                    "Sub-items (report lines / calculation; latest period)"]
        pers, fields, data = [], [], {}
        for r in rows[1:]:
            if r is None or r[0] is None:
                continue
            months = str(r[mon_i]) if r[mon_i] not in (None, "") else ""
            aud = str(r[aud_i]).strip() if aud_i is not None and r[aud_i] else ""
            head = str(r[per_i] or "")
            head += f" ({months}M)" if months else " (as at)"
            if aud:
                head += f" [{aud}]"
            if head not in pers:
                pers.append(head)
            k = tuple(r[i] for i in key_cols)
            if k not in data:
                data[k] = {"_sub": r[9] if sn != "Unmapped" and len(r) > 9 else "",
                           "_subp": {}}
                fields.append(k)
            data[k].setdefault(head, r[val_i])
            if sn != "Unmapped" and len(r) > 9 and r[9]:
                data[k]["_subp"].setdefault(head, str(r[9]))
        hdr = lead + pers + tail
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True)
        for k in fields:
            if sn == "Unmapped":
                stmt, scope, lab, denom, reason = k
                row = [stmt, scope, lab] + [data[k].get(p, "") for p in pers] + [denom, reason]
            else:
                fid, name, denom, cur, method = k
                row = ([fid, name] + [data[k].get(p, "") for p in pers]
                       + [denom, cur, method, data[k]["_sub"]])
            ws.append(row)
            if sn != "Unmapped":
                # every period cell carries ITS OWN calculation as a hover note
                ri = ws.max_row
                for j, p in enumerate(pers, len(lead) + 1):
                    sub = data[k]["_subp"].get(p)
                    if sub:
                        c = _sized_comment(f"{p}:\n{sub}", "engine")
                        ws.cell(row=ri, column=j).comment = c
        widths = [10, 48] if sn != "Unmapped" else [16, 12, 52]
        widths += [15] * len(pers) + ([13, 9, 10, 80] if sn != "Unmapped" else [13, 72])
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "C2" if sn != "Unmapped" else "D2"
        ws.auto_filter.ref = ws.dimensions
    out.save(wide_path)


def canonicalize_xlsx(path: str) -> None:
    """Normalize non-semantic XLSX metadata for byte-for-byte reproducibility.

    Excel workbooks are ZIP archives. openpyxl writes the current time into
    both ZIP members and core.xml on every save, making identical financial
    data hash differently. This changes metadata only, never workbook cells.
    """
    import os
    import zipfile

    fixed_time = b"2000-01-01T00:00:00Z"
    temp = path + ".canonical.tmp"
    with zipfile.ZipFile(path, "r") as source:
        with zipfile.ZipFile(temp, "w", compression=zipfile.ZIP_DEFLATED,
                             compresslevel=9) as target:
            target.comment = source.comment
            for name in sorted(source.namelist()):
                data = source.read(name)
                if name == "docProps/core.xml":
                    data = re.sub(
                        rb"(<dcterms:created\b[^>]*>)[^<]*(</dcterms:created>)",
                        rb"\g<1>" + fixed_time + rb"\g<2>", data)
                    data = re.sub(
                        rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                        rb"\g<1>" + fixed_time + rb"\g<2>", data)
                original = source.getinfo(name)
                info = zipfile.ZipInfo(name, (2000, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = original.create_system
                info.external_attr = original.external_attr
                info.internal_attr = original.internal_attr
                info.flag_bits = original.flag_bits
                target.writestr(info, data, compress_type=zipfile.ZIP_DEFLATED,
                                compresslevel=9)
    os.replace(temp, path)


def write_client_workbook_long(company: str, mapped: dict[tuple[str, str], "MappedStatement"],
                               template: dict[tuple[str, str], list[ClientField]],
                               out_path: str) -> None:
    """LONG-format workbook: one sheet per statement x scope (e.g. 'Income
    Statement - Standalone'), one row per field x period, with Period End /
    Months / Denomination columns, plus the full Audit sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    STMT_NAME = {"income": "Income Statement", "balance": "Balance Sheet",
                 "cashflow": "Cash Flow", "segment": "Segment Finance"}
    # per-share figures are rupees, never the statement denomination
    # (filings print 'Rs. in lakhs EXCEPT EPS')
    _EPS = re.compile(r"\bEPS\b|earnings per (equity )?share", re.IGNORECASE)
    STMT_ORDER = {"income": 0, "balance": 1, "cashflow": 2, "segment": 3}
    HDR = ["Field id", "Display Name", "Period End", "Months", "Audited",
           "Value", "Denomination", "Currency", "Method", "Sub-items (report lines / calculation)"]
    wb = Workbook()
    audit = wb.active
    audit.title = "Audit"
    audit.append(["Statement", "Scope", "Field id", "Display Name", "Method",
                  "Source report lines", "Denomination", "Currency", "Verification"])
    for c in audit[1]:
        c.font = Font(bold=True)

    keys = sorted(mapped, key=lambda k: (STMT_ORDER.get(k[0], 9), k[1] != "standalone"))
    for (stmt, scope) in keys:
        ms = mapped[(stmt, scope)]
        fields = template.get((stmt, scope))
        if not fields:
            continue
        ws = wb.create_sheet(f"{STMT_NAME.get(stmt, stmt)} - {scope.title()}"[:31])
        ws.append(HDR)
        for c in ws[1]:
            c.font = Font(bold=True)
        denom = ms.unit or "units"
        row2fid = {f.row: f.fid for f in fields}
        byfid = {f.fid: f for f in fields}
        prov = getattr(ms, "sources_vals", None) or {}

        _val_cache: dict = {}

        def field_value(fid, col, stack=()):
            """Resolve a field's value at a column: a reported fact if present,
            else its formula/ratio evaluated RECURSIVELY over the same resolver
            (so nested subtotals like Total Income from Operations → Net Sales →
            Gross Sales resolve). Memoised per (fid, col); cycle-safe. An additive
            formula emits when at least one term resolves (missing terms = 0), so
            single-child rollups (e.g. Total Other Income) are not suppressed."""
            if not fid or fid in stack:
                return None
            key = (fid, col)
            if key in _val_cache:
                return _val_cache[key]
            g = byfid.get(fid)
            if fid in ms.facts and col in ms.facts[fid]:
                v = ms.facts[fid][col]
            elif g and g.formula:
                tot, have = 0.0, 0
                for sg, rr in g.formula:
                    cv = field_value(row2fid.get(rr), col, stack + (fid,))
                    if cv is not None:
                        tot += sg * cv
                        have += 1
                v = round(tot, 2) if have >= 1 else None
            elif g and g.ratio:
                nr, dr, sc = g.ratio
                nv = field_value(row2fid.get(nr), col, stack + (fid,))
                dv = field_value(row2fid.get(dr), col, stack + (fid,))
                v = round(nv / dv * sc, 2) if (nv is not None and dv not in (None, 0)) else None
            else:
                v = None
            _val_cache[key] = v
            return v

        for f in sorted(fields, key=lambda x: x.order):
            per_vals = {}
            per_sub = {}
            method = ""
            if f.fid in ms.facts:
                method = "reported" if len(ms.sources.get(f.fid, [])) == 1 else "summed"
                if any("[adjusted]" in lab for items in prov.get(f.fid, {}).values()
                       for lab, _v in items):
                    method = "adjusted"
                for p in ms.periods:
                    if p.col in ms.facts[f.fid]:
                        per_vals[p.col] = ms.facts[f.fid][p.col]
                        items = prov.get(f.fid, {}).get(p.col, [])
                        per_sub[p.col] = "  +  ".join(f"{lab} = {v:,.2f}".rstrip("0").rstrip(".")
                                                      for lab, v in items)
            elif f.formula:
                for p in ms.periods:
                    total, have = 0.0, 0
                    parts = []
                    for sign, r in f.formula:
                        cf = row2fid.get(r)
                        cv = field_value(cf, p.col, (f.fid,))
                        if cv is not None:
                            total += sign * cv
                            have += 1
                            parts.append(f"{'+' if sign > 0 else '-'} {byfid[cf].name} "
                                         f"[{cf}] = {cv:,.2f}".rstrip("0").rstrip("."))
                    if have >= 1:
                        per_vals[p.col] = round(total, 2)
                        per_sub[p.col] = "  ".join(parts)
                        method = "computed"
            elif f.ratio:
                num_r, den_r, scale = f.ratio
                nfid, dfid = row2fid.get(num_r), row2fid.get(den_r)
                for p in ms.periods:
                    nv = field_value(nfid, p.col, (f.fid,))
                    dv = field_value(dfid, p.col, (f.fid,))
                    if nv is not None and dv not in (None, 0):
                        per_vals[p.col] = round(nv / dv * scale, 2)
                        nnm = byfid[nfid].name if nfid in byfid else f"C{num_r}"
                        dnm = byfid[dfid].name if dfid in byfid else f"C{den_r}"
                        per_sub[p.col] = (f"{nnm} [{nfid}] = {nv:,.2f}".rstrip("0").rstrip(".")
                                          + f"  ÷  {dnm} [{dfid}] = {dv:,.2f}".rstrip("0").rstrip(".")
                                          + f"  × {scale:g}")
                        method = "computed"
            if not per_vals:
                continue
            for p in ms.periods:
                if p.col not in per_vals:
                    continue
                # balance sheets are point-in-time ('As at ...'): no Months
                months = "" if stmt == "balance" else _SPAN_MONTHS.get(p.span, "")
                _ws_append(ws, [f.fid, f.name, p.end or p.raw,
                                months, p.audited,
                                per_vals[p.col],
                                "rupees" if _EPS.search(f.name) else denom,
                                ms.currency, method,
                                per_sub.get(p.col, "")])
            _ws_append(audit, [stmt, scope, f.fid, f.name, method,
                               "; ".join(ms.sources.get(f.fid, [])),
                               denom, ms.currency,
                               "\n".join(v for v in ms.verification if f"[{f.fid}]" in v)[:4000]])
        if ms.unmapped:
            _ws_append(audit, [stmt, scope, "", "UNMAPPED LINES", "",
                               "; ".join(ms.unmapped)[:300], "", "", ""])
        # statement-level review flags: visible on the sheet itself (orange
        # tab + header note), in the Audit sheet, and later on the Review
        # sheet — a flagged statement can never look verified
        flags = getattr(ms, "flags", None) or []
        if flags:
            ws.sheet_properties.tabColor = "ED8B00"
            hdr_cell = ws.cell(row=1, column=1)
            hdr_cell.value = "⚠ Field id"
            hdr_cell.comment = _sized_comment(
                "This statement needs review:\n" + "\n".join(f"• {f}" for f in flags),
                "Reports Radar")
            _ws_append(audit, [stmt, scope, "", "⚠ REVIEW FLAGS", "",
                               " | ".join(flags)[:2000], "", "", ""])
        for j, w in zip(range(1, 11), (10, 48, 12, 8, 11, 16, 13, 9, 10, 80)):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    # dedicated sheet for report lines with NO client field — visible, with values
    un = wb.create_sheet("Unmapped")
    un.append(["Statement", "Scope", "Report line (as printed)", "Period End", "Months",
               "Value", "Denomination", "Reason"])
    for c in un[1]:
        c.font = Font(bold=True)
    # filings routinely reprint the same line (IFRS mirror, revenue block +
    # segment-revenues block): one identical (line, period, value) fact is
    # listed ONCE, keeping the most informative reason (verified ✓ over info)
    entries: dict = {}
    order: list = []
    RANK = {"component": 0, "aggregate": 0, "unresolved": 1}
    for (stmt, scope) in keys:
        ms = mapped[(stmt, scope)]
        denom = ms.unit or "units"
        col2p = {p.col: p for p in ms.periods}
        klass = classify_unmapped(ms, template.get((stmt, scope), []))
        for k_i, (lab, vals) in enumerate(getattr(ms, "unmapped_vals", None) or []):
            cls, reason = klass[k_i]
            if stmt == "segment" and cls == "component":
                reason += " (per-segment member; template has no per-segment fields)"
            for col, v in vals.items():
                p = col2p.get(col)
                if p is None:
                    continue
                months = "" if stmt == "balance" else _SPAN_MONTHS.get(p.span, "")
                key = (stmt, scope, lab, p.end or p.raw, months, v)
                if key not in entries:
                    order.append(key)
                    entries[key] = (RANK.get(cls, 1), reason, denom)
                elif RANK.get(cls, 1) < entries[key][0]:
                    entries[key] = (RANK.get(cls, 1), reason, denom)
    for (stmt, scope, lab, pend, months, v) in order:
        _rk, reason, denom = entries[(stmt, scope, lab, pend, months, v)]
        _ws_append(un, [STMT_NAME.get(stmt, stmt), scope.title(), lab,
                        pend, months, v, denom, reason])
    for j, w in zip(range(1, 9), (16, 12, 52, 12, 8, 16, 13, 72)):
        un.column_dimensions[get_column_letter(j)].width = w
    un.freeze_panes = "A2"
    un.auto_filter.ref = un.dimensions

    from openpyxl.styles import Alignment
    for j, w in zip(range(1, 10), (12, 12, 10, 44, 10, 70, 12, 8, 110)):
        audit.column_dimensions[get_column_letter(j)].width = w
    for row in audit.iter_rows(min_row=2, min_col=9, max_col=9):   # wrap the Verification column
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    audit.freeze_panes = "A2"
    wb.save(out_path)


def write_client_workbook(company: str, mapped: dict[tuple[str, str], "MappedStatement"],
                          template: dict[tuple[str, str], list[ClientField]],
                          out_path: str) -> None:
    """client-template-layout workbook: template field order, one column per period,
    reported values where mapped, formula-computed where absent, plus a full
    audit sheet (fid -> source labels -> method -> verification)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    STMT_NAME = {"income": "Income Statement", "balance": "Balance Sheet",
                 "cashflow": "Cash Flow", "segment": "Segment Finance"}
    # per-share figures are rupees, never the statement denomination
    # (filings print 'Rs. in lakhs EXCEPT EPS')
    _EPS = re.compile(r"\bEPS\b|earnings per (equity )?share", re.IGNORECASE)
    wb = Workbook()
    audit = wb.active
    audit.title = "Audit"
    audit.append(["Statement", "Scope", "Field id", "Display Name", "Method",
                  "Source report lines", "Unit", "Currency", "Verification"])
    for c in audit[1]:
        c.font = Font(bold=True)

    for (stmt, scope), ms in mapped.items():
        fields = template.get((stmt, scope))
        if not fields:
            continue
        ws = wb.create_sheet(f"{STMT_NAME.get(stmt, stmt)} - {scope.title()}"[:31])
        hdr = ["Display Name", "Field id"] + [_period_label(p) for p in ms.periods]
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True)
        row2fid = {f.row: f.fid for f in fields}
        byfid = {f.fid: f for f in fields}

        _val_cache: dict = {}

        def field_value(fid, col, stack=()):
            """Recursive, memoised, cycle-safe field resolver (see the long-form
            writer for the rationale): reported fact, else formula/ratio evaluated
            over the same resolver; additive emits when ≥1 term resolves."""
            if not fid or fid in stack:
                return None
            key = (fid, col)
            if key in _val_cache:
                return _val_cache[key]
            g = byfid.get(fid)
            if fid in ms.facts and col in ms.facts[fid]:
                v = ms.facts[fid][col]
            elif g and g.formula:
                tot, have = 0.0, 0
                for sg, rr in g.formula:
                    cv = field_value(row2fid.get(rr), col, stack + (fid,))
                    if cv is not None:
                        tot += sg * cv
                        have += 1
                v = round(tot, 2) if have >= 1 else None
            elif g and g.ratio:
                nr, dr, sc = g.ratio
                nv = field_value(row2fid.get(nr), col, stack + (fid,))
                dv = field_value(row2fid.get(dr), col, stack + (fid,))
                v = round(nv / dv * sc, 2) if (nv is not None and dv not in (None, 0)) else None
            else:
                v = None
            _val_cache[key] = v
            return v

        for f in sorted(fields, key=lambda x: x.order):
            vals = []
            method = ""
            if f.fid in ms.facts:
                method = "reported" if len(ms.sources.get(f.fid, [])) == 1 else "summed"
                if any("[adjusted]" in lab for items in prov.get(f.fid, {}).values()
                       for lab, _v in items):
                    method = "adjusted"
                for p in ms.periods:
                    vals.append(ms.facts[f.fid].get(p.col))
            elif f.formula or f.ratio:
                for p in ms.periods:
                    vals.append(field_value(f.fid, p.col, ()))
                method = "computed" if any(v is not None for v in vals) else ""
            else:
                vals = [None] * len(ms.periods)
            if not any(v is not None for v in vals):
                continue
            ws.append([f.name, f.fid] + vals)
            if method:
                audit.append([stmt, scope, f.fid, f.name, method,
                              "; ".join(ms.sources.get(f.fid, [])),
                              ms.unit, ms.currency,
                              "; ".join(v for v in ms.verification if f"[{f.fid}]" in v)[:180]])
        ws.column_dimensions["A"].width = 52
        for j in range(3, 3 + len(ms.periods)):
            ws.column_dimensions[get_column_letter(j)].width = 20
        ws.freeze_panes = "C2"
        if ms.unmapped:
            audit.append([stmt, scope, "", "UNMAPPED LINES", "", "; ".join(ms.unmapped)[:300], "", "", ""])
    for j, w in zip(range(1, 10), (12, 12, 10, 44, 10, 70, 8, 8, 60)):
        audit.column_dimensions[get_column_letter(j)].width = w
    audit.freeze_panes = "A2"
    wb.save(out_path)
