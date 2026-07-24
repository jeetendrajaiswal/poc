"""Minimal web UI for the extraction pipeline.

Upload an annual report, pick the company (Nifty-50 for now), click Process. The pipeline
runs in a background thread (it makes model calls, so it takes a few minutes), then the
wide-format results (output/results_wide.xlsx) are rebuilt and shown on the page.

  .venv/bin/python -m src.webapp           # http://127.0.0.1:5000
"""
from __future__ import annotations

import csv
import glob
import hashlib
import json
import os
import re
import shutil
import threading
import uuid

from flask import (Flask, request, jsonify, send_file, render_template_string,
                   session, redirect, url_for)
from werkzeug.security import generate_password_hash, check_password_hash

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
os.chdir(ROOT)                                   # so relative output/ paths resolve
UPLOAD_DIR = os.path.join(ROOT, "uploads")
OUT_DIR = os.path.join(ROOT, "output")
os.makedirs(UPLOAD_DIR, exist_ok=True)

TABLES_DIR = os.path.join(OUT_DIR, "tables")
MDNA_DIR = os.path.join(OUT_DIR, "mdna")
CLIENT_DIR = os.path.join(OUT_DIR, "client")
QTR_RAW_DIR = os.path.join(OUT_DIR, "qtr_raw")
for _d in (TABLES_DIR, MDNA_DIR, CLIENT_DIR, QTR_RAW_DIR):
    os.makedirs(_d, exist_ok=True)

# ---------------------------------------------------------------------------
# Persistence — generated outputs and job state survive app-server restarts,
# redeploys and instance replacement. The S3 mirror is active when S3_BUCKET
# is set (it is on Elastic Beanstalk); local development runs disk-only.
# Best-effort by design: a persistence hiccup must never fail a job.
# ---------------------------------------------------------------------------
S3_BUCKET = os.getenv("S3_BUCKET", "")
S3_PREFIX = os.getenv("S3_PREFIX", "data-extraction/output")
JOBS_FILE = os.path.join(OUT_DIR, ".jobs.json")
CANONICAL_REPORT_DIR = os.path.join(OUT_DIR, ".canonical_reports")

# The long-format companion workbook (output/client/long/*_long.xlsx) is PAUSED:
# it is still built in-memory because the wide deliverable is derived from it,
# but it is no longer persisted to disk or mirrored to S3. Set EMIT_LONG=1 to
# re-enable the long deliverable later.
EMIT_LONG = os.getenv("EMIT_LONG", "").lower() in ("1", "true", "yes", "on")

# Re-processing the same filing must not ask probabilistic extractors to make
# the same decisions again.  A content-addressed manifest locks a verified
# report to the exact PDF + extraction code + client configuration that
# produced it.  Any change to any of those inputs invalidates the manifest.
_REPORT_LOCKS: dict[str, threading.Lock] = {}
_REPORT_LOCKS_GUARD = threading.Lock()


def _sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _pipeline_sha256(sector_config) -> str:
    """Fingerprint every local input capable of changing report contents."""
    paths = [sector_config.extraction_path,
             os.path.join(ROOT, "config", "default.yaml"),
             os.path.join(ROOT, "requirements.txt"),
             os.path.join(ROOT, "src", "webapp.py"),
             os.path.join(ROOT, "src", "config.py"),
             os.path.join(ROOT, "src", "llm.py")]
    paths += glob.glob(os.path.join(ROOT, "src", "engine", "*.py"))
    paths += glob.glob(
        os.path.join(
            os.path.dirname(sector_config.taxonomy_path), "**", "*.yaml"),
        recursive=True,
    )
    paths += [
        os.path.join(ROOT, "scripts", name)
        for name in ("repair_raw.py", "verify_raw.py", "verify_delivered.py")
    ]
    h = hashlib.sha256()
    for path in sorted(set(paths)):
        if not os.path.isfile(path):
            continue
        h.update(os.path.relpath(path, ROOT).encode("utf-8"))
        h.update(b"\0")
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                h.update(chunk)
    return h.hexdigest()


def _determinism_manifest(raw_name: str) -> str:
    return os.path.join(CLIENT_DIR, ".cache",
                        f"{raw_name}.determinism.json")


def _report_fingerprint(pdf_path: str, name: str, mode: str,
                        fin_only: bool, vision: bool, sector_config) -> dict:
    from importlib.metadata import PackageNotFoundError, version
    from src.config import config

    package_versions = {}
    for package in ("openai", "PyMuPDF", "openpyxl", "PyYAML"):
        try:
            package_versions[package] = version(package)
        except PackageNotFoundError:
            package_versions[package] = "missing"
    return {
        "version": 2,
        "name": name,
        "sector": sector_config.sector_id,
        "mode": mode,
        "financial_only": bool(fin_only),
        "vision": bool(vision),
        "pdf_sha256": _sha256_file(pdf_path),
        "pipeline_sha256": _pipeline_sha256(sector_config),
        "runtime": {
            "model_default": config.model_default,
            "model_large": config.model_large,
            "reasoning_effort": config.reasoning_effort,
            "self_consistency_n": config.self_consistency_n,
            "double_read": os.getenv("DOUBLE_READ", "1"),
            "packages": package_versions,
        },
    }


def _report_artifacts(name: str, raw_name: str) -> dict[str, str]:
    return {
        "workbook": os.path.join(CLIENT_DIR, f"{name}.xlsx"),
        "mapped": os.path.join(CLIENT_DIR, ".cache", f"{raw_name}.pkl"),
        "raw": os.path.join(QTR_RAW_DIR, f"{raw_name}.pkl"),
    }


def _deterministic_cache_hit(name: str, raw_name: str,
                             fingerprint: dict) -> bool:
    """True only for an intact, review-free artifact set for this exact input."""
    manifest_path = _determinism_manifest(raw_name)
    review = os.path.join(QTR_RAW_DIR, f"{raw_name}.review")
    if os.path.exists(review) or not os.path.exists(manifest_path):
        return False
    try:
        with open(manifest_path, encoding="utf-8") as fh:
            manifest = json.load(fh)
        if manifest.get("fingerprint") != fingerprint:
            return False
        artifacts = _report_artifacts(name, raw_name)
        expected = manifest.get("artifact_sha256", {})
        return all(
            os.path.isfile(path)
            and expected.get(kind) == _sha256_file(path)
            for kind, path in artifacts.items()
        )
    except (OSError, ValueError, TypeError):
        return False


def _canonical_report_path(fingerprint: dict) -> str:
    payload = json.dumps(
        fingerprint, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return os.path.join(
        CANONICAL_REPORT_DIR, hashlib.sha256(payload).hexdigest())


def _restore_canonical_report(name: str, raw_name: str,
                              fingerprint: dict) -> bool:
    """Restore a previously verified result for identical immutable inputs."""
    source_dir = _canonical_report_path(fingerprint)
    manifest_path = os.path.join(source_dir, "manifest.json")
    try:
        with open(manifest_path, encoding="utf-8") as fh:
            manifest = json.load(fh)
        if manifest.get("fingerprint") != fingerprint:
            return False
        destinations = _report_artifacts(name, raw_name)
        sources = {}
        for kind, destination in destinations.items():
            source = os.path.join(source_dir, kind)
            if (not os.path.isfile(source)
                    or _sha256_file(source)
                    != manifest["artifact_sha256"].get(kind)):
                return False
            sources[kind] = (source, destination)
        for source, destination in sources.values():
            os.makedirs(os.path.dirname(destination), exist_ok=True)
            temp = destination + f".{uuid.uuid4().hex}.tmp"
            shutil.copyfile(source, temp)
            os.replace(temp, destination)
        return _write_determinism_manifest(
            name, raw_name, fingerprint) is not None
    except (KeyError, OSError, TypeError, ValueError):
        return False


def _publish_canonical_report(name: str, raw_name: str,
                              fingerprint: dict) -> str | None:
    """Promote only a complete, review-free artifact set to immutable cache."""
    review = os.path.join(QTR_RAW_DIR, f"{raw_name}.review")
    artifacts = _report_artifacts(name, raw_name)
    if os.path.exists(review) or not all(
            os.path.isfile(path) for path in artifacts.values()):
        return None
    final_dir = _canonical_report_path(fingerprint)
    if os.path.isdir(final_dir):
        return final_dir
    os.makedirs(CANONICAL_REPORT_DIR, exist_ok=True)
    temp_dir = final_dir + f".{uuid.uuid4().hex}.tmp"
    os.makedirs(temp_dir)
    try:
        hashes = {}
        for kind, source in artifacts.items():
            destination = os.path.join(temp_dir, kind)
            shutil.copyfile(source, destination)
            hashes[kind] = _sha256_file(destination)
        with open(os.path.join(temp_dir, "manifest.json"), "w",
                  encoding="utf-8") as fh:
            json.dump({"fingerprint": fingerprint, "artifact_sha256": hashes},
                      fh, indent=2, sort_keys=True)
            fh.write("\n")
        try:
            os.replace(temp_dir, final_dir)
        except OSError:
            if not os.path.isdir(final_dir):
                raise
            shutil.rmtree(temp_dir)
        return final_dir
    except Exception:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise


def _write_determinism_manifest(name: str, raw_name: str,
                                fingerprint: dict) -> str | None:
    """Atomically publish the fingerprint after all verified artifacts exist."""
    review = os.path.join(QTR_RAW_DIR, f"{raw_name}.review")
    if os.path.exists(review):
        return None
    artifacts = _report_artifacts(name, raw_name)
    if not all(os.path.isfile(path) for path in artifacts.values()):
        return None
    path = _determinism_manifest(raw_name)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    payload = {
        "fingerprint": fingerprint,
        "artifact_sha256": {
            kind: _sha256_file(artifact)
            for kind, artifact in artifacts.items()
        },
    }
    tmp = path + f".{uuid.uuid4().hex}.tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, sort_keys=True)
        fh.write("\n")
    os.replace(tmp, path)
    return path


def _report_lock(raw_name: str) -> threading.Lock:
    with _REPORT_LOCKS_GUARD:
        return _REPORT_LOCKS.setdefault(raw_name, threading.Lock())


def _s3():
    import boto3
    return boto3.client("s3", region_name=os.getenv("AWS_REGION", "ap-south-1"))


def s3_upload(*paths):
    if not S3_BUCKET:
        return
    try:
        c = _s3()
        for p in paths:
            if p and os.path.exists(p):
                rel = os.path.relpath(p, OUT_DIR)
                c.upload_file(p, S3_BUCKET, f"{S3_PREFIX}/{rel}")
    except Exception:
        pass


def s3_delete(*paths):
    if not S3_BUCKET:
        return
    try:
        c = _s3()
        for p in paths:
            if not p:
                continue
            rel = os.path.relpath(p, OUT_DIR)
            c.delete_object(Bucket=S3_BUCKET, Key=f"{S3_PREFIX}/{rel}")
    except Exception:
        pass


def s3_restore():
    """On boot, pull anything the mirror has that the local disk doesn't."""
    if not S3_BUCKET:
        return
    try:
        c = _s3()
        for page in c.get_paginator("list_objects_v2").paginate(
                Bucket=S3_BUCKET, Prefix=S3_PREFIX + "/"):
            for o in page.get("Contents", []):
                rel = o["Key"][len(S3_PREFIX) + 1:]
                if not rel or rel.endswith("/"):
                    continue
                dst = os.path.join(OUT_DIR, rel)
                if os.path.exists(dst) and os.path.getsize(dst) == o["Size"]:
                    continue
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                c.download_file(S3_BUCKET, o["Key"], dst)
    except Exception:
        pass


def _save_jobs():
    try:
        import json
        with open(JOBS_FILE, "w") as fh:
            json.dump(jobs, fh)
        s3_upload(JOBS_FILE)
    except Exception:
        pass


s3_restore()
jobs: dict = {}   # job_id -> {state, company, message}
try:
    import json as _json
    with open(JOBS_FILE, encoding="utf-8") as _jobs_file:
        jobs = _json.load(_jobs_file)
    for _j in jobs.values():
        # a job that was 'running' when the server went down is gone — say so
        # instead of the confusing 'unknown job'
        if _j.get("state") == "running":
            _j["state"] = "error"
            _j["message"] = ("Processing was interrupted by a system restart — "
                             "please run this report again.")
except Exception:
    jobs = {}


app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "reports-radar-dev-secret-change-me")
app.config.update(SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE="Lax")

# ---------------------------------------------------------------------------
# Authentication — no database. A small set of in-app test accounts. Passwords
# are hashed at startup (never compared in plaintext). Override in production by
# supplying your own store and a real SECRET_KEY via the environment.
# ---------------------------------------------------------------------------
# username : (display name, password)
_TEST_ACCOUNTS = {
    "analyst": ("Analyst", "Radar@2026"),
}
USERS = {u: {"name": n, "pw": generate_password_hash(p)}
         for u, (n, p) in _TEST_ACCOUNTS.items()}

_PUBLIC_ENDPOINTS = {"login", "static", "favicon"}

# Reports Radar brand mark (radar sweep) — shared with the frp app.
LOGO_SVG = ('<svg class=logo viewBox="0 0 52 52" fill="none" width="30" height="30" '
            'xmlns="http://www.w3.org/2000/svg">'
            '<path d="M 0 52 V 0 A 52 52 0 0 1 52 52 Z" fill="#2563EB"/>'
            '<path d="M 0 14 A 38 38 0 0 1 38 52" fill="none" stroke="white" '
            'stroke-width="3" opacity="0.4" stroke-linecap="round"/>'
            '<path d="M 0 28 A 24 24 0 0 1 24 52" fill="none" stroke="white" '
            'stroke-width="7" stroke-linecap="round"/></svg>')


@app.route("/favicon.svg")
def favicon():
    return LOGO_SVG, 200, {"Content-Type": "image/svg+xml"}


@app.before_request
def _require_login():
    """Gate every route behind a session, except the login page / static files."""
    if request.endpoint in _PUBLIC_ENDPOINTS:
        return
    if not session.get("user"):
        # Async/JSON endpoints get a 401 so the UI can react; pages redirect.
        if request.path.startswith(("/tables/", "/mdna/")):
            return jsonify(error="Session expired — please log in again."), 401
        return redirect(url_for("login", next=request.path))


LOGIN_PAGE = r"""<!doctype html><html><head><meta charset=utf-8><title>Reports Radar — Sign in</title>
<meta name=viewport content="width=device-width,initial-scale=1">
<link rel=icon href="/favicon.svg" type="image/svg+xml">
<link rel=preconnect href="https://fonts.googleapis.com"><link rel=preconnect href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel=stylesheet>
<style>
 :root{--brand:#3B82F6;--brand-dark:#2563EB;--brand-subtle:#EFF6FF;--ink:#0F172A;--muted:#64748B;--line:#E5E7EB}
 *{box-sizing:border-box}
 body{font-family:Inter,-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif;margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;
      background:linear-gradient(160deg,#EFF6FF 0%,#DBEAFE 100%);color:var(--ink);-webkit-font-smoothing:antialiased}
 .box{background:#fff;width:min(400px,92vw);border-radius:16px;padding:34px 32px;box-shadow:0 1px 12px rgba(15,23,42,.08),0 24px 70px rgba(37,99,235,.14);border:1px solid var(--line)}
 .brandrow{display:flex;align-items:center;justify-content:center;gap:10px}
 .brandrow svg{flex-shrink:0}
 .brand{font-size:22px;font-weight:700;color:#1E40AF;letter-spacing:-.025em}
 .tag{text-align:center;color:var(--muted);font-size:13px;margin:6px 0 24px}
 label{font-weight:600;font-size:12.5px;display:block;margin:0 0 6px;color:#334155}
 .field{margin-bottom:16px}
 input{width:100%;padding:11px 12px;border:1px solid #D1D5DB;border-radius:9px;font-size:14px;font-family:inherit;transition:border-color .15s,box-shadow .15s}
 input:focus{outline:0;border-color:var(--brand);box-shadow:0 0 0 3px rgba(59,130,246,.15)}
 button{width:100%;margin-top:6px;background:var(--brand-dark);color:#fff;border:0;border-radius:9px;
        padding:12px;font-size:15px;font-weight:600;font-family:inherit;cursor:pointer;box-shadow:0 4px 14px rgba(37,99,235,.25)}
 button:hover{background:#1D4ED8}
 .err{background:#FEF2F2;border:1px solid #FECACA;color:#DC2626;font-size:13.5px;padding:10px 12px;border-radius:9px;margin-bottom:16px}
 .demo{margin-top:22px;padding-top:16px;border-top:1px solid var(--line);font-size:12px;color:var(--muted);line-height:1.7}
 .demo b{color:#334155}
 .demo code{background:var(--brand-subtle);color:var(--brand-dark);padding:1px 6px;border-radius:5px;font-size:11.5px}
</style></head><body>
<form class=box method=post action="/login">
 <div class=brandrow>__LOGO__<span class=brand>Reports Radar</span></div>
 <div class=tag>Sign in to continue</div>
 {% if error %}<div class=err>{{ error }}</div>{% endif %}
 <input type=hidden name=next value="{{ nxt }}">
 <div class=field>
  <label>Username</label>
  <input type=text name=username autofocus autocomplete=username required>
 </div>
 <div class=field>
  <label>Password</label>
  <input type=password name=password autocomplete=current-password required>
 </div>
 <button type=submit>Sign in</button>
</form>
</body></html>"""


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"):
        return redirect(url_for("tables_page"))
    error = None
    if request.method == "POST":
        u = (request.form.get("username") or "").strip().lower()
        p = request.form.get("password") or ""
        acct = USERS.get(u)
        if acct and check_password_hash(acct["pw"], p):
            session.clear()
            session["user"] = u
            session["name"] = acct["name"]
            nxt = request.form.get("next") or ""
            # Only allow same-site relative redirects (no open redirect).
            if not nxt.startswith("/") or nxt.startswith("//"):
                nxt = url_for("tables_page")
            return redirect(nxt)
        error = "Invalid username or password."
    html = render_template_string(LOGIN_PAGE, error=error,
                                  nxt=request.args.get("next", ""))
    return html.replace("__LOGO__", LOGO_SVG)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


TABLES_PAGE = r"""<!doctype html><html><head><meta charset=utf-8><title>Reports Radar</title>
<meta name=viewport content="width=device-width,initial-scale=1">
<link rel=icon href="/favicon.svg" type="image/svg+xml">
<link rel=preconnect href="https://fonts.googleapis.com"><link rel=preconnect href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel=stylesheet>
<style>
 :root{--brand:#2563EB;--brand2:#3B82F6;--brand-subtle:#EFF6FF;--ink:#0F172A;--muted:#64748B;--line:#E2E8F0;--bg:#F8FAFC;--topbar:64px}
 *{box-sizing:border-box}
 html,body{height:100%}
 body{font-family:Inter,-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif;margin:0;background:var(--bg);color:var(--ink);-webkit-font-smoothing:antialiased}
 /* top bar */
 .topbar{height:var(--topbar);background:#fff;color:var(--ink);display:flex;align-items:center;
         justify-content:space-between;padding:0 24px;border-bottom:1px solid var(--line);box-shadow:0 1px 3px rgba(15,23,42,.04);position:sticky;top:0;z-index:20}
 .brand{font-size:17.5px;font-weight:700;letter-spacing:-.025em;display:flex;align-items:center;gap:11px;color:var(--ink)}
 .brand svg{flex-shrink:0}
 .userbox{display:flex;align-items:center;gap:12px;font-size:13px;color:var(--muted)}
 .userbox .uname{color:var(--ink);font-weight:500}
 .userbox .avatar{width:30px;height:30px;border-radius:50%;background:var(--brand-subtle);color:var(--brand);display:flex;align-items:center;justify-content:center;font-weight:700;font-size:13px}
 .logout{color:var(--brand-dark,#1D4ED8);text-decoration:none;border:1px solid var(--line);border-radius:8px;padding:6px 13px;font-weight:600;font-size:12.5px;background:#fff;transition:all .15s}
 .logout:hover{background:var(--brand-subtle);border-color:var(--brand2)}
 /* two-column layout */
 .layout{display:grid;grid-template-columns:minmax(370px,410px) 1fr;gap:20px;padding:20px;align-items:start;
         height:calc(100vh - var(--topbar));max-width:1600px;margin:0 auto}
 .card{background:#fff;border:1px solid var(--line);border-radius:14px;padding:22px 24px;box-shadow:0 1px 3px rgba(15,23,42,.05)}
 label{font-weight:600;font-size:12.5px;display:block;margin:0 0 6px;color:#334155}
 input,select{width:100%;padding:10px 11px;border:1px solid #D1D5DB;border-radius:9px;font-size:14px;font-family:inherit;background:#fff;transition:border-color .15s,box-shadow .15s}
 input:focus,select:focus{outline:0;border-color:var(--brand2);box-shadow:0 0 0 3px rgba(59,130,246,.15)}
 .grid{display:grid;grid-template-columns:1fr 1fr;gap:12px}
 .field{margin-bottom:14px}
 .file-drop{border:1.5px dashed #CBD5E1;border-radius:11px;padding:16px;text-align:center;background:#F8FAFC;cursor:pointer;transition:border-color .15s,background .15s}
 .file-drop:hover,.file-drop.hover{border-color:var(--brand2);background:var(--brand-subtle)}
 .file-drop .fi{font-size:22px}
 .file-drop .ft{font-size:13px;color:var(--muted);margin-top:4px}
 .file-drop .fn{font-size:13px;color:var(--brand);font-weight:600;margin-top:4px;word-break:break-all}
 input[type=file]{display:none}
 button{background:var(--brand);color:#fff;border:0;border-radius:9px;padding:12px 22px;font-size:14px;font-weight:600;font-family:inherit;cursor:pointer;box-shadow:0 4px 14px rgba(37,99,235,.2)}
 button:hover{background:#1D4ED8}
 button:disabled{opacity:.5;cursor:not-allowed;background:var(--brand)}
 .actions{margin-top:16px}
 .actions button{width:100%}
 .idpreview{font-size:12px;color:var(--muted);margin-top:10px;text-align:center}
 .idpreview b{color:var(--brand);font-family:ui-monospace,SFMono-Regular,Menlo,monospace}
 #status{margin-top:14px;font-size:13.5px;padding:11px 13px;border-radius:9px;display:none;line-height:1.5}
 .run{background:#EFF6FF;border:1px solid #BFDBFE;color:#1E40AF}.ok{background:#F0FDF4;border:1px solid #BBF7D0;color:#166534}.err{background:#FEF2F2;border:1px solid #FECACA;color:#DC2626}.info{background:#EFF6FF;border:1px solid #BFDBFE;color:#1E40AF}
 .spin{display:inline-block;width:14px;height:14px;border:2px solid #BFDBFE;border-top-color:var(--brand);border-radius:50%;animation:s .8s linear infinite;vertical-align:-2px;margin-right:7px}
 @keyframes s{to{transform:rotate(360deg)}}
 .tabs{display:flex;gap:4px;margin-bottom:18px;border-bottom:2px solid var(--line)}
 .tab{background:none;border:0;border-bottom:3px solid transparent;margin:0;padding:9px 14px;font-size:14px;font-weight:600;font-family:inherit;color:var(--muted);cursor:pointer;border-radius:0;box-shadow:none}
 .tab:hover{color:var(--brand)}
 .tab.active{color:var(--brand);border-bottom-color:var(--brand)}
 /* results panel */
 .panel{background:#fff;border:1px solid var(--line);border-radius:14px;box-shadow:0 1px 3px rgba(15,23,42,.05);
        display:flex;flex-direction:column;height:calc(100vh - var(--topbar) - 40px);overflow:hidden}
 .panelhead{padding:16px 20px;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;gap:16px}
 .panelhead .pt{font-weight:700;font-size:15px;color:var(--ink);display:flex;align-items:center;gap:8px}
 .panelhead .cnt{font-size:12px;color:var(--brand);font-weight:600;background:var(--brand-subtle);border-radius:10px;padding:2px 9px}
 .search{position:relative;flex:1;max-width:280px}
 .search input{padding-left:32px}
 .search svg{position:absolute;left:10px;top:50%;transform:translateY(-50%);width:15px;height:15px;fill:none;stroke:var(--muted);stroke-width:2}
 .scrollarea{flex:1;overflow-y:auto;padding:16px 20px}
 .grp{margin-bottom:16px}
 .grphd{display:flex;align-items:center;gap:8px;font-weight:700;font-size:13px;color:var(--brand);margin:0 0 8px;
        position:sticky;top:-16px;background:#fff;padding:4px 0;z-index:1}
 .grphd .gn{background:var(--brand-subtle);color:var(--brand);border-radius:9px;padding:1px 8px;font-size:11px;font-weight:600}
 .wbgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:9px}
 .wbcard{background:#fff;border:1px solid var(--line);border-radius:10px;padding:9px 13px;display:flex;align-items:center;
         justify-content:space-between;gap:10px;transition:box-shadow .2s,border-color .2s,transform .2s}
 .wbcard:hover{border-color:var(--brand2);box-shadow:0 2px 12px rgba(59,130,246,.12);transform:translateY(-1px)}
 .wbname{font-weight:600;font-size:13px;color:var(--ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
 .wbright{display:flex;align-items:center;gap:10px;flex-shrink:0;font-size:12px;color:var(--muted)}
 .wbbadge{background:var(--brand-subtle);color:var(--brand);border-radius:10px;padding:1px 9px;font-weight:600;font-size:11.5px}
 .mdcard{cursor:pointer}
 .mdcard .go{color:var(--brand);font-weight:600;font-size:12.5px}
 .empty{color:var(--muted);font-size:13.5px;padding:30px 10px;text-align:center}
 a.dl{color:var(--brand);font-weight:600;text-decoration:none}
 a.dl:hover{text-decoration:underline}
 .del{color:var(--muted);background:none;border:none;cursor:pointer;font-size:13px;line-height:1;padding:0;font-family:inherit}
 .del:hover{color:#dc2626}
 /* confirm dialog */
 #confoverlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.5);z-index:60;align-items:center;justify-content:center}
 #confbox{background:#fff;border-radius:14px;width:min(380px,92vw);padding:24px 24px 20px;box-shadow:0 18px 60px rgba(15,23,42,.35);text-align:center}
 #confbox h3{margin:0 0 8px;font-size:17px;color:var(--ink)}
 #confbox p{margin:0 0 20px;font-size:13.5px;color:var(--muted);word-break:break-word}
 #confbox .row{display:flex;gap:10px}
 #confbox button{width:100%;box-shadow:none;margin:0}
 #confbox .cancel{background:var(--brand-subtle);color:var(--brand)}
 #confbox .danger{background:#dc2626}
 /* MD&A modal */
 #mdoverlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.5);z-index:50}
 #mdmodal{position:fixed;top:3vh;left:50%;transform:translateX(-50%);width:min(1420px,96vw);height:94vh;background:#fff;border-radius:14px;z-index:51;display:none;flex-direction:column;box-shadow:0 18px 60px rgba(15,23,42,.35)}
 #mdbody{display:flex;flex:1;min-height:0}
 #mdnav{width:250px;flex:0 0 250px;overflow-y:auto;border-right:1px solid var(--line);padding:14px 10px;background:#fafbff}
 #mdnav .tocitem{display:block;padding:7px 10px;margin:2px 0;border-radius:8px;font-size:12.5px;font-weight:600;color:#334155;cursor:pointer;line-height:1.3}
 #mdnav .tocitem:hover{background:#eef2ff;color:#3730a3}
 #mdnav .tocitem.active{background:#e0e7ff;color:#3730a3}
 #mdnav .tochead{font-size:10.5px;letter-spacing:.08em;text-transform:uppercase;color:var(--muted);padding:2px 10px 8px}
 @media (max-width:900px){#mdnav{display:none}}
 #mdhead{display:flex;align-items:center;justify-content:space-between;padding:14px 24px;border-bottom:1px solid var(--line)}
 #mdhead h1{margin:0;font-size:17px;color:var(--brand)}
 #mdclose{background:none;border:0;font-size:26px;color:var(--muted);cursor:pointer;margin:0;padding:0 4px;line-height:1;box-shadow:none}
 #mdview{overflow-y:auto;padding:10px 30px 30px;font-size:14px;line-height:1.55;flex:1}
 #mdview h2{color:var(--brand);border-bottom:2px solid var(--brand-subtle);padding-bottom:5px;margin-top:26px}
 #mdview h3{color:#1E40AF;margin-top:20px}
 #mdview h4{color:#3B82F6;margin-top:16px}
 #mdview table{border-collapse:collapse;margin:10px 0}
 #mdview th,#mdview td{border:1px solid var(--line);padding:5px 11px;text-align:right;font-size:13px}
 #mdview th{background:var(--brand-subtle);color:var(--brand)}
 #mdview td:first-child,#mdview th:first-child{text-align:left}
 #mdview li{margin:3px 0}
 @media(max-width:900px){.layout{grid-template-columns:1fr;height:auto}.panel{height:70vh}}
</style></head><body>
<div class=topbar>
 <div class=brand>__LOGO__<span>Reports Radar</span></div>
 <div class=userbox>
  <div class=avatar>__INITIAL__</div>
  <span class=uname>__USERNAME__</span>
  <a class=logout href="/logout">Log out</a>
 </div>
</div>
<div class=layout>
 <div class=card>
  <div class=tabs>
   <button type=button class="tab active" id=tab_qtr onclick="setMode('quarterly')">Quarterly filing</button>
   <button type=button class="tab" id=tab_mdna onclick="setMode('mdna')">MD&amp;A summary</button>
  </div>
  <form id=f>
   <input type=hidden name=mode id=mode value="quarterly">
   <div class=field>
    <label>Sector</label>
    <select name=sector id=sector>__SECTOR_OPTIONS__</select>
   </div>
   <div class=field>
    <label>Company</label>
    <input type=text name=company id=company placeholder="e.g. Wipro" autocomplete=off required oninput="onCompany()">
   </div>
   <div class=grid>
    <div class=field>
     <label>Financial year</label>
     <select name=fy id=fy onchange="updateId()">
      <option value=2025>2025</option>
      <option value=2026 selected>2026</option>
      <option value=2027 id=fy_2027>2027</option>
     </select>
    </div>
    <div class=field id=qtr_field>
     <label>Quarter</label>
     <select name=quarter id=quarter onchange="updateId()">
      <option value=Q1>Q1</option><option value=Q2>Q2</option>
      <option value=Q3>Q3</option><option value=Q4>Q4</option>
     </select>
    </div>
   </div>
   <div class=field>
    <label id=pdf_label>Quarterly results filing (PDF)</label>
    <div class=file-drop id=drop onclick="document.getElementById('pdf').click()">
     <div class=fi>📄</div>
     <div class=ft id=drop_txt>Click to choose a PDF, or drag &amp; drop here</div>
     <div class=fn id=drop_name></div>
    </div>
    <input type=file name=pdf id=pdf accept="application/pdf" required>
   </div>
   <div class=actions>
    <button id=go type=submit>Process</button>
    <div class=idpreview id=idprev></div>
   </div>
  </form>
  <div id=status></div>
 </div>

 <div class=panel>
  <div class=panelhead>
   <div class=pt><span id=panel_icon>📊</span><span id=panel_title>Generated reports</span><span class=cnt id=panel_cnt>0</span></div>
   <div class=search>
    <svg viewBox="0 0 24 24"><circle cx="11" cy="11" r="7"/><line x1="21" y1="21" x2="16.65" y2="16.65"/></svg>
    <input type=text id=filter placeholder="Filter by company…" oninput="renderReports()">
   </div>
  </div>
  <div class=scrollarea id=scrollarea>
   <div id=wbwrap><div class=wbgrid id=list></div></div>
   <div id=mdlist_wrap style="display:none"><div class=wbgrid id=mdlist></div></div>
  </div>
 </div>
</div>
<div id=mdoverlay onclick="closeMdna()"></div>
<div id=mdmodal>
 <div id=mdhead><h1 id=mdtitle></h1><button id=mdclose onclick="closeMdna()">&times;</button></div>
 <div id=mdbody><nav id=mdnav></nav>
<div id=mdview></div></div>
</div>
<div id=confoverlay onclick="if(event.target===this)closeConfirm()"><div id=confbox>
 <h3>Delete report?</h3>
 <p id=confname></p>
 <div class=row>
  <button type=button class=cancel onclick="closeConfirm()">Cancel</button>
  <button type=button class=danger id=confok>Delete</button>
 </div>
</div></div>
<script>
const f=document.getElementById('f'),st=document.getElementById('status'),go=document.getElementById('go');
const $=id=>document.getElementById(id);
let allWb=[], allMd=[];        // full report lists (unfiltered)
function normCompany(s){return s.trim().replace(/[^A-Za-z0-9]+/g,'_').replace(/^_+|_+$/g,'').toUpperCase();}
function normFy(s){const d=(s.match(/\d+/g)||[]).join('');return d?'FY'+d:'';}
function companyOf(id){return id.replace(/_(Q[1-4])?FY\d+.*$/,'');}   // strip period suffix
function canonName(){
 const m=$('mode').value, c=normCompany($('company').value), fy=normFy($('fy').value);
 if(!c||!fy)return '';
 return (m==='quarterly')?`${c}_${$('quarter').value}${fy}`:`${c}_${fy}`;
}
function updateId(){
 const n=canonName();
 $('idprev').innerHTML=n?`Output file: <b>${n}</b>`:'';
}
function onCompany(){ updateId(); $('filter').value=$('company').value; renderReports(); }
function setMode(m){
 $('mode').value=m;
 $('tab_qtr').className='tab'+(m==='quarterly'?' active':'');
 $('tab_mdna').className='tab'+(m==='mdna'?' active':'');
 $('qtr_field').style.display=(m==='quarterly')?'':'none';
 const opt2027=$('fy_2027'); opt2027.hidden=(m!=='quarterly');   // FY2027 = quarterly only
 if(m!=='quarterly'&&$('fy').value==='2027')$('fy').value='2026';
 $('pdf_label').textContent=(m==='quarterly')?'Quarterly results filing (PDF)':'Annual report (PDF)';
 $('go').textContent=(m==='mdna')?'Generate MD&A summary':'Process';
 $('panel_icon').textContent=(m==='mdna')?'📝':'📊';
 $('panel_title').textContent=(m==='mdna')?'MD&A summaries':'Generated reports';
 updateId(); renderReports();
}
// file drop
const drop=$('drop'),pdf=$('pdf');
pdf.onchange=()=>{ $('drop_name').textContent=pdf.files[0]?pdf.files[0].name:''; $('drop_txt').textContent=pdf.files[0]?'Selected file:':'Click to choose a PDF, or drag & drop here'; };
;['dragenter','dragover'].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.add('hover');}));
;['dragleave','drop'].forEach(ev=>drop.addEventListener(ev,e=>{e.preventDefault();drop.classList.remove('hover');}));
drop.addEventListener('drop',e=>{ if(e.dataTransfer.files.length){pdf.files=e.dataTransfer.files;pdf.onchange();} });

f.onsubmit=async e=>{e.preventDefault();
 if(!pdf.files.length){st.style.display='block';st.className='err';st.textContent='Please choose a PDF file.';return;}
 const fd=new FormData(f); go.disabled=true;
 st.style.display='block'; st.className='run'; st.innerHTML='<span class=spin></span>Uploading…';
 let r=await fetch('/tables/process',{method:'POST',body:fd}); let j=await r.json();
 if(j.error){st.className='err';st.textContent=j.error;go.disabled=false;return;}
 if(j.cached){
  st.className='info';st.innerHTML='♻️ '+j.message;go.disabled=false;
  await loadList();await loadMdna();
  if(j.kind==='mdna'&&j.doc){viewMdna(j.doc);}
  return;
 }
 poll(j.job);
};
async function poll(job){
 let r=await fetch('/tables/status/'+job); let j=await r.json();
 if(j.state==='running'){st.className='run';st.innerHTML='<span class=spin></span>'+j.message;setTimeout(()=>poll(job),30000);}
 else if(j.state==='done'){st.className='ok';st.innerHTML='✓ '+j.message;go.disabled=false;await loadList();await loadMdna();if(j.kind==='mdna'&&j.doc){viewMdna(j.doc);}}
 else{st.className='err';st.textContent='✗ '+j.message;go.disabled=false;}
}
// group a list of {id, ...} rows by company; returns [[company, rows], ...] sorted
function groupByCompany(rows){
 const map=new Map();
 rows.forEach(r=>{const c=companyOf(r.id)||r.id; if(!map.has(c))map.set(c,[]); map.get(c).push(r);});
 return [...map.entries()].sort((a,b)=>a[0].localeCompare(b[0]))
   .map(([c,rs])=>[c, rs.sort((a,b)=>a.id.localeCompare(b.id))]);
}
function renderReports(){
 const m=$('mode').value, q=normCompany($('filter').value);
 if(m==='mdna'){ $('wbwrap').style.display='none'; $('mdlist_wrap').style.display='block';
  const items=allMd.map(n=>({id:n.replace('_MDNA.md',''), file:n}))
                   .filter(r=>!q||normCompany(r.id).includes(q));
  $('panel_cnt').textContent=items.length;
  $('mdlist').innerHTML=items.length?groupByCompany(items).map(([c,rs])=>
   `<div class=grp><div class=grphd>${c}<span class=gn>${rs.length}</span></div><div class=wbgrid>`+
   rs.map(r=>{const period=r.id.slice(c.length).replace(/^_/,'')||'summary';
    return `<div class="wbcard mdcard" onclick="viewMdna('${r.file}')"><span class=wbname>${period}</span>`+
     `<span class=wbright><span class=wbbadge>MD&amp;A</span><span class=go>view →</span></span></div>`;}).join('')+
   `</div></div>`).join('')
   :`<div class=empty>${allMd.length?'No summaries match "'+$('filter').value+'".':'No MD&A summaries yet.'}</div>`;
 } else { $('mdlist_wrap').style.display='none'; $('wbwrap').style.display='block';
  const items=allWb.map(x=>({id:x.name.replace('.xlsx',''), file:x.name, tables:x.tables, review:x.review}))
                   .filter(r=>!q||normCompany(r.id).includes(q));
  $('panel_cnt').textContent=items.length;
  $('list').innerHTML=items.length?groupByCompany(items).map(([c,rs])=>
   `<div class=grp><div class=grphd>${c}<span class=gn>${rs.length}</span></div><div class=wbgrid>`+
   rs.map(r=>{const period=r.id.slice(c.length).replace(/^_/,'')||r.id;
    return `<div class=wbcard><span class=wbname title="${r.id}">${period}</span>`+
     `<span class=wbright>`+`<span class=wbbadge>${r.tables??'?'} sheets</span>`+
     `<a class=dl href="/tables/download/${encodeURIComponent(r.file)}" title="download">⬇</a>`+
     `<button class=del onclick="deleteReport('${r.file}')" title="delete">✕</button></span></div>`;}).join('')+
   `</div></div>`).join('')
   :`<div class=empty>${allWb.length?'No reports match "'+$('filter').value+'".':'No reports yet.'}</div>`;
 }
}
async function loadList(){
 let r=await fetch('/tables/list'); allWb=await r.json();
 renderReports();
}
function deleteReport(file){
 $('confname').textContent=file.replace('.xlsx','');
 $('confoverlay').style.display='flex';
 $('confok').onclick=async()=>{
  closeConfirm();
  let r=await fetch('/tables/delete/'+encodeURIComponent(file),{method:'POST'});
  if(!r.ok){alert('Delete failed.');return;}
  loadList();
 };
}
function closeConfirm(){$('confoverlay').style.display='none';}
function esc(s){return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
function inline(s){return s.replace(/\*\*(.+?)\*\*/g,'<b>$1</b>').replace(/(^|[^*])\*([^*]+)\*/g,'$1<i>$2</i>');}
function mdToHtml(md){
 const lines=md.split('\n'); let out=[],i=0,inList=0;
 const closeList=()=>{while(inList){out.push('</ul>');inList--;}};
 while(i<lines.length){
  let l=lines[i];
  if(l.startsWith('|')){ closeList();
   let rows=[]; while(i<lines.length&&lines[i].trim().startsWith('|')){rows.push(lines[i].trim());i++;}
   let html='<table>';
   rows.forEach((rw,ri)=>{
    const cells=rw.replace(/^\||\|$/g,'').split('|').map(c=>c.trim());
    if(cells.every(c=>/^[-: ]*$/.test(c)))return;
    const tag=(ri===0)?'th':'td';
    html+='<tr>'+cells.map(c=>`<${tag}>${inline(esc(c))}</${tag}>`).join('')+'</tr>';
   });
   out.push(html+'</table>'); continue;
  }
  let m;
  if(m=l.match(/^(#{1,6})\s+(.*)/)){ closeList(); out.push(`<h${m[1].length}>${inline(esc(m[2]))}</h${m[1].length}>`); }
  else if(/^\s*[-•]\s+/.test(l)){
   const depth=Math.floor((l.match(/^\s*/)[0].length)/2)+1;
   while(inList<depth){out.push('<ul>');inList++;}
   while(inList>depth){out.push('</ul>');inList--;}
   out.push('<li>'+inline(esc(l.replace(/^\s*[-•]\s+/,'')))+'</li>');
  }
  else if(/^---+$/.test(l.trim())){ closeList(); out.push('<hr>'); }
  else if(l.trim()===''){ closeList(); }
  else { closeList(); out.push('<p>'+inline(esc(l))+'</p>'); }
  i++;
 }
 closeList(); return out.join('\n');
}
async function loadMdna(){
 let r=await fetch('/mdna/list'); allMd=await r.json();
 renderReports();
}
async function viewMdna(n){
 let r=await fetch('/mdna/view/'+encodeURIComponent(n)); let md=await r.text();
 document.getElementById('mdtitle').textContent=n.replace('_MDNA.md','')+' — MD&A Summary';
 const html=mdToHtml(md);
 document.getElementById('mdview').innerHTML=html;
 // docs-style table of contents: one entry per theme, click to jump,
 // active section highlighted while scrolling
 const mv=document.getElementById('mdview');
 const nav=document.getElementById('mdnav');
 const heads=[...mv.querySelectorAll('h1,h2')];
 nav.innerHTML='<div class=tochead>Contents</div>';
 const items=heads.map((h,i)=>{h.id='theme_'+i;
  const it=document.createElement('a');it.className='tocitem';it.textContent=h.textContent;
  it.onclick=()=>{h.scrollIntoView({behavior:'smooth',block:'start'});};
  nav.appendChild(it);return it;});
 nav.style.display=heads.length>1?'block':'none';
 mv.onscroll=()=>{let a=0;
  heads.forEach((h,i)=>{if(h.getBoundingClientRect().top-mv.getBoundingClientRect().top<90)a=i;});
  items.forEach((it,i)=>it.classList.toggle('active',i===a));};
 mv.onscroll();
 document.getElementById('mdoverlay').style.display='block';
 document.getElementById('mdmodal').style.display='flex';
 document.getElementById('mdview').scrollTop=0;
 document.body.style.overflow='hidden';
}
function closeMdna(){
 document.getElementById('mdoverlay').style.display='none';
 document.getElementById('mdmodal').style.display='none';
 document.body.style.overflow='';
}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeMdna();});
updateId(); loadList(); loadMdna();
setMode('quarterly');
</script></body></html>"""


def _job_logger(raw_name: str):
    """Per-job debug log: output/logs/<raw_name>.log — every stage writes what
    it did and why (pages, corrections, flags, paid-call decisions, findings),
    so a wrong workbook can be diagnosed from the log alone. Thread-safe;
    also mirrors to stdout for the server console."""
    import datetime
    log_dir = os.path.join(ROOT, "output", "logs")
    os.makedirs(log_dir, exist_ok=True)
    path = os.path.join(log_dir, f"{raw_name}.log")
    lock = threading.Lock()

    def log(msg):
        line = f"{datetime.datetime.now().strftime('%H:%M:%S')} {msg}"
        with lock:
            with open(path, "a", encoding="utf-8") as fh:
                fh.write(line + "\n")
        print(f"[{raw_name}] {line}", flush=True)
    return log, path


def _run_tables_job(job_id: str, name: str, pdf_path: str, fin_only: bool,
                    vision: bool, mode: str = "auto", sector_id: str | None = None):
    jobs[job_id] = {"state": "running", "company": name, "message": "Processing…"}
    _save_jobs()
    run_lock = None
    try:
        if mode == "mdna":
            from src.engine.mdna import summarize_mdna
            md = summarize_mdna(pdf_path, log=lambda m: None)
            with open(os.path.join(MDNA_DIR, f"{name}_MDNA.md"), "w") as fh:
                fh.write(md)
            jobs[job_id] = {"state": "done", "company": name, "kind": "mdna",
                            "doc": f"{name}_MDNA.md",
                            "message": "Done — MD&A summary generated"}
            _save_jobs()
            s3_upload(os.path.join(MDNA_DIR, f"{name}_MDNA.md"))
            return
        # quarterly filing -> FULL client pipeline:
        # extract raw statements -> verify identities -> map to the client
        # taxonomy -> final client workbook (wide; long kept alongside)
        import pickle
        import sys as _sys
        from src.engine.tables_llm import extract_tables_smart
        from src.engine.client_map import (
            company_unit,
            map_quarter,
            propose_unmapped_mappings,
            to_wide,
            write_client_workbook_long,
            write_mapping_proposals,
        )
        from src.engine.sector_config import load_sector_assets
        sector_config, template, taxonomy = load_sector_assets(sector_id)
        m = re.match(r"(.+)_(Q[1-4])(FY\d+)$", name)
        raw_name = f"{m.group(1).lower()}_{m.group(2).lower()}{m.group(3)}" if m else name.lower()
        log, log_path = _job_logger(raw_name)
        # Serialise same-report jobs and re-check the fingerprint only after
        # acquiring the lock.  If two requests arrive together, the second one
        # waits for the first to publish its verified artifacts, then reuses
        # those exact bytes instead of starting another probabilistic read.
        run_lock = _report_lock(raw_name)
        run_lock.acquire()
        fingerprint = _report_fingerprint(
            pdf_path, name, mode, fin_only, vision, sector_config)
        if _deterministic_cache_hit(name, raw_name, fingerprint):
            out = os.path.join(CLIENT_DIR, f"{name}.xlsx")
            log("CACHE: exact PDF/config/pipeline fingerprint hit — "
                "reusing verified report byte-for-byte")
            jobs[job_id] = {
                "state": "done", "company": name,
                "message": f"<b>{os.path.basename(out)}</b> is ready.",
            }
            _save_jobs()
            return
        if _restore_canonical_report(name, raw_name, fingerprint):
            out = os.path.join(CLIENT_DIR, f"{name}.xlsx")
            log("CACHE: identical PDF/config/pipeline restored from the "
                "verified content-addressed canonical report")
            jobs[job_id] = {
                "state": "done", "company": name,
                "message": f"<b>{os.path.basename(out)}</b> is ready.",
            }
            _save_jobs()
            return
        import time as _time
        import traceback as _tb
        _t0 = _time.time()
        log(f"=== job start: {name} | sector {sector_config.sector_id} | "
            f"pdf {os.path.getsize(pdf_path):,} bytes ===")

        jobs[job_id]["message"] = "Processing…"
        from src.engine.tables_llm import maybe_trim_large_filing
        pdf_in = maybe_trim_large_filing(
            pdf_path, log=log,
            extraction_policy=sector_config.extraction_policy)
        log("EXTRACT: whole-file upload + per-statement questions")
        tables = extract_tables_smart(
            pdf_in, mode="quarterly", log=log,
            extraction_policy=sector_config.extraction_policy)
        rows = [(t.page, t.n, t.title, t.scope, t.section, t.grid) for t in tables]
        # Deterministic un-mangling of the boxed-negative paren/comma glyph
        # ('(21,914)' -> text layer '121.914)'). Runs on EVERY statement — the
        # signature (unbalanced bracket + comma-rendered-as-dot + stray '1') is
        # self-proving, so it is a no-op on clean cells and fires even where the
        # printed subtotals read correctly and no identity fails.
        from src.engine.source_align import repair_bracket_glyphs as _rbg
        _nfix = 0
        _rows2 = []
        for _pg, _n2, _t2, _sc2, _se2, _g2 in rows:
            _g3 = _rbg(_g2)
            _nfix += sum(1 for a, b in zip(sum(_g2, []), sum(_g3, [])) if a != b)
            _rows2.append((_pg, _n2, _t2, _sc2, _se2, _g3))
        rows = _rows2
        if _nfix:
            log(f"EXTRACT: bracket-glyph repair fixed {_nfix} cell(s) deterministically")
        pickle.dump(rows, open(os.path.join(QTR_RAW_DIR, f"{raw_name}.pkl"), "wb"))
        log(f"EXTRACT: {len(rows)} tables cached -> {raw_name}.pkl")

        # free completeness tripwire: a statement the PDF prints but the model
        # never returned must surface, not vanish
        from src.engine.filing_chat import unextracted_statements
        missing_stmts = unextracted_statements(
            pdf_in, tables,
            extraction_policy=sector_config.extraction_policy)
        for kind in missing_stmts:
            log(f"⚠ COMPLETENESS: '{kind}' heading printed in the PDF but no "
                "table extracted")

        _sys.path.insert(0, os.path.join(ROOT, "scripts"))
        import repair_raw as _rr

        jobs[job_id]["message"] = "Processing…"
        note = ""
        review = os.path.join(QTR_RAW_DIR, f"{raw_name}.review")
        try:
            os.remove(review)
        except OSError:
            pass
        failing = _rr._failing_statements(rows)
        log(f"VERIFY: {len(failing)} statement(s) failing identities"
            + ("".join(f"\n    · [{sc[:4]}] {t[:50]}: {'; '.join(b)[:90]}"
                       for _pg, t, sc, _se, b in failing) if failing else ""))
        if failing:
            # the printed arithmetic pinpoints which statements — free
            # positional repair first, pixels only where text cannot decide
            jobs[job_id]["message"] = "Processing…"
            try:
                _rr.repair(raw_name, pdf_path=pdf_path, log=lambda m_: log(f"REPAIR: {m_}"))
                rows = pickle.load(open(os.path.join(QTR_RAW_DIR, f"{raw_name}.pkl"), "rb"))
            except Exception:
                log("REPAIR: crashed —\n" + _tb.format_exc())
            failing = _rr._failing_statements(rows)
            log(f"REPAIR: {len(failing)} statement(s) still failing after repair")

        # TARGETED double-read (DOUBLE_READ=0 disables). The file-upload Q&A
        # already reads pages provider-side, and statements printed on DIGITAL
        # pages are verified cell-by-cell against the text layer. A second
        # independent read is reserved for statements with no text authority,
        # or statements that still fail after deterministic repair: repeated
        # values can appear elsewhere on a digital page and therefore pass
        # digit-coverage authority while occupying the wrong row.
        suspects, xread_notes, broad = [], [], []
        if os.getenv("DOUBLE_READ", "1") != "0":
            try:
                from src.engine import source_align as _sa
                from src.engine.client_map import (adopt_verified_second_reads,
                                                   compare_reads)
                from src.engine.client_map import statement_of as _sof
                from src.engine.filing_chat import _page_number_forms as _pnf
                _forms = _pnf(pdf_path)
                _untrusted = _sa.untrusted_text_pages(pdf_path)
                _still_bad = {(_sc2, _sof(_sec2, _t3))
                              for _pg2, _t3, _sc2, _sec2, _b2 in failing}
                _needs = {(_sc, _sof(_sec, _tt))
                          for _pg, _n, _tt, _sc, _sec, _g in rows
                          if _sof(_sec, _tt)
                          and ((_sc, _sof(_sec, _tt)) in _still_bad
                               or not _sa.has_text_authority(
                                   _g, _forms, _untrusted))}
                if _needs:
                    log(f"DOUBLE-READ: {len(_needs)} failing or source-unverified "
                        f"statement(s) -> paid cross-read: {sorted(_needs)}")
                    jobs[job_id]["message"] = "Processing…"
                    _t2 = extract_tables_smart(
                        pdf_in, mode="quarterly",
                        log=lambda m_: log(f"XREAD: {m_}"),
                        extraction_policy=sector_config.extraction_policy)
                    _lines = _sa.page_word_lines(pdf_path)
                    rows, _adopted = adopt_verified_second_reads(
                        rows, _t2, _needs, _forms, _lines, _untrusted)
                    if _adopted:
                        with open(os.path.join(QTR_RAW_DIR, f"{raw_name}.pkl"),
                                  "wb") as _raw_fh:
                            pickle.dump(rows, _raw_fh)
                        failing = _rr._failing_statements(rows)
                        for _a in _adopted:
                            log("DOUBLE-READ: adopted source-grounded "
                                f"[{_a['scope'][:4]}] {_a['stmt']} "
                                f"({_a['old_failures']} -> {_a['new_failures']})")
                    suspects, xread_notes, broad = compare_reads(rows, _t2, _needs)
                    log(f"DOUBLE-READ: {len(suspects)} unproven cell "
                        f"disagreement(s), {len(broad)} statement(s) disagree broadly")
                else:
                    log("DOUBLE-READ: skipped — every statement is verified "
                        "against the text layer (no paid second read)")
            except Exception:
                log("DOUBLE-READ: crashed —\n" + _tb.format_exc())

        # Report generation is a pure function of this PDF and the selected
        # sector configuration. Historical filings may be compared in a
        # separate QA process, but must never mutate or flag this deliverable:
        # otherwise an identical clean run changes merely because a sibling
        # cache happens to exist.
        review_items = ([f"[{sc}] {t}: {'; '.join(bad)}"
                         for _pg, t, sc, _sec, bad in failing]
                        + xread_notes
                        + [f"COMPLETENESS: '{k}' printed in the PDF but not extracted"
                           for k in missing_stmts])
        if review_items:
            with open(review, "w") as fh:
                fh.write("\n".join(review_items))
            note = ""                       # flags are ops-facing: the .review
                                            # sidecars are the internal queue
        else:
            note = ""

        jobs[job_id]["message"] = "Processing…"
        unit = company_unit(pdf_path, pages=sorted({r_[0] for r_ in rows}))
        log(f"UNIT: filing denomination = '{unit or '?'}'")
        # period hint from the job's own metadata (Q4 FY2026 …) — used only
        # when a statement's header cannot resolve its periods (bare-year
        # cash-flow columns); every use is flagged for review
        hint = (int(m.group(2)[1]), int(m.group(3)[2:])) if m else None
        log(f"MAP: taxonomy mapping (period hint {hint})")
        mapped = map_quarter(rows, template, taxonomy, default_unit=unit,
                             period_hint=hint, log=lambda m_: log(f"MAP: {m_}"))
        # Semantic suggestions are a separate, explicitly non-authoritative
        # review artifact. They never enter `mapped`, the workbook, its
        # verification, or the canonical-report cache.
        proposal_path = os.path.join(
            CLIENT_DIR, ".proposals", f"{raw_name}.json")
        try:
            proposal_payload = propose_unmapped_mappings(
                mapped, taxonomy, template)
            write_mapping_proposals(proposal_path, proposal_payload)
            suggested = sum(
                bool(item["suggested_fid"])
                for item in proposal_payload["proposals"])
            log("MAP-PROPOSALS: "
                f"{suggested}/{len(proposal_payload['proposals'])} "
                "unreviewed suggestion(s) written to review sidecar; "
                "0 promoted to report facts")
        except Exception:
            proposal_path = None
            log("MAP-PROPOSALS: advisory generation failed; authoritative "
                "report is unaffected\n" + _tb.format_exc())
        cache = os.path.join(CLIENT_DIR, ".cache", f"{raw_name}.pkl")
        os.makedirs(os.path.dirname(cache), exist_ok=True)
        pickle.dump(mapped, open(cache, "wb"))
        if EMIT_LONG:
            long_dir = os.path.join(CLIENT_DIR, "long")
            os.makedirs(long_dir, exist_ok=True)
            long_out = os.path.join(long_dir, f"{name}_long.xlsx")
        else:
            # PAUSED: build the long workbook to a throwaway temp path so the
            # wide deliverable can still be derived from it, then discard it.
            import tempfile
            long_out = os.path.join(tempfile.gettempdir(), f"{name}_long.xlsx")
        write_client_workbook_long(name.split("_")[0], mapped, template, long_out)
        out = os.path.join(CLIENT_DIR, f"{name}.xlsx")
        to_wide(long_out, out)
        # flag propagation + post-write verification: every extraction/mapping
        # flag and every identity the DELIVERED file breaks is stamped onto
        # the workbook itself (Review sheet + orange tabs) — nothing that
        # failed verification can look verified
        from src.engine.client_map import (
            annotate_flags,
            annotate_review,
            canonicalize_xlsx,
        )
        from src.engine.client_map import statement_of as _sof3
        n_flags = annotate_flags(out, mapped)
        import verify_delivered as _vd
        _findings = _vd.verify_workbook(
            out, raw_rows=rows, template=template, taxonomy=taxonomy,
            pdf_path=pdf_path)
        for f_ in _findings:
            log(f"POST-WRITE ⚠ [{f_['kind']}] {f_['stmt']}/{f_['scope']}: {f_['detail'][:140]}")
        _vd.annotate_findings(out, _findings)
        blocking_findings = [
            finding for finding in _findings
            if finding.get("kind") in ("identity", "empty")
        ]
        if blocking_findings:
            with open(review, "a", encoding="utf-8") as fh:
                if os.path.getsize(review):
                    fh.write("\n")
                fh.write("\n".join(
                    f"POST-WRITE [{finding['kind']}] "
                    f"{finding['stmt']}/{finding['scope']}: {finding['detail']}"
                    for finding in blocking_findings))
        annotate_review(out, suspects,
                        [{"stmt": _sof3(sec_, t_), "scope": sc_, "title": t_, "page": pg_}
                         for pg_, t_, sc_, sec_, _bad in failing] + broad
                        + [{"stmt": k, "scope": "?", "title": k, "page": ""}
                           for k in missing_stmts])
        canonicalize_xlsx(out)
        determinism_path = _write_determinism_manifest(
            name, raw_name, fingerprint)
        canonical_path = _publish_canonical_report(
            name, raw_name, fingerprint)
        log(f"DONE: {os.path.basename(out)} written | {n_flags} statement flag(s), "
            f"{len(_findings)} post-write finding(s), {len(suspects)} review cell(s) "
            f"| {_time.time() - _t0:.0f}s")
        jobs[job_id] = {"state": "done", "company": name,
                        "message": f"<b>{os.path.basename(out)}</b> is ready.{note}"}
        _save_jobs()
        s3_upload(out, cache,
                  os.path.join(QTR_RAW_DIR, f"{raw_name}.pkl"), review, log_path,
                  determinism_path,
                  *([proposal_path] if proposal_path else []),
                  *([os.path.join(canonical_path, name_)
                     for name_ in ("manifest.json", "workbook", "mapped", "raw")]
                    if canonical_path else []))
        if EMIT_LONG:
            s3_upload(long_out)
        else:
            try:                       # discard the throwaway long intermediate
                os.remove(long_out)
            except OSError:
                pass
    except Exception as e:
        import traceback
        traceback.print_exc()          # full detail stays in server logs only
        try:
            log("JOB FAILED —\n" + traceback.format_exc())
        except Exception:
            pass
        jobs[job_id] = {"state": "error", "company": name,
                        "message": "Processing failed. Please try again."}
        _save_jobs()
    finally:
        if run_lock is not None and run_lock.locked():
            run_lock.release()
        # Privacy: the uploaded report is deleted as soon as processing ends.
        try:
            os.remove(pdf_path)
        except OSError:
            pass


@app.route("/")
@app.route("/tables")
def tables_page():
    from src.engine.sector_config import (
        available_sector_ids,
        default_sector_id,
        load_sector_config,
    )

    name = session.get("name", "User")
    default = default_sector_id()
    options = "".join(
        f'<option value="{sid}"{" selected" if sid == default else ""}>'
        f'{load_sector_config(sid).name}</option>'
        for sid in available_sector_ids()
    )
    return (TABLES_PAGE
            .replace("__LOGO__", LOGO_SVG)
            .replace("__USERNAME__", name)
            .replace("__INITIAL__", (name[:1] or "U").upper())
            .replace("__SECTOR_OPTIONS__", options))


def _canonical_name(company: str, fy: str, quarter: str, mode: str) -> str:
    """Build a stable output identifier from the structured fields so the same
    company + period always maps to the same file (and can be de-duplicated)."""
    comp = re.sub(r"[^A-Za-z0-9]+", "_", company or "").strip("_").upper()
    digits = "".join(re.findall(r"\d+", fy or ""))
    fy_tag = f"FY{digits}" if digits else ""
    if not comp or not fy_tag:
        return ""
    if mode == "quarterly":
        q = (quarter or "").upper()
        if q not in ("Q1", "Q2", "Q3", "Q4"):
            return ""
        return f"{comp}_{q}{fy_tag}"
    return f"{comp}_{fy_tag}"          # annual / MD&A


@app.route("/tables/process", methods=["POST"])
def tables_process():
    from src.engine.sector_config import load_sector_config

    mode = request.form.get("mode") or "quarterly"
    if mode not in ("annual", "quarterly", "mdna", "auto"):
        mode = "quarterly"

    name = _canonical_name(request.form.get("company", ""),
                           request.form.get("fy", ""),
                           request.form.get("quarter", ""),
                           mode)
    if not name:
        need_q = " and quarter" if mode == "quarterly" else ""
        return jsonify(error=f"Please provide a company, fiscal year{need_q}."), 400
    try:
        sector_config = load_sector_config(request.form.get("sector") or None)
    except (OSError, TypeError, ValueError):
        return jsonify(error="Please select a configured sector."), 400

    # De-duplication: an existing filename is reusable only after checking the
    # newly uploaded PDF. This prevents a stale workbook produced by an older
    # extractor from permanently hiding a statement that the filing prints.
    if mode == "mdna":
        existing = os.path.join(MDNA_DIR, f"{name}_MDNA.md")
        if os.path.exists(existing):
            return jsonify(cached=True, kind="mdna", name=name,
                           doc=f"{name}_MDNA.md",
                           message=f"Already generated — showing existing MD&A summary for <b>{name}</b>.")
    else:
        existing = os.path.join(CLIENT_DIR, f"{name}.xlsx")

    pdf = request.files.get("pdf")
    if not pdf or not pdf.filename.lower().endswith(".pdf"):
        return jsonify(error="Please upload a PDF file."), 400
    # A request-specific upload path prevents simultaneous runs for the same
    # report from overwriting the file another worker is still reading.
    pdf_path = os.path.join(
        UPLOAD_DIR, f"tables_{name}_{uuid.uuid4().hex}.pdf")
    pdf.save(pdf_path)
    if mode != "mdna" and os.path.exists(existing):
        m_raw = re.match(r"(.+)_(Q[1-4])(FY\d+)$", name)
        raw = (f"{m_raw.group(1).lower()}_{m_raw.group(2).lower()}{m_raw.group(3)}"
               if m_raw else name.lower())
        fingerprint = _report_fingerprint(
            pdf_path, name, mode, True, True, sector_config)
        # Filename equality is not evidence of equality. Reuse only an intact
        # artifact set produced from these exact bytes and this exact sector
        # configuration/code fingerprint.
        if _deterministic_cache_hit(name, raw, fingerprint):
            try:
                os.remove(pdf_path)
            except OSError:
                pass
            return jsonify(cached=True, kind="tables", name=name,
                           message=f"Already generated — client workbook <b>{name}</b> is ready below.")
    fin_only = True   # financial-statements section only, always
    vision = True     # scanned pages are handled automatically in both modes
    job_id = uuid.uuid4().hex[:8]
    jobs[job_id] = {"state": "running", "company": name, "message": "Queued…"}
    _save_jobs()
    threading.Thread(target=_run_tables_job,
                     args=(job_id, name, pdf_path, fin_only, vision, mode,
                           sector_config.sector_id),
                     daemon=True).start()
    return jsonify(job=job_id)


@app.route("/tables/status/<job_id>")
def tables_status(job_id):
    return jsonify(jobs.get(job_id, {"state": "error", "message": "unknown job"}))


@app.route("/tables/list")
def tables_list():
    out = []
    for fn in sorted(os.listdir(CLIENT_DIR)):
        if not re.match(r"[A-Z0-9_]+_Q[1-4]FY\d+\.xlsx$", fn):
            continue
        ntab = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(os.path.join(CLIENT_DIR, fn), read_only=True)
            ntab = len(wb.sheetnames)
            wb.close()
        except Exception:
            pass
        m2 = re.match(r"(.+)_(Q[1-4])(FY\d+)\.xlsx$", fn)
        raw = f"{m2.group(1).lower()}_{m2.group(2).lower()}{m2.group(3)}" if m2 else fn.lower()
        needs_review = os.path.exists(os.path.join(QTR_RAW_DIR, f"{raw}.review"))
        out.append({"name": fn, "tables": ntab, "review": needs_review})
    return jsonify(out)


@app.route("/tables/download/<path:name>")
def tables_download(name):
    name = os.path.basename(name)                  # no path traversal
    path = os.path.join(CLIENT_DIR, name)
    if not name.endswith(".xlsx") or not os.path.exists(path):
        return "not found", 404
    return send_file(path, as_attachment=True, download_name=name)


@app.route("/tables/delete/<path:name>", methods=["POST"])
def tables_delete(name):
    name = os.path.basename(name)                  # no path traversal
    if not name.endswith(".xlsx"):
        return jsonify(error="bad name"), 400
    stem = name[:-len(".xlsx")]                     # e.g. WIPRO_Q1FY2026
    m = re.match(r"(.+)_(Q[1-4])(FY\d+)$", stem)
    raw = f"{m.group(1).lower()}_{m.group(2).lower()}{m.group(3)}" if m else stem.lower()

    # Every local artifact produced for this report (see _run_tables_job).
    targets = [
        os.path.join(CLIENT_DIR, name),                              # wide deliverable
        os.path.join(CLIENT_DIR, "long", f"{stem}_long.xlsx"),       # long companion
        os.path.join(CLIENT_DIR, ".cache", f"{raw}.pkl"),            # mapped-data cache
        _determinism_manifest(raw),                                  # input/artifact hashes
        os.path.join(QTR_RAW_DIR, f"{raw}.pkl"),                     # raw extracted tables
        os.path.join(QTR_RAW_DIR, f"{raw}.review"),                  # review sidecar
    ]
    if not os.path.exists(targets[0]):
        return jsonify(error="not found"), 404
    for p in targets:
        try:
            os.remove(p)
        except OSError:
            pass
    s3_delete(*targets)
    return jsonify(ok=True, name=name)


@app.route("/mdna/list")
def mdna_list():
    out = [fn for fn in sorted(os.listdir(MDNA_DIR)) if fn.endswith(".md")]
    return jsonify(out)


@app.route("/mdna/view/<path:name>")
def mdna_view(name):
    name = os.path.basename(name)                  # no path traversal
    path = os.path.join(MDNA_DIR, name)
    if not name.endswith(".md") or not os.path.exists(path):
        return "not found", 404
    with open(path) as fh:
        return fh.read(), 200, {"Content-Type": "text/plain; charset=utf-8"}


if __name__ == "__main__":
    # macOS AirPlay Receiver squats on :5000, so default to 8000 (override with PORT=...).
    port = int(os.getenv("PORT", "8005"))
    print(f" * Open  http://127.0.0.1:{port}")
    app.run(host="127.0.0.1", port=port, debug=False)
