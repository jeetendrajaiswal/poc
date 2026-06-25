"""Wide-format results export (no API calls).

Pivots the long-format pipeline outputs into one ROW per (company, scope, year) with the
59 datapoints as COLUMNS — the analysis-friendly layout:

    Year | Company | Type | <datapoint 1> | <datapoint 2> | ...

Value precedence per cell:  page-verified truth (data/verified_truth.csv)  >  pipeline value
(output/<company>_full.json)  >  "Not disclosed" (item absent)  /  "N/A" (scope not filed).
Year is auto-detected per report from the latest "March 31, 20XX" on the financial pages.

  .venv/bin/python -m src.export_wide                 # all companies with artifacts
  .venv/bin/python -m src.export_wide adani infosys
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys

import yaml

OUT = "output"
TAX = os.path.join(os.path.dirname(__file__), "..", "taxonomy", "definitions.yaml")
# search the web-app upload dir first, then ~/Downloads, so both uploaded and local PDFs resolve
PDF_DIRS = [os.path.join(os.path.dirname(__file__), "..", "uploads"), os.path.expanduser("~/Downloads/")]
VERIFIED = os.path.join(os.path.dirname(__file__), "..", "data", "verified_truth.csv")
ALL = ["reliance", "hindalco", "reddy", "itc", "infosys", "adani"]

_YEAR_RE = re.compile(r"(?:31st?\s+march|march\s+31)[,\s]+(20\d\d)", re.I)


def _taxonomy_keys() -> list[str]:
    return [i["key"] for i in yaml.safe_load(open(TAX))["items"]]


def _load_verified() -> dict:
    out: dict = {}
    if os.path.exists(VERIFIED):
        for r in csv.DictReader(open(VERIFIED)):
            out[(r["company"].lower(), r["key"], r["scope"].lower())] = r["verified_value"]
    return out


def detect_year(company: str) -> str:
    """Latest fiscal-year-ending found on the report's financial pages, as 'FY2025-26'."""
    pdf = next((os.path.join(d, f"{company}.pdf") for d in PDF_DIRS
                if os.path.exists(os.path.join(d, f"{company}.pdf"))), None)
    if not pdf:
        return ""
    try:
        import fitz
        doc = fitz.open(pdf)
        years: list[int] = []
        for i in range(doc.page_count):
            years += [int(y) for y in _YEAR_RE.findall(doc[i].get_text())]
        doc.close()
        if not years:
            return ""
        # The financial statements are two-column (current + prior year), so the reporting
        # period is the MAX of the two MOST-FREQUENT years — this ignores rare future dates
        # (bond/lease maturities like 2061) that a naive max() would wrongly pick up.
        from collections import Counter
        top = [yr for yr, _ in Counter(years).most_common(2)]
        y = max(top)
        return f"FY{y - 1}-{y % 100:02d}"
    except Exception:
        return ""


def _cell(company: str, key: str, scope: str, rec: dict, verified: dict) -> str:
    vk = (company.lower(), key, scope)
    if vk in verified:
        return verified[vk]
    if rec is None:
        return ""
    if rec.get("status"):           # N/A (sector / no consolidated statements)
        return "N/A"
    if rec.get("found"):
        return rec.get("value", "")
    return "Not disclosed"


def build(companies: list[str]):
    keys = _taxonomy_keys()
    verified = _load_verified()
    header = ["Year", "Company", "Type"] + keys
    rows = []
    for c in companies:
        path = os.path.join(OUT, f"{c}_full.json")
        if not os.path.exists(path):
            print(f"  {c}: no artifacts (skipped)")
            continue
        recs = json.load(open(path))
        idx = {(r["key"], r["scope"]): r for r in recs}
        # only the two real scopes; drop any stray 'both'/'front' rows
        scopes = [s for s in ("standalone", "consolidated") if any(r["scope"] == s for r in recs)]
        year = detect_year(c)
        for scope in scopes:
            row = [year, c.title(), scope.title()]
            for k in keys:
                row.append(_cell(c, k, scope, idx.get((k, scope)), verified))
            rows.append(row)
        print(f"  {c:10} {year or '(year?)':9} scopes={scopes}")
    return header, rows


def main(companies):
    header, rows = build(companies)
    if not rows:
        print("no rows"); return
    # CSV
    with open(os.path.join(OUT, "results_wide.csv"), "w", newline="") as fh:
        w = csv.writer(fh); w.writerow(header); w.writerows(rows)
    # Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Results (wide)"
    ws.append(header)
    hf = PatternFill("solid", fgColor="1F4E78"); hfont = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = hf; c.font = hfont; c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    cofill = {"Adani": "FDE9D9", "Infosys": "E2EFDA", "Hindalco": "DDEBF7",
              "Reliance": "FFF2CC", "Reddy": "EDEDED", "Itc": "F2DCDB"}
    for row in rows:
        ws.append(row); rr = ws.max_row
        for c in ws[rr]:
            c.border = bd; c.alignment = Alignment(vertical="center")
        ws.cell(rr, 2).fill = PatternFill("solid", fgColor=cofill.get(row[1], "FFFFFF"))
        for cc in (1, 2, 3):
            ws.cell(rr, cc).font = Font(bold=True)
        for cc in range(4, len(row) + 1):
            ws.cell(rr, cc).alignment = Alignment(horizontal="right", vertical="center")
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 13
    from openpyxl.utils import get_column_letter
    for i in range(4, len(header) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.freeze_panes = "D2"           # freeze Year/Company/Type + header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    wb.save(os.path.join(OUT, "results_wide.xlsx"))
    print(f"\nwrote output/results_wide.xlsx and output/results_wide.csv  "
          f"({len(rows)} rows x {len(header)} cols)")


if __name__ == "__main__":
    main([a for a in sys.argv[1:] if not a.startswith("--")] or ALL)
