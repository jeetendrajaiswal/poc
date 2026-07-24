"""Post-write verification of a DELIVERED client workbook (wide format).

The last line of defence: after the workbook is written, read the file back
and check the numbers the client will actually see —

  * every template-formula field whose value was REPORTED from the filing
    must equal the sum of its mapped components (when enough components are
    mapped for the sum to be meaningful);
  * Total Assets == Total Equity And Liabilities (the one identity the
    template's own formulas cannot express);
  * no statement sheet may be EMPTY when the raw extraction contains that
    statement (an empty sheet is silent data loss, the worst failure class).

Offline and deterministic. Used two ways:
  * library — webapp._run_tables_job calls verify_workbook() and stamps every
    finding onto the workbook's Review sheet before publishing;
  * CLI     — python scripts/verify_delivered.py [PATH ...|--all] audits
    existing deliverables in output/client/.
"""
import glob
import os
import pickle
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PKL_DIR = os.path.join(ROOT, "output", "qtr_raw")

_SHEETK = {"Income Statement": "income", "Balance Sheet": "balance",
           "Cash Flow": "cashflow"}


def _num(v):
    return v if isinstance(v, (int, float)) else None


def _close(a, b):
    return abs(a - b) <= max(1.5, 0.005 * abs(b))


def _raw_has(raw_rows, stmt, scope) -> bool:
    from src.engine.client_map import statement_of
    for _p, _n, t, sc, sec, g in raw_rows or []:
        s = f"{sec} {t}".lower()
        if sc == scope and statement_of(sec, t) == stmt and "ifrs" not in s and len(g) >= 3:
            return True
    return False


def _printed_digitset(pdf_path=None, raw_rows=None):
    """Digit-strings (>=4 digits, separator-agnostic) printed in the source.
    From the PDF text layer when available (independent), else from the raw
    extracted grids. Corruption-tolerant: '242.363' and '242,363' both -> 242363."""
    dig = set()
    text = ""
    if pdf_path and os.path.exists(os.path.expanduser(pdf_path)):
        import pymupdf
        doc = pymupdf.open(os.path.expanduser(pdf_path))
        text = "\n".join(p.get_text() for p in doc)
        doc.close()
    elif raw_rows:
        text = "\n".join(str(c) for _p, _n, _t, _s, _se, g in raw_rows
                         for row in g for c in row)
    for m in re.finditer(r"\d[\d,.]*\d|\d", text):
        d = re.sub(r"\D", "", m.group(0))
        if len(d) >= 4:
            dig.add(d)
    return dig


def _grounds(v, dig):
    """Does the value appear in the source? Try every plausible printed form —
    a printed '650.30' yields digits '65030', but the same value can also print
    as '650.3' ('6503') or the integer part; match if ANY (>=4 digits) is in
    the source digit-set. Avoids the trailing-zero mismatch that would falsely
    flag a correct decimal total."""
    a = abs(v)
    forms = {
        re.sub(r"\D", "", f"{a:.2f}"),                        # 650.30 -> 65030
        re.sub(r"\D", "", f"{a:.2f}".rstrip("0").rstrip(".")),  # -> 6503
        re.sub(r"\D", "", f"{a:.1f}"),                        # 650.3  -> 6503
        str(int(round(a))),                                   # 650
    }
    return any(len(d) >= 4 and d in dig for d in forms)


def verify_workbook(path: str, raw_rows=None, template=None, taxonomy=None,
                    pdf_path=None) -> list[dict]:
    """Findings for one wide workbook. Each: {stmt, scope, kind, detail}.

    kind='identity' — a real value error:
      * a PRINTED CROSS-IDENTITY between reported totals is broken, OR
      * a template total's mapped components don't reconstruct it AND the
        reported total does NOT ground to the source (its digits are printed
        nowhere) — i.e. the total itself is wrong (e.g. Wipro TNCA=468).
      Both -> flagged (orange tab + Review).
    kind='empty' — a statement present in the raw extraction is missing from
      the workbook (silent data loss) -> flagged.
    kind='info' — components don't reconstruct a total but the total DOES
      ground (it is a correct printed figure; a component line just lacks a
      dedicated field). NOT a value error -> Audit note only, no flag. This is
      what stops a 100%-correct statement (TCS, HM) from being orange-tabbed.
    """
    import openpyxl
    if template is None or taxonomy is None:
        from src.engine.sector_config import load_sector_assets
        _sector, loaded_template, loaded_taxonomy = load_sector_assets()
        template = template or loaded_template
        taxonomy = taxonomy or loaded_taxonomy
    dig = _printed_digitset(pdf_path, raw_rows)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    findings: list[dict] = []
    for scope in ("standalone", "consolidated"):
        for disp, stmt in _SHEETK.items():
            sn = f"{disp} - {scope.title()}"
            in_raw = _raw_has(raw_rows, stmt, scope) if raw_rows is not None else None
            if sn not in wb.sheetnames or len(rows := list(wb[sn].iter_rows(values_only=True))) <= 1:
                if in_raw:
                    findings.append({"stmt": stmt, "scope": scope, "kind": "empty",
                                     "detail": f"{sn} is empty, but the statement WAS "
                                               "extracted from the filing — data was lost "
                                               "on the way to the workbook"})
                continue
            hdr = rows[0]
            pcols = [j for j, h in enumerate(hdr)
                     if isinstance(h, str) and re.search(r"\d{4}-\d{2}-\d{2}|\(as at\)|\(\d+M\)", h)]
            mcol = next((j for j, h in enumerate(hdr) if h == "Method"), None)
            byfid, method = {}, {}
            for r in rows[1:]:
                if r and r[0] is not None:
                    fid = str(r[0]).lstrip("⚠ ").strip()
                    byfid[fid] = r
                    method[fid] = str(r[mcol]) if mcol is not None and mcol < len(r) else ""
            fields = template.get((stmt, scope), [])
            row2fid = {f.row: f.fid for f in fields}

            def val(fid, j):
                r = byfid.get(str(fid))
                return _num(r[j]) if r and j < len(r) else None

            # (1) printed cross-identities between reported totals
            for identity in getattr(taxonomy, "identities", {}).get(stmt, []):
                if scope not in identity["scopes"]:
                    continue
                name = str(identity["name"])
                required = [
                    (float(term["coefficient"]), str(term["fid"]))
                    for term in identity["terms"]
                    if term["presence"] == "required"
                ]
                optional = [
                    (float(term["coefficient"]), str(term["fid"]))
                    for term in identity["terms"]
                    if term["presence"] == "optional"
                ]
                result_fid = str(identity["result_fid"])
                if result_fid not in byfid:
                    continue
                for j in pcols:
                    rv = val(result_fid, j)
                    if rv is None:
                        continue
                    if not all(
                            val(fid, j) is not None
                            for _coefficient, fid in required):
                        continue
                    present_optional = [
                        (coefficient, fid) for coefficient, fid in optional
                        if val(fid, j) is not None
                    ]
                    tot = sum(
                        coefficient * val(fid, j)
                        for coefficient, fid in required + present_optional
                    )
                    if not _close(tot, rv):
                        findings.append({"stmt": stmt, "scope": scope, "kind": "identity",
                                         "detail": (f"{name} broken (col {j}): reported {rv:,.2f} "
                                                    f"vs {tot:,.2f} from the other reported totals")})
                        break

            # (2) template-total reconstruction, GATED ON GROUNDING: a mismatch
            # is a real error only when the reported total is not itself printed
            for f in fields:
                if not f.formula or str(f.fid) not in byfid or method.get(str(f.fid)) == "computed":
                    continue
                n_present = sum(1 for _sg, r_ in f.formula if str(row2fid.get(r_)) in byfid)
                for j in pcols:
                    tv = val(f.fid, j)
                    if tv is None:
                        continue
                    tot = have = 0
                    for sg, r_ in f.formula:
                        cv = val(row2fid.get(r_), j)
                        if cv is not None:
                            tot += sg * cv
                            have += 1
                    if have >= 2 and have >= 0.6 * max(1, n_present) and not _close(tot, tv):
                        grounded = _grounds(tv, dig)
                        findings.append({
                            "stmt": stmt, "scope": scope,
                            "kind": "info" if grounded else "identity",
                            "detail": (f"{f.name} [{f.fid}] col {j}: reported {tv:,.2f}, "
                                       f"components sum {tot:,.2f}"
                                       + ("" if grounded else
                                          " — reported total is NOT printed in the source "
                                          "(likely a misread/mis-mapped total)"))})
    wb.close()
    return findings


def annotate_findings(path: str, findings: list[dict]) -> None:
    """Stamp post-write findings onto the workbook: Review rows + orange tabs."""
    if not findings:
        return
    import openpyxl
    from src.engine.client_map import _review_append, _review_sheet
    STMT_NAME = {"income": "Income Statement", "balance": "Balance Sheet",
                 "cashflow": "Cash Flow", "segment": "Segment Finance"}
    real = [f for f in findings if f.get("kind") != "info"]   # info = Audit-only
    if not real:
        return
    wb = openpyxl.load_workbook(path)
    rv = _review_sheet(wb)
    for f in real:
        _review_append(rv, [STMT_NAME.get(f["stmt"], f["stmt"]), str(f["scope"]).title(),
                            "⚠ " + f["detail"], "", "", "", ""])
        sn = f"{STMT_NAME.get(f['stmt'], f['stmt'])} - {str(f['scope']).title()}"[:31]
        if sn in wb.sheetnames:
            wb[sn].sheet_properties.tabColor = "ED8B00"
    wb.save(path)


_PDFDIR = os.path.expanduser("~/Downloads/qtr_reports")


def _pdf_for(path: str):
    name = os.path.basename(path).rsplit(".", 1)[0]
    m = re.match(r"(.+)_(Q[1-4])(FY\\d+)$", name)
    raw = f"{m.group(1).lower()}_{m.group(2).lower()}{m.group(3)}" if m else name.lower()
    for cand in (os.path.join(_PDFDIR, f"{raw}.pdf"),
                 os.path.join(_PDFDIR, "full", f"{raw}_full.pdf")):
        if os.path.exists(cand):
            return cand
    return None


def _raw_rows_for(path: str):
    name = os.path.basename(path).rsplit(".", 1)[0]
    m = re.match(r"(.+)_(Q[1-4])(FY\d+)$", name)
    raw = f"{m.group(1).lower()}_{m.group(2).lower()}{m.group(3)}" if m else name.lower()
    p = os.path.join(PKL_DIR, f"{raw}.pkl")
    return pickle.load(open(p, "rb")) if os.path.exists(p) else None


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        sys.exit(__doc__)
    if args == ["--all"]:
        args = sorted(glob.glob(os.path.join(ROOT, "output", "client", "*.xlsx")))
    from src.engine.sector_config import load_sector_assets
    _sector, template, _taxonomy = load_sector_assets()
    n_bad = 0
    for path in args:
        base = os.path.basename(path)
        if base.startswith("~$") or base == "REMAINING_DIFFS.xlsx":
            continue
        fs = verify_workbook(
            path, raw_rows=_raw_rows_for(path), template=template,
            taxonomy=_taxonomy, pdf_path=_pdf_for(path))
        if fs:
            n_bad += int(any(finding["kind"] != "info" for finding in fs))
            print(f"\n{base}: {len(fs)} finding(s)")
            for f in fs:
                print(f"   [{f['kind']}] {f['stmt']}/{f['scope']}: {f['detail']}")
        else:
            print(f"{base}: OK")
    sys.exit(1 if n_bad else 0)
