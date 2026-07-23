"""Offline regression harness for the extraction pipeline — no paid calls.

Proves, against the REAL corpus (cached raws in output/qtr_raw + PDFs in
~/Downloads/qtr_reports), that:

  1. positional source reconciliation fixes the known failures
     (Infosys Travelling Cost 15.96→1,596; the corrupted OCI cell) and NEVER
     increases a statement's identity failures anywhere in the corpus;
  2. the period parser handles the corpus' header pathologies
     ('31March 2026', 'Jun 30, 2025', banner + bare years) and
     resolve_bare_periods recovers bare-year cash-flow columns from the
     filing's other statements / job metadata — never an empty sheet;
  3. the identity suite detects the known structural corruptions
     (duplicate period columns in infosys_q4 / techm_q2);
  4. verify_delivered finds the known broken deliverables (does not run by
     default on the whole output/client — use verify_delivered.py --all).

Usage: python scripts/regression_offline.py
Exit 0 = all assertions hold.
"""
import os
import pickle
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.engine import identities, source_align as sa
from src.engine.client_map import (MappedStatement, Period, _parse_period,
                                   infer_spans, resolve_bare_periods)
from src.engine.filing_chat import _page_number_forms

PKL = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "output", "qtr_raw")
PDF = os.path.expanduser("~/Downloads/qtr_reports")

PASS = FAIL = 0


def check(name, ok, detail=""):
    global PASS, FAIL
    PASS += ok
    FAIL += not ok
    print(f"  {'PASS' if ok else 'FAIL'}  {name}" + (f" — {detail}" if detail and not ok else ""))


def _pdf_for(name):
    full = os.path.join(PDF, "full", f"{name}_full.pdf")
    return full if os.path.exists(full) else os.path.join(PDF, f"{name}.pdf")


def _reconciled(name):
    """All statements of one cached raw, positionally reconciled."""
    rows = pickle.load(open(os.path.join(PKL, f"{name}.pkl"), "rb"))
    pdf = _pdf_for(name)
    forms = _page_number_forms(pdf)
    lines = sa.page_word_lines(pdf)
    scans = sa.untrusted_text_pages(pdf)
    out = []
    for _p, _n, title, scope, section, grid in rows:
        g2, rep = sa.reconcile_with_source(grid, section, title, forms, lines,
                                           scan_pages=scans)
        out.append((title, scope, section, grid, g2, rep))
    return out


print("== 1. named decimal / wrong-column repairs (infosys_q4 standalone) ==")
_infosys_pdf = _pdf_for("infosys_q4FY2026")
if os.path.exists(_infosys_pdf):
    rec = _reconciled("infosys_q4FY2026")
    for title, scope, section, g0, g2, rep in rec:
        if scope == "standalone" and "Results" in section and "IFRS" not in section:
            travel = next((r for r in g2 if any("ravel" in str(c) for c in r)), None)
            check("Travelling Cost 15.96 corrected to 1,596",
                  travel is not None and "1,596" in travel, str(travel))
            check("standalone results identities tie after reconcile",
                  not identities.failing(section, title, g2),
                  str(identities.failing(section, title, g2)))
else:
    print(f"  SKIP  Infosys Q4 source PDF is not present: {_infosys_pdf}")

print("== 2. corpus-wide: reconciliation never worsens identities ==")
tot_before = tot_after = 0
worse = []
for fn in sorted(os.listdir(PKL)):
    if not fn.endswith(".pkl") or fn.startswith("_"):
        continue
    name = fn[:-4]
    if not os.path.exists(_pdf_for(name)):
        continue
    for title, scope, section, g0, g2, rep in _reconciled(name):
        b = len(identities.failing(section, title, g0))
        a = len(identities.failing(section, title, g2))
        tot_before += b
        tot_after += a
        if a > b:
            worse.append((name, scope, title[:40]))
check(f"identity failures do not increase (before={tot_before}, after={tot_after})",
      tot_after <= tot_before and not worse, str(worse[:4]))

print("== 2b. glyph-corruption recovery, gated on printed identities ==")
# Boxed-total text-layer corruption seen on Latent View's consolidated cash
# flow: '(' extracts as the digit '1' ('(726.99)'->'(1726.99)'), commas
# extract as spaces ('1,645.32'->'1 645.32'). The repair is accepted only
# because op+inv+fin=net ties across BOTH period columns after it.
from src.engine.source_align import repair_glyph_by_identity as _rgi
_glyph = [["Net cash generated from operating activities", "1 645.32", "1 304.74"],
          ["Net cash used in investing activities (b)", "(1726.99)", "(12 416.55)"],
          ["Net cash used in financing activities (c)", "(67.98)", "(45.75)"],
          ["Net increase/(decrease) in cash and cash equivalents", "850.35", "(11157.56)"],
          ["Cash and cash equivalents at the beginning of the year", "694.43", "1,648.64"],
          ["Cash from business combination", "-", "200.25"],
          ["Effect of exchange rate fluctuations on cash held", "40.36", "3.10"],
          ["Cash and cash equivalents at the end of the year", "1,585.14", "694.43"]]
_gfix, _gn = _rgi(_glyph, "Statement of Cash Flows", "Cash Flow")
from src.engine.identities import run_checks as _rc
check("glyph repair ties out the corrupt cash flow (all identities pass)",
      _gn >= 1 and all(ok for _n, ok in _rc("Statement of Cash Flows", "Cash Flow", _gfix)),
      f"fixed={_gn}; " + " ".join(f"{n}={ok}" for n, ok in _rc("Statement of Cash Flows", "Cash Flow", _gfix)))
_inv = next((r for r in _gfix if "investing" in r[0].lower()), None)
check("investing subtotal recovered to (726.99)/(2416.55), not the corrupt 1-prefixed form",
      _inv is not None and _inv[1].strip("()") == "726.99" and _inv[2].strip("()") == "2416.55",
      str(_inv))
# a clean grid must be left untouched (no false repair)
_clean = [["Net cash generated from operating activities", "100.00", "90.00"],
          ["Net cash used in investing activities (b)", "(30.00)", "(20.00)"],
          ["Net cash used in financing activities (c)", "(10.00)", "(5.00)"],
          ["Net increase/(decrease) in cash and cash equivalents", "60.00", "65.00"],
          ["Cash and cash equivalents at the beginning of the year", "40.00", "-25.00"],
          ["Cash and cash equivalents at the end of the year", "100.00", "40.00"]]
_cfix, _cn = _rgi(_clean, "Statement of Cash Flows", "Cash Flow")
check("clean cash flow is not touched by the glyph repair", _cn == 0 and _cfix == _clean)

print("== 2c. comparative-column reconciliation from sibling filings ==")
# A scanned filing's prior-year column is corrected from the company's other
# filings that reported the same period — but ONLY when corroborated (weight>=2)
# AND it makes a broken printed identity tie, never on an ambiguous '?' span.
import verify_raw as _vrr
_atishay = os.path.join(PKL, "atishay_q4FY2026.pkl")
if os.path.exists(_atishay) and os.path.exists(os.path.join(PKL, "atishay_q3FY2026.pkl")):
    import copy as _copy
    from src.engine.client_map import statement_of as _so2
    _rows = pickle.load(open(_atishay, "rb"))
    _rows = _copy.deepcopy(_rows)
    # re-introduce the real scan misreads in the FY2025 income column
    for _p, _n, _t, _sc, _sec, _g in _rows:
        if _sc == "standalone" and _so2(_sec, _t) == "income":
            for _r in _g:
                _lab = " ".join(str(c) for c in _r if c and not str(c).replace(",", "").replace(".", "").replace("(", "").replace(")", "").replace("-", "").isdigit()).lower()
                if _lab.startswith("total expenses"):
                    _r[5] = "4,567.75"
                elif _lab.startswith("profit before tax"):
                    _r[5] = "757.89"
                elif _lab.startswith("purchase"):
                    _r[5] = "2,877.40"
    _fixed, _notes = _vrr.reconcile_comparatives("atishay_q4FY2026", _rows)
    _vals = {}
    for _p, _n, _t, _sc, _sec, _g in _fixed:
        if _sc == "standalone" and _so2(_sec, _t) == "income":
            for _r in _g:
                _lab = str(_r[0]).lower()
                if _lab.startswith("profit before tax"):
                    _vals["pbt"] = _r[5]
                elif _lab.startswith("total expenses"):
                    _vals["exp"] = _r[5]
    check("reconciliation restores the FY2025 misreads from sibling filings",
          len(_notes) == 3 and str(_vals.get("pbt")) == "957.89" and str(_vals.get("exp")) == "4367.75",
          f"notes={len(_notes)} vals={_vals}")
    # idempotent: reconciling the ALREADY-corrected rows makes no further change
    check("reconciliation is idempotent (a second pass corrects nothing)",
          _vrr.reconcile_comparatives("atishay_q4FY2026", _fixed)[1] == [],
          "a second pass over corrected rows should yield 0 corrections")
else:
    print("  (skipped — atishay fixtures not present)")

print("== 2d. bracket-glyph repair (unbalanced paren + comma-as-dot) ==")
from src.engine.source_align import _repair_bracket_glyph as _rbgc
check("'(21,914)' misread as '121.914)' is recovered",
      _rbgc("121.914)") == "(21,914)", _rbgc("121.914)"))
check("'(26,245)' misread as '126.245)' is recovered",
      _rbgc("126.245)") == "(26,245)", _rbgc("126.245)"))
check("ambiguous lone bracket '(6061' is NOT guessed",
      _rbgc("(6061") == "(6061", _rbgc("(6061"))
check("lost-digit '3.532)' is NOT silently 'fixed'",
      _rbgc("3.532)") == "3.532)", _rbgc("3.532)"))
check("positive comma-as-dot is recovered ('826.354' -> '826,354')",
      _rbgc("826.354") == "826,354" and _rbgc("20.977") == "20,977"
      and _rbgc("165.912") == "165,912", _rbgc("826.354"))
check("balanced-negative comma-as-dot is recovered ('(1.234)' -> '(1,234)')",
      _rbgc("(1.234)") == "(1,234)", _rbgc("(1.234)"))
check("clean values are untouched by glyph repair",
      all(_rbgc(v) == v for v in ["(23,432)", "29,107", "(393)", "1,833.09",
                                   "10,834.12", "2,31,511", "70,698", "37.92",
                                   "1,45,575.77", "(5)"]),
      "a clean cell was altered")

print("== 2e. parse_periods skips enumerator/label columns (no spurious period) ==")
from src.engine.client_map import parse_periods as _pp
_slno = [["Sl No", "Particulars", "As at March 31, 2026", "As at March 31, 2025"],
         ["", "Property, plant and equipment", "13,476", "14,096"],
         ["A", "Goodwill", "77,728", "76,230"],
         ["", "Total assets", "3,62,511", "3,35,906"]]
_pers = _pp(_slno)
check("a leading 'Sl No' column yields NO spurious '(unresolved)' period",
      len(_pers) == 2 and all(p.end for p in _pers)
      and [p.col for p in _pers] == [2, 3],
      f"periods={[(p.span, p.end, p.col) for p in _pers]}")

print("== 2f. period-header row with a trailing bare year is not eaten as data ==")
# LLM expands most columns but leaves the last as a bare '2025'; that single
# year cell must NOT make _header_and_data treat the whole period-header row as
# the first data row (which would leave parse_periods with no header to read).
_hdrbare = [["INFOSYS LIMITED", "", "", "", ""],
            ["Particulars", "Quarter ended March 31, 2026", "Quarter ended December 31, 2025",
             "Year ended March 31, 2026", "2025"],
            ["Revenue from operations", "38,641", "37,996", "148,819", "136,592"],
            ["Total expenses", "30,000", "29,000", "120,000", "110,000"]]
_hp = _pp(_hdrbare)
check("period header with a trailing bare year still parses (>=3 resolved periods)",
      sum(1 for p in _hp if p.end) >= 3,
      f"periods={[(p.span, p.end, p.col) for p in _hp]}")

print("== 3. structural detection (duplicate period columns) ==")
# synthetic fixtures — independent of the mutable cached pkls (a fresh run may
# legitimately fix a filing's dup column, which must not fail the harness)
_dupgrid = [["Particulars", "Mar-26", "Dec-25", "Mar-25"],
            ["Revenue", "46402", "45479", "45479"],
            ["Other income", "1159", "974", "974"],
            ["Total income", "47561", "46453", "46453"],
            ["Employee cost", "24688", "24103", "24103"],
            ["Total expenses", "36764", "35868", "35868"],
            ["Profit before tax", "10797", "10585", "10585"]]
_okgrid = [["Particulars", "Mar-26", "Dec-25", "Mar-25"],
           ["Revenue", "46402", "45479", "40925"],
           ["Total income", "47561", "46453", "42115"],
           ["Total expenses", "36764", "35868", "32452"],
           ["Profit before tax", "10797", "10585", "9663"]]
check("dup-column detector fires on a duplicated period column",
      identities.dup_columns(_dupgrid) is True)
check("dup-column detector clean on distinct columns",
      identities.dup_columns(_okgrid) is False)

print("== 4. period parsing pathologies ==")
p = _parse_period("Three months ended 31March 2026 (Audited)", 1)
check("'31March 2026' parses", (p.span, p.end) == ("3M", "2026-03-31"), str((p.span, p.end)))
p = _parse_period("Jun 30, 2025", 1)
check("'Jun 30, 2025' parses", p.end == "2025-06-30", p.end)
pers = infer_spans([_parse_period("Quarter ended June 30, 2025 (Unaudited)", 1),
                    _parse_period("Quarter ended June 30, 2024 (Unaudited)", 2)])
check("plain quarter headers still parse", all(q.end for q in pers))
# a document TITLE carrying a date must NOT be treated as a period banner that
# stamps one date onto every column (the Wipro period-collapse bug)
from src.engine.client_map import parse_periods as _pp
_grid = [["AUDITED STANDALONE FINANCIAL RESULTS FOR THE THREE MONTHS AND YEAR ENDED MARCH 31, 2026", "", "", "", "", ""],
         ["", "Three months ended", "Three months ended", "Three months ended", "Year ended", "Year ended"],
         ["Particulars", "March 31, 2026", "December 31, 2025", "March 31, 2025", "March 31, 2026", "March 31, 2025"],
         ["Revenue from operations", "100", "90", "80", "370", "340"]]
_pk = [(p.span, p.end) for p in _pp(_grid)]
check("title-with-date is not a banner (5 distinct periods, no collapse)",
      len(set(_pk)) == 5 and ("3M", "2026-03-31") in _pk and ("FY", "2026-03-31") in _pk,
      str(_pk))

print("== 5. bare-year cash flow resolved from sibling statements ==")
income = MappedStatement(
    periods=[Period("3M", "2026-03-31", "audited", 1),
             Period("FY", "2026-03-31", "audited", 4),
             Period("FY", "2025-03-31", "audited", 5)],
    facts={"256": {1: 1.0, 4: 4.0, 5: 5.0}}, sources={}, unmapped=[], verification=[])
cf = MappedStatement(
    periods=[Period("?", "", "", 1, raw="2026"), Period("?", "", "", 2, raw="2025")],
    facts={"17538": {1: 10.0, 2: 20.0}}, sources={}, unmapped=[], verification=[])
mapped = {("income", "standalone"): income, ("cashflow", "standalone"): cf}
resolve_bare_periods(mapped, period_hint=(4, 2026))
check("CF '2026' -> FY ended 2026-03-31",
      (cf.periods[0].span, cf.periods[0].end) == ("FY", "2026-03-31"),
      str([(q.span, q.end) for q in cf.periods]))
check("CF '2025' -> FY ended 2025-03-31", cf.periods[1].end == "2025-03-31")
check("resolution is flagged for review", bool(getattr(cf, "flags", None)))

cf2 = MappedStatement(
    periods=[Period("?", "", "", 1, raw=""), Period("?", "", "", 2, raw="")],
    facts={"17538": {1: 1.0, 2: 2.0}}, sources={}, unmapped=[], verification=[])
resolve_bare_periods({("cashflow", "standalone"): cf2}, period_hint=(2, 2026))
check("hint-only: Q2 FY2026 CF -> 6M ended 2025-09-30",
      (cf2.periods[0].span, cf2.periods[0].end) == ("6M", "2025-09-30"),
      str([(q.span, q.end) for q in cf2.periods]))

cf3 = MappedStatement(periods=[], facts={"17538": {1: 1.0, 2: 2.0}},
                      sources={}, unmapped=[], verification=[])
resolve_bare_periods({("cashflow", "standalone"): cf3}, period_hint=None)
check("no evidence at all: placeholder periods + loud flag (never empty)",
      len(cf3.periods) == 2 and any("could NOT" in f for f in (cf3.flags or [])))

print("== 5b. taxonomy 'Owners' alias: '-Owners' attribution rows pin deterministically ==")
import src.llm as _llm
_llm.extract_json = lambda **k: {"assignments": []}     # deterministic pins only
import importlib as _il
import src.engine.client_map as _cm
_il.reload(_cm)
_TPL = _cm.load_template(os.path.join(os.path.dirname(PKL), "..", "config", "client_template_software.xlsx"))
_TAX = _cm.load_taxonomy(os.path.join(os.path.dirname(PKL), "..", "config", "client_taxonomy_software.yaml"))
_lvg = next((g for _p, _n, t, sc, sec, g in pickle.load(open(os.path.join(PKL, "latent_view_q4FY2026.pkl"), "rb"))
             if sc == "consolidated" and _cm.statement_of(sec, t) == "income"), None)
if _lvg:
    _lms = _cm.map_statement(_lvg, "income", _TAX, _TPL[("income", "consolidated")])
    _po = sorted(_lms.facts.get("22299", {}).values())
    _to = sorted(_lms.facts.get("22303", {}).values())
    check("'-Owners' profit-attribution pins to 22299 (not collapsed into TCI)",
          527.52 in [round(v, 2) for v in _po] and 734.26 in [round(v, 2) for v in _to]
          and 1468.52 not in [round(v, 2) for v in _to], f"22299={_po[:2]} 22303={_to[:2]}")

print("== 6. verify_mapped catches an unbalanced mapped balance sheet ==")
from src.engine.client_map import load_template, verify_mapped
template = load_template(os.path.join(os.path.dirname(PKL), "..", "config",
                                      "client_template_software.xlsx"))
bs = MappedStatement(periods=[Period("?", "2026-03-31", "", 1)],
                     facts={"13771": {1: 115541.51}, "13776": {1: 101198.69}},
                     sources={}, unmapped=[], verification=[])
verify_mapped(bs, template[("balance", "standalone")], "balance")
check("Genesis-style TA≠TEL raises a statement flag",
      bool(getattr(bs, "flags", None))
      and any("assets = equity" in f.lower() for f in bs.flags))

print("== 7. paid work is gated: double-read only without text authority ==")
techm = pickle.load(open(os.path.join(PKL, "techm_q2FY2026.pkl"), "rb"))
forms = _page_number_forms(os.path.join(PDF, "techm_q2FY2026.pdf"))
untr = sa.untrusted_text_pages(os.path.join(PDF, "techm_q2FY2026.pdf"))
check("digital filing: NO statement needs a paid double-read",
      all(sa.has_text_authority(g, forms, untr) for _p, _n, t, sc, sec, g in techm))
ati = pickle.load(open(os.path.join(PKL, "atishay_q4FY2026.pkl"), "rb"))
forms_a = _page_number_forms(os.path.join(PDF, "atishay_q4FY2026.pdf"))
untr_a = sa.untrusted_text_pages(os.path.join(PDF, "atishay_q4FY2026.pdf"))
check("scanned filing: every statement needs the double-read",
      not any(sa.has_text_authority(g, forms_a, untr_a) for _p, _n, t, sc, sec, g in ati))

print("== 8. cross-read: provable disagreements resolved, unprovable flagged ==")
from src.engine.client_map import compare_reads, statement_of
from src.engine.tables import RawTable
t2 = []
for pg, n, tt, sc, sec, g in techm:
    g2 = [list(r) for r in g]
    if sc == "standalone" and "Results" in sec:
        for r in g2:
            lab = " ".join(str(c) for c in r).lower()
            if "subcontracting" in lab:
                r[2] = "13,109"           # identity-covered: provable
            if "basic" in lab and "dilut" not in lab:
                for j, c in enumerate(r):  # EPS: no identity covers it
                    if any(ch.isdigit() for ch in str(c)):
                        r[j] = "99.99"
                        break
    t2.append(RawTable(page=pg, n=n, title=tt, scope=sc, section=sec,
                       page_head="", units="", grid=g2))
keys = {(sc, statement_of(sec, tt)) for _p, _n, tt, sc, sec, _g in techm}
sus, _notes, _broad = compare_reads(techm, t2, keys)
check("exactly the unprovable cell is flagged (EPS), the provable one is not",
      len(sus) == 1 and "basic" in sus[0]["label"].lower(), str(sus))

print("== 9. cross-quarter consistency ==")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import verify_raw as _vr
# span-ambiguous INCOME columns must not produce false flags (a '?' span can be
# FY in one filing and a quarter in another). A genuine balance-sheet
# disagreement on a repeated 'as at' column IS legitimate and may appear.
_txq = _vr.cross_quarter_flags("techm_q2FY2026", techm)
check("techm q2 span-ambiguous income columns produce no false flags",
      all(d["stmt"] != "income" for d in _txq), str([d["note"][:50] for d in _txq]))
# cross-quarter mechanics on a SYNTHETIC pair (independent of live pkls, which
# a fresh run may legitimately clean): a repeated March column that disagrees
# between two filings must be flagged.
_bs = lambda ta: [["Particulars", "March 31, 2025"], ["ASSETS", ""],
                  ["Property plant and equipment", "1000"], ["Investments", "2000"],
                  ["Trade receivables", "3000"], ["Cash", "500"],
                  ["Total assets", ta], ["Total equity", "4000"],
                  ["Total liabilities", "2500"], ["Total equity and liabilities", ta]]
_q4 = [(1, 1, "Standalone Financial Results — Balance Sheet", "standalone",
        "Statement of Assets and Liabilities", _bs("6500"))]
_q2 = [(1, 1, "Standalone Financial Results — Balance Sheet", "standalone",
        "Statement of Assets and Liabilities", _bs("9999"))]  # disagrees on repeated col
import types as _types
_orig_listdir = os.listdir
_orig_open = open
def _fake_flags():
    # feed the synthetic q2 as the sibling of the synthetic q4
    import pickle as _pk
    saved = {}
    for nm, rws in [("xqtest_q2FY2026", _q2)]:
        p = os.path.join(_vr.PKL_DIR, f"{nm}.pkl"); _pk.dump(rws, open(p, "wb")); saved[nm] = p
    try:
        return _vr.cross_quarter_flags("xqtest_q4FY2026", _q4)
    finally:
        for p in saved.values():
            try: os.remove(p)
            except OSError: pass
_xf = _fake_flags()
check("cross-quarter flags a disagreeing repeated column",
      any("disagrees" in d["note"] for d in _xf), str([d["note"][:60] for d in _xf]))

print("== 10. process hygiene: prompts, truncation plumbing, completeness ==")
import inspect

from src.engine import filing_chat as _fc
from src.llm import ask_text as _at
check("extraction has its own prompt without the Q&A 'prefer consolidated' bias",
      "Prefer consolidated" not in _fc.EXTRACT_PROMPT
      and "NEVER invent data" in _fc.EXTRACT_PROMPT
      and "parentheses are NEGATIVE" in _fc.EXTRACT_PROMPT)
check("questions request the denomination line + period banner",
      "denomination line" in _fc._DETAIL and "period banner" in _fc._DETAIL)
_balance_questions = [(sc, lab) for sc, lab, _q in _fc._QUESTIONS
                      if "Balance Sheet" in lab]
check("standalone and consolidated balance sheets use isolated calls",
      {sc for sc, _lab in _balance_questions} == {"standalone", "consolidated"}
      and len(_balance_questions) == 2)
check("ask_text exposes the truncation signal",
      "with_status" in inspect.signature(_at).parameters)

tabs = [type("T", (), {"section": sec, "title": t})()
        for _p, _n, t, _sc, sec, _g in techm]
stripped = [x for x in tabs if "Assets and Liab" not in x.section]
check("completeness tripwire: stripped balance sheet is reported missing",
      "balance sheet" in _fc.unextracted_statements(
          os.path.join(PDF, "techm_q2FY2026.pdf"), stripped))
check("completeness tripwire: full extraction reports nothing missing",
      _fc.unextracted_statements(os.path.join(PDF, "techm_q2FY2026.pdf"), tabs) == [])

print("== 10a. incomplete cached workbooks are not reused ==")
_cache_pdf = os.path.join(PDF, "techm_q2FY2026.pdf")
if os.path.exists(_cache_pdf):
    import tempfile
    from openpyxl import Workbook
    from src.webapp import _cached_workbook_missing_statements

    _tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    _tmp.close()
    try:
        _wb = Workbook()
        _wb.active.title = "Income Statement - Standalone"
        _wb.active.append(["Field id", "Display Name", "2026-03-31"])
        _wb.active.append(["1", "Revenue", 1])
        _cf = _wb.create_sheet("Cash Flow - Standalone")
        _cf.append(["Field id", "Display Name", "2026-03-31"])
        _cf.append(["2", "Net cash flow", 1])
        _wb.save(_tmp.name)
        check("cache guard rejects a workbook missing a printed balance sheet",
              "balance sheet" in
              _cached_workbook_missing_statements(_tmp.name, _cache_pdf))
        _bs = _wb.create_sheet("Balance Sheet - Standalone")
        _bs.append(["Field id", "Display Name", "2026-03-31"])
        _bs.append(["3", "Total assets", 1])
        _wb.save(_tmp.name)
        check("cache guard accepts the balance-sheet tab once populated",
              "balance sheet" not in
              _cached_workbook_missing_statements(_tmp.name, _cache_pdf))
    finally:
        try:
            os.remove(_tmp.name)
        except OSError:
            pass
else:
    print(f"  SKIP  cache completeness source PDF is not present: {_cache_pdf}")

print("== 10b. filing-level scope is resolved before extraction and mapping ==")
_single_entity = """
The Company confirms that it does not have any subsidiary, associate, or
joint venture company as of March 31, 2026.
The board also approved a split/consolidation of equity shares.
"""
check("single-entity disclosure resolves the filing to standalone",
      _fc.document_scope_from_text(_single_entity) == "standalone")
check("share split/consolidation wording is not treated as consolidated financials",
      _fc.document_scope_from_text(
          "The company approved split/consolidation of its equity shares."
      ) == "unknown")
check("an explicit consolidated-financial heading prevents a global override",
      _fc.document_scope_from_text(
          _single_entity + "\nStatement of Consolidated Financial Results"
      ) == "unknown")
_scoped_questions = _fc._questions_for_document_scope("standalone")
check("standalone-only filing never makes a consolidated extraction call",
      not any(sc == "consolidated" for sc, _lab, _q in _scoped_questions))
check("combined statement questions are constrained to standalone",
      all("STANDALONE statement only" in q
          for sc, _lab, q in _scoped_questions if sc == "unknown"))

print(f"\n{PASS} passed, {FAIL} failed")
sys.exit(1 if FAIL else 0)
