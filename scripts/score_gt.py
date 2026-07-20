"""Score engine mapping decisions against the human GT client files.

GT files (~/Downloads/software/<Company>-Q-2 2025.xlsx) hold the human's
label -> field-id decision per statement sheet. The engine's decision per
label is reconstructed from the generated workbook's provenance: Sub-items
lists the exact report lines summed into each mapped field; the Unmapped
sheet lists lines the engine assigned no field.

A GT line scores as a match when the engine put that label (normalised) on
the same fid. Labels repeated with different fids (e.g. 'Billed' under
receivables and payables) are compared as multisets within a statement.

Usage:
  python scripts/score_gt.py <workbook_dir> [<workbook_dir> ...]
"""
import collections
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.engine.client_map import norm_label

GT_DIR = os.path.expanduser("~/Downloads/software")
GT_FILES = {
    "ATISHAY": "Atishay Ltd- Q-2 2025.xlsx",
    "HCL": "HCL Technologies Ltd-Q-2 2025.xlsx",
    "HM": "Happiest Mind - Q-2 2025.xlsx",
    "TCS": "Tata Consultancy Services Ltd-Q-2 2025.xlsx",
    "WIPRO": "Wipro Q-2-2025.xlsx",
}
SHEET2KEY = {"P&L": "Income Statement", "BS": "Balance Sheet",
             "CFS": "Cash Flow", "SR": "Segment Finance"}


def load_gt(path):
    """{(stmt, scope): [(norm_label, fid), ...]}"""
    wb = openpyxl.load_workbook(path, read_only=True)
    out = collections.defaultdict(list)
    for sn in wb.sheetnames:
        m = re.match(r"(P&L|BS|CFS|SR)_(Standalone|Consolidated)", sn)
        if not m:
            continue
        key = (SHEET2KEY[m.group(1)], m.group(2))
        for r in wb[sn].iter_rows(values_only=True):
            lab, fid = str(r[0] or "").strip(), r[2]
            if not lab or fid is None or lab.lower() == "particulars":
                continue
            out[key].append((norm_label(lab), str(fid).split(".")[0]))
    return out


def load_engine(path):
    """{(stmt, scope): [(norm_label, fid-or-''), ...]} from provenance."""
    wb = openpyxl.load_workbook(path, read_only=True)
    out = collections.defaultdict(set)          # dedup (label, fid, period) noise
    for sn in wb.sheetnames:
        if sn == "Audit":
            continue
        if sn == "Unmapped":
            for r in list(wb[sn].iter_rows(values_only=True))[1:]:
                stmt, scope, lab = r[0], r[1], str(r[2] or "")
                out[(stmt, scope)].add((norm_label(lab), ""))
            continue
        m = re.match(r"(.+) - (Standalone|Consolidated)", sn)
        if not m:
            continue
        key = (m.group(1), m.group(2))
        for r in list(wb[sn].iter_rows(values_only=True))[1:]:
            fid, method, sub = str(r[0]), str(r[8] or ""), str(r[9] or "")
            if method == "computed":
                continue
            for part in sub.split("  +  "):
                lab = part.rsplit(" = ", 1)[0].strip()
                if lab:
                    out[key].add((norm_label(lab.replace(" [adjusted]", "")), fid))
    return {k: list(v) for k, v in out.items()}


def score(gt, eng):
    """multiset match per (stmt, scope); returns (ok, total, misses)."""
    ok = tot = 0
    misses = []
    for key, pairs in gt.items():
        epairs = eng.get(key, [])
        by_lab = collections.defaultdict(list)
        for lab, fid in epairs:
            by_lab[lab].append(fid)
        for lab, fid in pairs:
            if lab not in by_lab:               # line absent from filing/extraction
                continue
            tot += 1
            if fid in by_lab[lab]:
                ok += 1
                by_lab[lab].remove(fid)
            else:
                misses.append((key, lab, fid, sorted(set(by_lab[lab]))))
    return ok, tot, misses


if __name__ == "__main__":
    dirs = sys.argv[1:] or ["output/client"]
    for d in dirs:
        print(f"### {d}")
        g_ok = g_tot = 0
        all_misses = []
        for comp, gtf in sorted(GT_FILES.items()):
            wbp = os.path.join(d, f"{comp}_Q2FY2026_dummytest.xlsx")
            if not os.path.exists(wbp):
                continue
            gt = load_gt(os.path.join(GT_DIR, gtf))
            eng = load_engine(wbp)
            ok, tot, misses = score(gt, eng)
            g_ok += ok; g_tot += tot
            all_misses += [(comp,) + m for m in misses]
            print(f"  {comp:8s} {ok}/{tot}")
        pct = 100.0 * g_ok / g_tot if g_tot else 0
        print(f"  TOTAL    {g_ok}/{g_tot}  ({pct:.1f}%)")
        for comp, key, lab, fid, got in all_misses:
            print(f"    MISS {comp} {key[0][:14]}/{key[1][:4]}: '{lab[:40]}' GT->{fid} engine->{got or 'unmapped'}")
