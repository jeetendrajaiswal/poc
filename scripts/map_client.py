"""Map verified raw extractions to the client template workbooks.

Reads output/qtr_raw/<name>.pkl (verified raw tables), maps every statement
to the client taxonomy (LLM-assisted -> API cost ~$0.10-0.26 per filing),
and writes output/client/<NAME>_dummytest.xlsx in long format.

Usage:
  python scripts/map_client.py <name> [<name> ...]   # e.g. wipro_q4FY2026
  python scripts/map_client.py --all                 # all filings in output/qtr_raw
"""
import os
import pickle
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.engine.client_map import (load_template, load_taxonomy, map_quarter,
                                   write_client_workbook_long)
from src.llm import usage_cost, reset_usage

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PKL_DIR = os.path.join(ROOT, "output", "qtr_raw")
OUT_DIR = os.path.join(ROOT, "output", "client")
PDF_DIR = os.path.expanduser("~/Downloads/qtr_reports")
TEMPLATE = os.path.expanduser("~/Downloads/software/Software Template.xlsx")

_UNITS = {"lac": "lakhs", "lakh": "lakhs", "crore": "crores",
          "million": "millions", "billion": "billions", "thousand": "thousands"}


def filing_unit(pdf_path: str) -> str:
    """The filing's printed denomination ('(Rs. in lakhs)', '(₹ crore)').
    Anchored on the parenthetical unit text, which survives dirty OCR even
    when the currency symbol does not."""
    import collections
    import re
    import pymupdf
    votes = collections.Counter()
    doc = pymupdf.open(pdf_path)
    for page in doc:
        t = " ".join(page.get_text().split())
        if len(re.findall(r"\d[\d,]{2,}", t)) < 15:
            continue                              # statement pages only
        for m in re.finditer(r"\(.{0,20}\b(lakh|lac|crore|million|billion|thousand)s?\b.{0,20}\)", t, re.I):
            span = m.group(0).lower()
            if "$" in span or "usd" in span or "us$" in span:
                continue                          # foreign-currency note, not the denomination
            votes[_UNITS[m.group(1).lower()]] += 1
    doc.close()
    return votes.most_common(1)[0][0] if votes else ""

TAXONOMY = os.path.join(ROOT, "config", "client_taxonomy_software.yaml")


def to_wide(long_path: str, wide_path: str) -> None:
    """Pivot a long-format workbook (one row per field x period) into the
    DEFAULT wide format: one row per field, one column per period.
    Deterministic — works offline from the long file alone."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    src = openpyxl.load_workbook(long_path, read_only=True)
    out = openpyxl.Workbook()
    out.remove(out.active)
    for sn in src.sheetnames:
        rows = list(src[sn].iter_rows(values_only=True))
        ws = out.create_sheet(sn)
        if sn == "Audit" or not rows:
            for r in rows:
                ws.append(list(r))
            for c in ws[1]:
                c.font = Font(bold=True)
            ws.freeze_panes = "A2"
            continue
        if sn == "Unmapped":
            key_cols, val_i = (0, 1, 2, 6, 7), 5      # stmt, scope, label, denom, reason
            per_i, mon_i, aud_i = 3, 4, None
            lead = ["Statement", "Scope", "Report line (as printed)"]
            tail = ["Denomination", "Reason"]
        else:
            key_cols, val_i = (0, 1, 6, 7, 8), 5      # fid, name, denom, cur, method
            per_i, mon_i, aud_i = 2, 3, 4
            lead = ["Field id", "Display Name"]
            tail = ["Denomination", "Currency", "Method",
                    "Sub-items (report lines / calculation; latest period)"]
        pers, fields, data = [], [], {}
        for r in rows[1:]:
            if r is None or r[0] is None:
                continue
            months = str(r[mon_i]) if r[mon_i] not in (None, "") else ""
            aud = str(r[aud_i]).strip() if aud_i is not None and r[aud_i] else ""
            head = str(r[per_i] or "")
            head += f" ({months}M)" if months else " (as at)"
            if aud:
                head += f" [{aud}]"
            if head not in pers:
                pers.append(head)
            k = tuple(r[i] for i in key_cols)
            if k not in data:
                data[k] = {"_sub": r[9] if sn != "Unmapped" and len(r) > 9 else "",
                           "_subp": {}}
                fields.append(k)
            data[k].setdefault(head, r[val_i])
            if sn != "Unmapped" and len(r) > 9 and r[9]:
                data[k]["_subp"].setdefault(head, str(r[9]))
        hdr = lead + pers + tail
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True)
        for k in fields:
            if sn == "Unmapped":
                stmt, scope, lab, denom, reason = k
                row = [stmt, scope, lab] + [data[k].get(p, "") for p in pers] + [denom, reason]
            else:
                fid, name, denom, cur, method = k
                row = ([fid, name] + [data[k].get(p, "") for p in pers]
                       + [denom, cur, method, data[k]["_sub"]])
            ws.append(row)
            if sn != "Unmapped":
                # every period cell carries ITS OWN calculation as a hover note
                from openpyxl.comments import Comment
                ri = ws.max_row
                for j, p in enumerate(pers, len(lead) + 1):
                    sub = data[k]["_subp"].get(p)
                    if sub:
                        c = Comment(f"{p}:\n{sub}", "engine", height=140, width=420)
                        ws.cell(row=ri, column=j).comment = c
        widths = [10, 48] if sn != "Unmapped" else [16, 12, 52]
        widths += [15] * len(pers) + ([13, 9, 10, 80] if sn != "Unmapped" else [13, 72])
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "C2" if sn != "Unmapped" else "D2"
        ws.auto_filter.ref = ws.dimensions
    out.save(wide_path)


def run(name: str, template, taxonomy) -> str:
    cache = os.path.join(OUT_DIR, ".cache", f"{name}.pkl")
    rows = pickle.load(open(os.path.join(PKL_DIR, f"{name}.pkl"), "rb"))
    unit = filing_unit(os.path.join(PDF_DIR, f"{name}.pdf"))
    mapped = map_quarter(rows, template, taxonomy, default_unit=unit)
    os.makedirs(os.path.dirname(cache), exist_ok=True)
    pickle.dump(mapped, open(cache, "wb"))     # offline re-writes / debugging
    long_dir = os.path.join(OUT_DIR, "long")
    os.makedirs(long_dir, exist_ok=True)
    long_out = os.path.join(long_dir, f"{name.upper()}_dummytest_long.xlsx")
    write_client_workbook_long(name.split("_")[0].upper(), mapped, template, long_out)
    out = os.path.join(OUT_DIR, f"{name.upper()}_dummytest.xlsx")
    to_wide(long_out, out)                     # WIDE is the default deliverable
    return out


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if "--all" in sys.argv:
        args = sorted(fn[:-4] for fn in os.listdir(PKL_DIR)
                      if fn.endswith(".pkl") and not fn.startswith("_"))
    if not args:
        sys.exit(__doc__)
    template = load_template(TEMPLATE)
    taxonomy = load_taxonomy(TAXONOMY)
    for name in args:
        reset_usage()
        out = run(name, template, taxonomy)
        print(f"{name:22s} -> {os.path.basename(out)}  (${usage_cost():.2f})")
